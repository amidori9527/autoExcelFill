from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch
import zipfile

from openpyxl import Workbook, load_workbook

from autoexcel.add_cards import (
    _read_card_numbers_windows,
    add_cards_to_workbook,
    parse_card_numbers,
)
from autoexcel.fast_xlsx import list_sheets


class AddCardsTest(unittest.TestCase):
    def test_parses_multiline_card_numbers(self) -> None:
        self.assertEqual(
            parse_card_numbers("1234\n3121\n\n 1341 \n"),
            ["1234", "3121", "1341"],
        )

    def test_reads_windows_multiline_input_until_blank_line(self) -> None:
        with patch("builtins.print"), patch(
            "builtins.input", side_effect=["1234", " 3121 ", ""]
        ):
            self.assertEqual(_read_card_numbers_windows(), ["1234", "3121"])

    def test_inserts_cards_before_first_non_chinese_colored_template(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            workbook_path = Path(temporary_directory) / "cards.xlsx"
            workbook = Workbook()
            chinese_sheet = workbook.active
            chinese_sheet.title = "试算平衡"
            chinese_sheet.sheet_properties.tabColor = "3157D5"

            template = workbook.create_sheet("01833770033")
            template.sheet_properties.tabColor = "3157D5"
            template.merge_cells("A1:M1")
            template["A1"] = "=SHEETSNAME(A1)"
            template["A2"] = "日期"
            template["B2"] = "期初余额"
            template["C2"] = "代收金额"
            template["M2"] = "差值"
            template["N2"] = "增量"
            template["A4"] = 46210
            template["B4"] = 123
            template["C4"] = "=B4+10"
            template["N4"] = "=C4"
            workbook.save(workbook_path)

            result = add_cards_to_workbook(workbook_path, ["1234", "3121", "1234"])

            self.assertEqual(result.created, ["1234", "3121"])
            self.assertEqual(result.skipped, [("1234", "duplicate input")])
            updated = load_workbook(workbook_path, data_only=False)
            self.assertEqual(
                updated.sheetnames[:4], ["试算平衡", "1234", "3121", "01833770033"]
            )
            for card in ("1234", "3121"):
                sheet = updated[card]
                self.assertEqual(sheet["B3"].value, 0)
                self.assertEqual(sheet["N3"].value, "=M3")
                self.assertEqual(sheet["C3"].value, "=B3+10")
                self.assertEqual(sheet["A1"].value, "=SHEETSNAME(A1)")
                self.assertEqual(sheet.max_row, 3)
                self.assertIn("A1:M1", {str(item) for item in sheet.merged_cells.ranges})

    def test_avoids_orphaned_sheet_relationship_paths(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            workbook_path = Path(temporary_directory) / "cards.xlsx"
            workbook = Workbook()
            workbook.active.title = "试算平衡"
            template = workbook.create_sheet("01833770033")
            template.sheet_properties.tabColor = "3157D5"
            template["A1"] = "模板"
            template["A2"] = "日期"
            template["B2"] = "期初余额"
            template["M2"] = "差值"
            template["N2"] = "增量"
            template["A3"] = 46210
            workbook.save(workbook_path)
            with zipfile.ZipFile(workbook_path, "a") as archive:
                archive.writestr(
                    "xl/worksheets/_rels/sheet3.xml.rels",
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"/>',
                )

            add_cards_to_workbook(workbook_path, ["1234"])

            entries = {entry.name: entry.path for entry in list_sheets(workbook_path)}
            self.assertEqual(entries["1234"], "xl/worksheets/sheet4.xml")


if __name__ == "__main__":
    unittest.main()
