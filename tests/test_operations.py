from __future__ import annotations

from datetime import date
import unittest

from openpyxl import Workbook

from autoexcel.operations import freeze_colored_sheets_next_day, freeze_next_day_row


class ColoredSheetOperationsTest(unittest.TestCase):
    def test_new_current_date_row_resets_j_and_clears_o(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Merchant A"
        sheet.sheet_properties.tabColor = "3157D5"
        sheet["A1"] = date(2026, 7, 13)
        sheet["J1"] = 88.25
        sheet["J1"].number_format = "0.00"
        sheet["O1"] = 42.5
        sheet["O1"].number_format = "0.00"

        values_workbook = Workbook()
        values_sheet = values_workbook.active
        values_sheet.title = "Merchant A"
        values_sheet["A1"] = date(2026, 7, 13)
        values_sheet["J1"] = 88.25
        values_sheet["O1"] = 42.5

        changed, skipped = freeze_colored_sheets_next_day(
            workbook,
            values_workbook,
            date(2026, 7, 14),
        )

        self.assertEqual(changed, [("Merchant A", 1)])
        self.assertEqual(skipped, [])
        self.assertEqual(sheet["J1"].value, 88.25)
        self.assertEqual(sheet["O1"].value, 42.5)
        self.assertEqual(sheet["J2"].value, 0)
        self.assertEqual(sheet["J2"].number_format, "0.00")
        self.assertIsNone(sheet["O2"].value)
        self.assertEqual(sheet["O2"].number_format, "0.00")

    def test_single_sheet_insert_keeps_j_and_o_unchanged(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Merchant A"
        sheet["A1"] = date(2026, 7, 13)
        sheet["J1"] = 88.25
        sheet["O1"] = 42.5

        values_workbook = Workbook()
        values_sheet = values_workbook.active
        values_sheet.title = "Merchant A"
        values_sheet["A1"] = date(2026, 7, 13)
        values_sheet["J1"] = 88.25
        values_sheet["O1"] = 42.5

        freeze_next_day_row(
            workbook,
            values_workbook,
            "Merchant A",
            date(2026, 7, 14),
        )

        self.assertEqual(sheet["J2"].value, 88.25)
        self.assertEqual(sheet["O2"].value, 42.5)


if __name__ == "__main__":
    unittest.main()
