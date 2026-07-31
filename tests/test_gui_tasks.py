from __future__ import annotations

from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from autoexcel.add_cards import AddCardsResult
from autoexcel.daily_long_sync import DailyLongSyncResult
from autoexcel.flow_sync import (
    FlowSyncFiles,
    FullFlowSyncResult,
    TpCollectionSyncResult,
    TpPayoutSyncResult,
    TpWithdrawalSyncResult,
)
from autoexcel.gui_tasks import (
    _jobs_from_directory,
    run_add_cards_task,
    run_daily_long_sync_task,
    run_fill_task,
    run_full_flow_sync_task,
    run_tp_collection_sync_task,
    run_tp_payout_sync_task,
    run_tp_withdrawal_sync_task,
    run_wallet_flow_sync_task,
)


class GuiTasksTest(unittest.TestCase):
    def test_add_cards_task_passes_selected_color_mode(self) -> None:
        workbook = Path("/tmp/cards.xlsx")
        with (
            patch(
                "autoexcel.gui_tasks.add_cards.parse_card_numbers",
                return_value=["1234", "3121"],
            ),
            patch(
                "autoexcel.gui_tasks.add_cards.add_cards_to_workbook",
                return_value=AddCardsResult(
                    ["1234", "3121"],
                    [],
                    "01833770033",
                ),
            ) as add_cards,
        ):
            result = run_add_cards_task(
                workbook,
                "1234\n3121",
                lambda _message: None,
                sheet_color=None,
                random_sheet_colors=True,
            )

        add_cards.assert_called_once_with(
            workbook,
            ["1234", "3121"],
            sheet_color=None,
            random_sheet_colors=True,
        )
        self.assertEqual(result.title, "增卡完成")

    def test_daily_long_sync_task_reports_inserted_and_skipped_rows(self) -> None:
        result = DailyLongSyncResult(
            source_rows=12,
            inserted_rows=10,
            skipped_rows=2,
            old_summary_row=100,
            new_summary_row=110,
            repaired_formula_rows=99,
        )

        with patch(
            "autoexcel.gui_tasks.daily_long_sync.sync_daily_long",
            return_value=result,
        ) as sync:
            task_result = run_daily_long_sync_task(
                Path("/tmp/workbook.xlsx"),
                lambda _message: None,
            )

        sync.assert_called_once()
        self.assertEqual(task_result.title, "当日长款数据同步完成")
        self.assertIn("识别当日长款 12 条", task_result.summary)
        self.assertIn("新增 10 条", task_result.summary)
        self.assertIn("重复跳过 2 条", task_result.summary)
        self.assertIn("修复“是否补单”公式 99 行", task_result.summary)
        self.assertEqual(task_result.output_path, Path("/tmp/workbook.xlsx"))

    def test_full_flow_sync_task_reports_all_row_counts(self) -> None:
        directory = Path("/tmp/flow-sync")
        workbook = directory / "对账结果.xlsx"
        payout = TpPayoutSyncResult(10, 14, 11, 15, 18, 4)
        collection = TpCollectionSyncResult(20, 12, 21, 13, 30, 0, 10, 2)
        withdrawal = TpWithdrawalSyncResult(16, 3, 17, 4, 19, 0, 2)
        wallet = TpPayoutSyncResult(30, 18, 31, 19, 40, 0)
        result = FullFlowSyncResult(
            files=FlowSyncFiles(
                directory=directory,
                workbook=workbook,
                payment_orders=directory / "付款订单.xlsx",
                collection_orders=(
                    directory / "收款订单1.xlsx",
                    directory / "收款订单2.xlsx",
                ),
                withdrawal_orders=directory / "商户提现申请.xlsx",
                wallet_flow=directory / "平台钱包流水.xlsx",
            ),
            payout=payout,
            collection=collection,
            withdrawal=withdrawal,
            wallet=wallet,
        )

        with patch(
            "autoexcel.gui_tasks.flow_sync.sync_all_flows",
            return_value=result,
        ) as sync:
            task_result = run_full_flow_sync_task(
                directory,
                lambda _message: None,
            )

        sync.assert_called_once()
        self.assertEqual(task_result.title, "一键流水同步完成")
        self.assertIn("TP代付 14 条", task_result.summary)
        self.assertIn("TP代收 12 条", task_result.summary)
        self.assertIn("TP提现 3 条", task_result.summary)
        self.assertIn("钱包流水 18 条", task_result.summary)
        self.assertEqual(task_result.output_path, workbook)

    def test_full_flow_sync_task_reports_missing_withdrawal_as_skipped(self) -> None:
        directory = Path("/tmp/flow-sync")
        workbook = directory / "对账结果.xlsx"
        payout = TpPayoutSyncResult(10, 14, 11, 15, 18, 4)
        collection = TpCollectionSyncResult(20, 12, 21, 13, 30, 0, 10, 2)
        wallet = TpPayoutSyncResult(30, 18, 31, 19, 40, 0)
        result = FullFlowSyncResult(
            files=FlowSyncFiles(
                directory=directory,
                workbook=workbook,
                payment_orders=directory / "付款订单.xlsx",
                collection_orders=(
                    directory / "收款订单1.xlsx",
                    directory / "收款订单2.xlsx",
                ),
                withdrawal_orders=None,
                wallet_flow=directory / "平台钱包流水.xlsx",
            ),
            payout=payout,
            collection=collection,
            withdrawal=None,
            wallet=wallet,
        )

        with patch(
            "autoexcel.gui_tasks.flow_sync.sync_all_flows",
            return_value=result,
        ):
            task_result = run_full_flow_sync_task(
                directory,
                lambda _message: None,
            )

        self.assertIn("TP提现 跳过（当日无提现文件）", task_result.summary)
        self.assertIn("钱包流水 18 条", task_result.summary)

    def test_wallet_flow_sync_task_reports_row_counts(self) -> None:
        result = TpPayoutSyncResult(
            removed_rows=20,
            inserted_rows=18,
            old_detail_end_row=21,
            new_detail_end_row=19,
            summary_row=30,
            shifted_rows=0,
        )

        with patch(
            "autoexcel.gui_tasks.flow_sync.sync_wallet_flow",
            return_value=result,
        ) as sync:
            task_result = run_wallet_flow_sync_task(
                Path("/tmp/workbook.xlsx"),
                Path("/tmp/wallet-flow.xlsx"),
                lambda _message: None,
            )

        sync.assert_called_once()
        self.assertEqual(task_result.title, "钱包流水同步完成")
        self.assertIn("清除 20 条", task_result.summary)
        self.assertIn("写入 18 条", task_result.summary)

    def test_tp_collection_sync_task_reports_status_counts(self) -> None:
        result = TpCollectionSyncResult(
            removed_rows=20,
            inserted_rows=12,
            old_detail_end_row=21,
            new_detail_end_row=13,
            summary_row=30,
            shifted_rows=0,
            successful_rows=10,
            partial_rows=2,
        )

        with patch(
            "autoexcel.gui_tasks.flow_sync.sync_tp_collection",
            return_value=result,
        ) as sync:
            task_result = run_tp_collection_sync_task(
                Path("/tmp/workbook.xlsx"),
                Path("/tmp/collection-1.xlsx"),
                Path("/tmp/collection-2.xlsx"),
                lambda _message: None,
            )

        sync.assert_called_once()
        self.assertEqual(task_result.title, "TP代收同步完成")
        self.assertIn("写入 12 条", task_result.summary)
        self.assertIn("支付成功 10 条", task_result.summary)
        self.assertIn("部分支付 2 条", task_result.summary)

    def test_tp_withdrawal_sync_task_reports_status_counts(self) -> None:
        result = TpWithdrawalSyncResult(
            removed_rows=16,
            inserted_rows=3,
            old_detail_end_row=17,
            new_detail_end_row=4,
            summary_row=19,
            shifted_rows=0,
            skipped_rows=2,
        )

        with patch(
            "autoexcel.gui_tasks.flow_sync.sync_tp_withdrawal",
            return_value=result,
        ) as sync:
            task_result = run_tp_withdrawal_sync_task(
                Path("/tmp/workbook.xlsx"),
                Path("/tmp/withdrawal.xlsx"),
                lambda _message: None,
            )

        sync.assert_called_once()
        self.assertEqual(task_result.title, "TP提现同步完成")
        self.assertIn("写入 3 条", task_result.summary)
        self.assertIn("跳过 2 条", task_result.summary)

    def test_tp_payout_sync_task_reports_dynamic_ranges(self) -> None:
        result = TpPayoutSyncResult(
            removed_rows=10,
            inserted_rows=14,
            old_detail_end_row=11,
            new_detail_end_row=15,
            summary_row=18,
            shifted_rows=4,
        )

        with patch(
            "autoexcel.gui_tasks.flow_sync.sync_tp_payout",
            return_value=result,
        ) as sync:
            task_result = run_tp_payout_sync_task(
                Path("/tmp/workbook.xlsx"),
                Path("/tmp/payment-orders.xlsx"),
                lambda _message: None,
            )

        sync.assert_called_once()
        self.assertEqual(task_result.title, "TP代付同步完成")
        self.assertIn("清除 10 条", task_result.summary)
        self.assertIn("写入 14 条", task_result.summary)
        self.assertIn("下移 4 行", task_result.summary)

    def test_jobs_from_directory_matches_one_pair_for_date(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            (directory / "TranDetailReport_123_202607140000.xlsx").touch()
            (directory / "收款订单_202607140001.xlsx").touch()

            jobs = _jobs_from_directory(directory, date(2026, 7, 14))

            self.assertEqual(len(jobs), 1)
            self.assertTrue(jobs[0].upstream_path.name.startswith("TranDetailReport"))
            self.assertTrue(jobs[0].backend_path.name.startswith("收款订单"))

    def test_jobs_from_directory_rejects_empty_directory(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            with self.assertRaisesRegex(FileNotFoundError, "没有找到 Excel"):
                _jobs_from_directory(Path(temporary_directory), date(2026, 7, 14))

    def test_fill_task_rejects_non_xlsx_file_before_processing(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            workbook = Path(temporary_directory) / "orders.xls"
            workbook.touch()

            with self.assertRaisesRegex(ValueError, "xlsx"):
                run_fill_task(workbook, date(2026, 7, 14), lambda _message: None)

    def test_fill_task_updates_colored_sheet(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            workbook_path = directory / "ledger.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Merchant A"
            sheet.sheet_properties.tabColor = "3157D5"
            sheet["A1"] = date(2026, 7, 13)
            sheet["J1"] = 88.25
            sheet["J1"].number_format = "0.00"
            sheet["K1"] = 100
            sheet["M1"] = 120
            sheet["O1"] = 42.5
            sheet["O1"].number_format = "0.00"
            workbook.save(workbook_path)

            with patch(
                "autoexcel.gui_tasks.create_process_log_path",
                return_value=directory / "fill.log",
            ):
                result = run_fill_task(
                    workbook_path, date(2026, 7, 14), lambda _message: None
                )

            updated = load_workbook(workbook_path, data_only=False)
            self.assertEqual(result.title, "Excel 增行完成")
            self.assertEqual(updated["Merchant A"]["A2"].value.date(), date(2026, 7, 14))
            self.assertEqual(updated["Merchant A"]["J1"].value, 88.25)
            self.assertEqual(updated["Merchant A"]["O1"].value, 42.5)
            self.assertEqual(updated["Merchant A"]["J2"].value, 0)
            self.assertEqual(updated["Merchant A"]["J2"].number_format, "0.00")
            self.assertIsNone(updated["Merchant A"]["O2"].value)
            self.assertEqual(updated["Merchant A"]["O2"].number_format, "0.00")

    def test_fill_task_updates_b2b_and_income_after_colored_sheets(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            workbook_path = directory / "ledger.xlsx"
            workbook = Workbook()
            colored = workbook.active
            colored.title = "Merchant A"
            colored.sheet_properties.tabColor = "3157D5"
            colored["A1"] = date(2026, 7, 13)

            b2b = workbook.create_sheet("B2B支出")
            b2b.append([None] * 15)
            b2b.append(["日期", *[f"字段{index}" for index in range(2, 16)]])
            b2b.append([date(2026, 7, 13), *([1] * 14)])
            b2b.append(["汇总", *([None] * 14)])
            b2b_table = Table(displayName="B2BTable", ref="A2:O4")
            b2b_table.totalsRowCount = 1
            b2b.add_table(b2b_table)

            income = workbook.create_sheet("收入")
            income.append(["日期", *[f"字段{index}" for index in range(2, 14)]])
            income.append([date(2026, 7, 13), *([1] * 12)])
            income.append(["汇总", *([None] * 12)])
            table = Table(displayName="IncomeTable", ref="A1:M3")
            table.totalsRowCount = 1
            income.add_table(table)
            workbook.save(workbook_path)

            with patch(
                "autoexcel.gui_tasks.create_process_log_path",
                return_value=directory / "fill.log",
            ):
                run_fill_task(workbook_path, date(2026, 7, 14), lambda _message: None)

            updated = load_workbook(workbook_path, data_only=False)
            updated_b2b = updated["B2B支出"]
            self.assertEqual(updated_b2b["A3"].value.date(), date(2026, 7, 13))
            self.assertEqual(updated_b2b["A4"].value.date(), date(2026, 7, 14))
            self.assertEqual(updated_b2b["A5"].value, "汇总")
            self.assertEqual(updated_b2b.tables["B2BTable"].ref, "A2:O5")
            updated_income = updated["收入"]
            self.assertEqual(updated_income["A2"].value.date(), date(2026, 7, 13))
            self.assertEqual(updated_income["A3"].value.date(), date(2026, 7, 14))
            self.assertEqual(updated_income["A4"].value, "汇总")
            self.assertEqual(updated_income.tables["IncomeTable"].ref, "A1:M4")


if __name__ == "__main__":
    unittest.main()
