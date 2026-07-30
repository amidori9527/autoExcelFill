from __future__ import annotations

from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from autoexcel.fast_xlsx import (
    MAIN_NS,
    _advance_income_sheet_xml,
    advance_income_sheet_fast,
    advance_summary_table_sheet_fast,
)


def tag(name: str) -> str:
    return f"{{{MAIN_NS}}}{name}"


class IncomeSheetFastTest(unittest.TestCase):
    def test_income_row_is_frozen_and_formula_row_advances(self) -> None:
        xml = b"""<?xml version="1.0" encoding="utf-8"?>
        <worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
          <dimension ref="A1:M3"/>
          <sheetData>
            <row r="1"><c r="A1" t="inlineStr"><is><t>Date</t></is></c></row>
            <row r="2">
              <c r="A2" s="1"><v>46210</v></c>
              <c r="B2" s="2"><f>Source!B3</f><v>100</v></c>
              <c r="C2" s="2"><f>B2*0.004</f><v>0.4</v></c>
              <c r="L2" s="2"><f>SUM(B2:K2)</f><v>100.4</v></c>
            </row>
            <row r="3"><c r="A3" t="inlineStr"><is><t>Total</t></is></c><c r="L3"><f>SUBTOTAL(109,L2:L2)</f><v>100.4</v></c></row>
          </sheetData>
        </worksheet>"""

        updated_xml, inserted_row, reason = _advance_income_sheet_xml(
            xml, date(2026, 7, 8)
        )

        self.assertEqual(inserted_row, 2)
        self.assertIsNone(reason)
        root = ET.fromstring(updated_xml)
        rows = root.findall(f"{tag('sheetData')}/{tag('row')}")
        self.assertEqual([row.attrib["r"] for row in rows], ["1", "2", "3", "4"])
        self.assertEqual(root.find(tag("dimension")).attrib["ref"], "A1:M4")

        frozen_cells = {cell.attrib["r"]: cell for cell in rows[1].findall(tag("c"))}
        self.assertEqual(frozen_cells["B2"].find(tag("v")).text, "100")
        self.assertIsNone(frozen_cells["B2"].find(tag("f")))
        self.assertEqual(frozen_cells["C2"].find(tag("v")).text, "0.4")
        self.assertIsNone(frozen_cells["C2"].find(tag("f")))

        formula_cells = {cell.attrib["r"]: cell for cell in rows[2].findall(tag("c"))}
        self.assertEqual(formula_cells["A3"].find(tag("v")).text, "46211")
        self.assertEqual(formula_cells["B3"].find(tag("f")).text, "Source!B4")
        self.assertEqual(formula_cells["C3"].find(tag("f")).text, "B3*0.004")
        self.assertEqual(formula_cells["L3"].find(tag("f")).text, "SUM(B3:K3)")

    def test_income_table_range_expands_with_inserted_row(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            workbook_path = Path(temporary_directory) / "income.xlsx"
            workbook = Workbook()
            income = workbook.active
            income.title = "收入"
            income.append(["日期", *[f"字段{index}" for index in range(2, 14)]])
            income.append([date(2026, 7, 7), *([1] * 12)])
            income.append(["汇总", *([None] * 12)])
            table = Table(displayName="IncomeTable", ref="A1:M3")
            table.totalsRowCount = 1
            income.add_table(table)
            workbook.save(workbook_path)

            result = advance_income_sheet_fast(workbook_path, date(2026, 7, 8))

            self.assertTrue(result.changed)
            self.assertEqual(result.inserted_row, 2)
            updated = load_workbook(workbook_path, data_only=False)
            updated_income = updated["收入"]
            self.assertEqual(updated_income["A2"].value.date(), date(2026, 7, 7))
            self.assertEqual(updated_income["A3"].value.date(), date(2026, 7, 8))
            self.assertEqual(updated_income["A4"].value, "汇总")
            self.assertEqual(updated_income.tables["IncomeTable"].ref, "A1:M4")

    def test_b2b_uses_table_total_row_instead_of_physical_last_row(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            workbook_path = Path(temporary_directory) / "b2b.xlsx"
            workbook = Workbook()
            b2b = workbook.active
            b2b.title = "B2B支出"
            b2b.append([None] * 15)
            b2b.append(["日期", *[f"字段{index}" for index in range(2, 16)]])
            b2b.append([date(2026, 7, 7), *([1] * 14)])
            b2b.append(["汇总", *([None] * 14)])
            b2b["E6"] = None
            b2b["E6"].number_format = "0.00"
            table = Table(displayName="B2BTable", ref="A2:O4")
            table.totalsRowCount = 1
            b2b.add_table(table)
            workbook.save(workbook_path)

            result = advance_summary_table_sheet_fast(
                workbook_path, "B2B支出", date(2026, 7, 8)
            )

            self.assertTrue(result.changed)
            self.assertEqual(result.inserted_row, 3)
            updated = load_workbook(workbook_path, data_only=False)
            updated_b2b = updated["B2B支出"]
            self.assertEqual(updated_b2b["A3"].value.date(), date(2026, 7, 7))
            self.assertEqual(updated_b2b["A4"].value.date(), date(2026, 7, 8))
            self.assertEqual(updated_b2b["A5"].value, "汇总")
            self.assertEqual(updated_b2b.tables["B2BTable"].ref, "A2:O5")
            self.assertTrue(updated_b2b["E7"].has_style)

    def test_daily_balance_opening_balance_references_previous_row_column_r(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            workbook_path = Path(temporary_directory) / "daily-balance.xlsx"
            workbook = Workbook()
            daily_balance = workbook.active
            daily_balance.title = "每日余额监测"
            daily_balance.append(["日期", *[f"字段{index}" for index in range(2, 17)]])
            daily_balance.append([date(2026, 7, 7), *([1] * 15)])
            daily_balance.append(["汇总", *([None] * 15)])
            table = Table(displayName="DailyBalanceTable", ref="A1:P3")
            table.totalsRowCount = 1
            daily_balance.add_table(table)
            workbook.save(workbook_path)

            result = advance_summary_table_sheet_fast(
                workbook_path,
                "每日余额监测",
                date(2026, 7, 8),
                opening_balance_increment=True,
            )

            self.assertTrue(result.changed)
            updated = load_workbook(workbook_path, data_only=False)["每日余额监测"]
            self.assertEqual(updated["B3"].value, "=R2")


if __name__ == "__main__":
    unittest.main()
