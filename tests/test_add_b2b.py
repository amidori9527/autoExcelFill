from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
import zipfile

from openpyxl import Workbook, load_workbook

from autoexcel.add_b2b import (
    FieldMapping,
    append_b2b_to_workbook,
    guess_field_mapping,
    parse_input_text,
    validate_and_build_records,
)


STANDARD_TEXT = (
    "2026-07-12 22:04:23  75NVDJEI  01635548053  01850801086  50000\n"
    "2026-07-12 21:49:26  75NV9QFJ  01642530541  01604767483  60000"
)
STANDARD_MAPPING = FieldMapping(
    date_time=0,
    trx_id=1,
    outgoing_card=2,
    amount=4,
)


class AddB2BTest(unittest.TestCase):
    def test_groups_date_and_time_and_guesses_mapping(self) -> None:
        lines = parse_input_text(STANDARD_TEXT)

        self.assertEqual(
            [field.value for field in lines[0].fields],
            [
                "2026-07-12 22:04:23",
                "75NVDJEI",
                "01635548053",
                "01850801086",
                "50000",
            ],
        )
        self.assertEqual(guess_field_mapping(lines[0].fields), STANDARD_MAPPING)

    def test_supports_a_different_field_order_selected_from_first_line(self) -> None:
        lines = parse_input_text(
            "75NVDJEI 50000 01850801086 2026-07-12 22:04:23 01635548053"
        )
        mapping = FieldMapping(
            date_time=3,
            trx_id=0,
            outgoing_card=4,
            amount=1,
        )

        record = validate_and_build_records(lines, mapping)[0]

        self.assertEqual(record.date_time, "2026-07-12 22:04:23")
        self.assertEqual(record.trx_id, "75NVDJEI")
        self.assertEqual(record.outgoing_card, "01635548053")
        self.assertEqual(str(record.amount), "50000")

    def test_converts_negative_amount_to_positive(self) -> None:
        lines = parse_input_text(
            "2026-07-12 22:04:23 75NVDJEI 01635548053 01850801086 -50000"
        )

        record = validate_and_build_records(lines, STANDARD_MAPPING)[0]

        self.assertEqual(record.amount, 50000)
        self.assertTrue(record.amount_was_negative)

    def test_rejects_zero_amount(self) -> None:
        lines = parse_input_text(
            "2026-07-12 22:04:23 75NVDJEI 01635548053 01850801086 0"
        )

        with self.assertRaisesRegex(ValueError, "金额必须大于 0"):
            validate_and_build_records(lines, STANDARD_MAPPING)

    def test_rejects_non_numeric_amount(self) -> None:
        lines = parse_input_text(
            "2026-07-12 22:04:23 75NVDJEI 01635548053 01850801086 错误金额"
        )

        with self.assertRaisesRegex(ValueError, "不是数字"):
            validate_and_build_records(lines, STANDARD_MAPPING)

    def test_zero_amount_does_not_modify_workbook(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            workbook_path = Path(temporary_directory) / "unchanged.xlsx"
            workbook = Workbook()
            workbook.active.title = "提取B2B"
            workbook.active.append(
                ["原始数据", "日期", "提取日期", "转出卡号", "金额"]
            )
            workbook.save(workbook_path)
            original = workbook_path.read_bytes()
            lines = parse_input_text(
                "2026-07-12 22:04:23 75NVDJEI 01635548053 01850801086 0"
            )

            with self.assertRaisesRegex(ValueError, "金额必须大于 0"):
                append_b2b_to_workbook(workbook_path, lines, STANDARD_MAPPING)

            self.assertEqual(workbook_path.read_bytes(), original)

    def test_rejects_rows_that_do_not_match_the_first_line(self) -> None:
        lines = parse_input_text(
            STANDARD_TEXT
            + "\n2026-07-12 21:35:30 75NV54HX 01768458715 50000"
        )

        with self.assertRaisesRegex(ValueError, "第 3 行：字段数"):
            validate_and_build_records(lines, STANDARD_MAPPING)

    def test_appends_formulas_cached_values_styles_and_filter_range(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            workbook_path = Path(temporary_directory) / "b2b.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "提取B2B"
            headers = [
                "ctrl+shift+v粘贴",
                "日期",
                "提取日期",
                "转出卡号",
                "提取对应金额",
                "B2B佣金扣除",
                "提取收款卡号",
                "测试账号",
                "提取TRXID",
            ]
            sheet.append(headers)
            sheet.append(
                [
                    "2026-07-11 10:00:00 OLD123 01600000000 01800000000 10000",
                    "2026-07-11 10:00:00",
                    "20260711",
                    "01600000000",
                    10000,
                    40,
                    "01800000000",
                    None,
                    "OLD123",
                ]
            )
            sheet["C2"].number_format = "0"
            sheet["D2"].number_format = "@"
            sheet["E2"].number_format = "#,##0"
            sheet["G2"].number_format = "@"
            sheet["I2"].number_format = "@"
            sheet.auto_filter.ref = "A1:I2"
            workbook.create_sheet("其他")
            workbook.save(workbook_path)

            lines = parse_input_text(STANDARD_TEXT)
            result = append_b2b_to_workbook(workbook_path, lines, STANDARD_MAPPING)

            self.assertEqual(result.inserted_count, 2)
            self.assertEqual((result.start_row, result.end_row), (3, 4))
            formulas = load_workbook(workbook_path, data_only=False)
            updated = formulas["提取B2B"]
            self.assertEqual(updated["A3"].value, lines[0].raw)
            self.assertEqual(
                updated["B3"].value,
                '=_xlfn.TEXTBEFORE(TRIM(A3)," ",2)',
            )
            self.assertEqual(updated["C3"].value, '=SUBSTITUTE(LEFT(B3,10),"-","")')
            self.assertEqual(
                updated["D3"].value,
                '=_xlfn.TEXTBEFORE(_xlfn.TEXTAFTER(TRIM(A3)," ",3)," ")',
            )
            self.assertEqual(updated["E3"].value, 50000)
            self.assertEqual(updated["E3"].data_type, "n")
            self.assertEqual(updated["F3"].value, "=E3*0.004")
            self.assertEqual(updated.auto_filter.ref, "A1:I4")
            self.assertEqual(updated["D3"].style_id, updated["D2"].style_id)
            self.assertIsNone(updated["G3"].value)

            values = load_workbook(workbook_path, data_only=True)["提取B2B"]
            self.assertEqual(values["B3"].value, "2026-07-12 22:04:23")
            self.assertEqual(values["C3"].value, "20260712")
            self.assertEqual(values["D3"].value, "01635548053")
            self.assertEqual(values["E3"].value, 50000)
            self.assertEqual(values["F3"].value, 200)
            self.assertIsNone(values["G3"].value)
            self.assertEqual(values["I3"].value, "75NVDJEI")

            with zipfile.ZipFile(workbook_path) as archive:
                self.assertIsNone(archive.testzip())

    def test_negative_amount_is_written_positive_and_reported(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            workbook_path = Path(temporary_directory) / "negative.xlsx"
            workbook = Workbook()
            workbook.active.title = "提取B2B"
            workbook.active.append(
                ["原始数据", "日期", "提取日期", "转出卡号", "金额", "佣金"]
            )
            workbook.save(workbook_path)
            lines = parse_input_text(
                "2026-07-12 22:04:23 75NVDJEI 01635548053 01850801086 -50000"
            )

            result = append_b2b_to_workbook(workbook_path, lines, STANDARD_MAPPING)

            self.assertEqual(result.converted_negative_count, 1)
            formulas = load_workbook(workbook_path, data_only=False)["提取B2B"]
            self.assertEqual(formulas["E2"].value, 50000)
            self.assertEqual(formulas["E2"].data_type, "n")
            values = load_workbook(workbook_path, data_only=True)["提取B2B"]
            self.assertEqual(values["E2"].value, 50000)
            self.assertEqual(values["F2"].value, 200)
            self.assertIsNone(values["G2"].value)


if __name__ == "__main__":
    unittest.main()
