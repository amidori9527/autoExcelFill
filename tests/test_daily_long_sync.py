from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table

from autoexcel.daily_long_sync import CUMULATIVE_HEADERS, sync_daily_long
from autoexcel.flow_sync import WALLET_FLOW_HEADERS


def make_daily_row(transaction_id: str, amount: float) -> list[object]:
    return [
        "",
        "",
        "正常",
        "2026-07-24 12:00:00",
        "2026-07-24 11:59:58",
        transaction_id,
        "wallet-1",
        "user-wallet-1",
        "代理户",
        "Nagad",
        "长款",
        100.5,
        amount,
        0,
        100.5 + amount,
    ]


def make_workbook(path: Path) -> None:
    workbook = Workbook()
    daily = workbook.active
    daily.title = "长款(当日)"
    daily.append(list(WALLET_FLOW_HEADERS))
    daily.append(make_daily_row("existing-1", 10))
    daily.append(make_daily_row("new-1", 20))
    daily.append(make_daily_row("new-1", 20))
    daily.append([])
    daily.append(
        [
            None,
            "钱包ID",
            "求和项:交易金额",
            "计数项:交易类型",
        ]
    )
    daily.append([None, "wallet-1", 50, 3])
    daily.append([None, "总计", 50, 3])

    cumulative = workbook.create_sheet("长款累计")
    cumulative.append(list(CUMULATIVE_HEADERS))
    cumulative.append(
        [
            "正常",
            "2026-07-23 12:00:00",
            "2026-07-23 11:59:58",
            "existing-1",
            "wallet-old",
            "user-old",
            "代理户",
            None,
            "Nagad",
            "长款",
            10,
            20,
            0,
            30,
            '=IF(COUNTIF(B2B回款!B:B,D2),"YES","NO")',
            '=IF(COUNTIF(TP代收!L:L,#REF!),"YES","NO")',
            '=TEXT(INT(B2),"YYYYMMDD")',
            None,
        ]
    )
    cumulative.append(
        [
            "汇总",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "=SUBTOTAL(109,表_5[交易金额])",
        ]
    )
    table = Table(displayName="表_5", ref="A1:R3")
    table.totalsRowCount = 1
    cumulative.add_table(table)
    cumulative["A4"].fill = PatternFill("solid", fgColor="FFFF00")
    workbook.save(path)


class DailyLongSyncTest(unittest.TestCase):
    def test_sync_inserts_unique_rows_before_total_and_repairs_formulas(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            workbook_path = Path(temporary_directory) / "workbook.xlsx"
            make_workbook(workbook_path)

            result = sync_daily_long(workbook_path)

            self.assertEqual(result.source_rows, 3)
            self.assertEqual(result.inserted_rows, 1)
            self.assertEqual(result.skipped_rows, 2)
            self.assertEqual(result.old_summary_row, 3)
            self.assertEqual(result.new_summary_row, 4)
            self.assertEqual(result.repaired_formula_rows, 1)

            workbook = load_workbook(workbook_path, data_only=False)
            cumulative = workbook["长款累计"]
            self.assertEqual(cumulative["D3"].value, "new-1")
            self.assertEqual(cumulative["A3"].value, "正常")
            self.assertIsNone(cumulative["H3"].value)
            self.assertEqual(cumulative["K3"].value, 100.5)
            self.assertEqual(cumulative["L3"].value, 20)
            self.assertEqual(cumulative["N3"].value, 120.5)
            self.assertEqual(
                cumulative["O3"].value,
                '=IF(COUNTIF(B2B回款!B:B,D3),"YES","NO")',
            )
            self.assertEqual(
                cumulative["P2"].value,
                '=IF(COUNTIF(TP代收!L:L,D2),"YES","NO")',
            )
            self.assertEqual(
                cumulative["P3"].value,
                '=IF(COUNTIF(TP代收!L:L,D3),"YES","NO")',
            )
            self.assertEqual(
                cumulative["Q3"].value,
                '=TEXT(INT(B3),"YYYYMMDD")',
            )
            self.assertEqual(cumulative["A4"].value, "汇总")
            self.assertTrue(cumulative["A5"].has_style)
            self.assertEqual(cumulative.tables["表_5"].ref, "A1:R4")
            self.assertEqual(cumulative.tables["表_5"].autoFilter.ref, "A1:R3")

            second_result = sync_daily_long(workbook_path)

            self.assertEqual(second_result.inserted_rows, 0)
            self.assertEqual(second_result.skipped_rows, 3)
            self.assertEqual(second_result.new_summary_row, 4)
