from __future__ import annotations

import csv
from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from openpyxl import Workbook

from autoexcel import diff_orders, payout_diff


class PayoutDiffTest(unittest.TestCase):
    def write_upstream(self, path: Path) -> None:
        with path.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.writer(stream)
            writer.writerow(["TRANS_ID", "TRX_STATUS", "TRX_AMT", "FEE", "FED"])
            writer.writerow(["100", "COMPLETED", "100", "0.30", "0.04"])
            writer.writerow(["200", "FAILED", "75", "0.20", "0.03"])
            writer.writerow(["300", "COMPLETED", "50", "0.15", "0.02"])

    def write_backend(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "付款订单"
        sheet.append(
            [
                "transactionId",
                "交易状态",
                "付款金额(PKR)",
                "手续费(PKR)",
                "支付方式名称",
            ]
        )
        sheet.append(["100", "上游已打款", 100, 1.50, "jazzcash"])
        sheet.append(["300", "处理中", 50, 0.75, "jazzcash"])
        sheet.append(["400", "上游已打款", 25, 0.38, "easypaisa"])
        workbook.save(path)

    def write_finerbit_collection(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Transaction"
        sheet.append(
            [
                "TransactionId",
                "ReferenceId",
                "ServiceName",
                "ChannelName",
                "Route",
                "CustomerName",
                "PurchaseAmount",
                "ConsumerNumber",
                "Email",
                "CNIC",
                "TransactionStatus",
                "Created Date",
            ]
        )
        sheet.append([1, "C1", "E-Wallet", "EasyPaisa", "", "TarsPay", 100, "", "", "", "Success", "19-07-2026"])
        sheet.append([2, "C2", "E-Wallet", "JazzCash", "", "TarsPay", 200, "", "", "", "Success", "19-07-2026"])
        sheet.append([3, "C3", "E-Wallet", "EasyPaisa", "", "TarsPay", 999, "", "", "", "Failed", "19-07-2026"])
        workbook.save(path)

    def write_finerbit_disbursement(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Transaction"
        headers = [f"Column {index}" for index in range(1, 29)]
        headers[1] = "Created Date Time"
        headers[4] = "Reference Id"
        headers[16] = "Received Amount"
        headers[26] = "Status"
        sheet.append(headers)
        for order_id, created_at, amount, status in (
            ("P1", "19-07-2026 10:00:00", 100, "Success"),
            ("P2", "19-07-2026 11:00:00", 200, "Success"),
            ("P3", "19-07-2026 12:00:00", 999, "Failed"),
            ("P5", "20-07-2026 10:00:00", 400, "Success"),
        ):
            row = [None] * 28
            row[1] = created_at
            row[4] = order_id
            row[16] = amount
            row[26] = status
            sheet.append(row)
        workbook.save(path)

    def write_finerbit_backend(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "付款订单"
        headers = [f"列{index}" for index in range(1, 26)]
        headers[0] = "平台订单号"
        headers[6] = "付款金额(PKR)"
        headers[7] = "手续费(PKR)"
        headers[9] = "渠道成本(PKR)"
        headers[21] = "交易状态"
        sheet.append(headers)
        for order_id, amount, fee, cost, status in (
            ("P1", 100, 5, 2, "上游已打款"),
            ("P4", 50, 2, 1, "上游已打款"),
            ("P3", 999, 9, 9, "处理中"),
        ):
            row = [None] * 25
            row[0] = order_id
            row[6] = amount
            row[7] = fee
            row[9] = cost
            row[21] = status
            sheet.append(row)
        workbook.save(path)

    def test_reads_successful_orders_and_calculates_payout_fees(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            upstream = directory / "上游.csv"
            backend = directory / "付款订单.xlsx"
            self.write_upstream(upstream)
            self.write_backend(backend)

            result = payout_diff.read_job_diff_result(
                payout_diff.make_job(upstream, backend)
            ).result

            self.assertEqual(result.special_mode, "payout")
            self.assertEqual(result.a_count, 2)
            self.assertEqual(result.b_count, 2)
            self.assertEqual(diff_orders.unique_order_ids(result.a_only), ["300"])
            self.assertEqual(diff_orders.unique_order_ids(result.b_only), ["400"])
            self.assertEqual(result.mismatched, [])
            self.assertEqual(result.a_fee, Decimal("0.51"))
            self.assertEqual(result.b_fee, Decimal("1.88"))
            self.assertEqual(result.channel_cost, Decimal("0.52"))

    def test_directory_recognition_uses_headers(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            upstream = directory / "renamed.csv"
            backend = directory / "renamed.xlsx"
            self.write_upstream(upstream)
            self.write_backend(backend)

            jobs = payout_diff.find_jobs_in_directory(directory)

            self.assertEqual(len(jobs), 1)
            self.assertEqual(jobs[0].upstream_path, upstream)
            self.assertEqual(jobs[0].backend_path, backend)

    def test_payout_html_contains_both_profit_metrics(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            upstream = directory / "上游.csv"
            backend = directory / "付款订单.xlsx"
            self.write_upstream(upstream)
            self.write_backend(backend)
            job_result = payout_diff.read_job_diff_result(
                payout_diff.make_job(upstream, backend)
            )

            html = diff_orders.render_stats_html([job_result])

            self.assertIn("毛利润（我方付款金额×1.5%）", html)
            self.assertIn("渠道利润参考（平台手续费-渠道成本）", html)

    def test_payout_gross_profit_is_one_point_five_percent_of_own_amount(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            upstream = directory / "上游.csv"
            backend = directory / "付款订单.xlsx"
            self.write_upstream(upstream)
            self.write_backend(backend)
            job_result = payout_diff.read_job_diff_result(
                payout_diff.make_job(upstream, backend)
            )

            html = diff_orders.render_stats_html([job_result])

            # 我方成功付款金额为 125 PKR，毛利润应为 125 * 1.5% = 1.875，展示 1.88。
            self.assertIn("毛利润（我方付款金额×1.5%）", html)
            self.assertIn(
                '<td class="num">1.88</td>\n            '
                '<td class="num">1.88</td>\n            '
                '<td class="num">1.36</td>',
                html,
            )

    def test_finerbit_uses_collection_channels_and_payout_columns(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            collection = directory / "Transaction Details.csv"
            disbursement = directory / "Disbursement Transaction Details.csv"
            backend = directory / "付款订单.xlsx"
            self.write_finerbit_collection(collection)
            self.write_finerbit_disbursement(disbursement)
            self.write_finerbit_backend(backend)

            job = payout_diff.make_job(
                disbursement,
                backend,
                algorithm="finerbit",
                collection_path=collection,
            )
            with patch(
                "autoexcel.payout_diff.load_finerbit_fee_rates",
                return_value=(Decimal("0.013"), Decimal("0.0125")),
            ):
                result = payout_diff.read_job_diff_result(job).result

            self.assertEqual(result.special_mode, "payout_finerbit")
            self.assertEqual(result.a_amount, Decimal("300"))
            self.assertEqual(result.b_amount, Decimal("150"))
            self.assertEqual(result.a_fee, Decimal("3.8000"))
            self.assertEqual(result.b_fee, Decimal("7"))
            self.assertEqual(result.channel_cost, Decimal("3"))
            self.assertEqual(diff_orders.unique_order_ids(result.a_only), ["P2"])
            self.assertEqual(diff_orders.unique_order_ids(result.b_only), ["P4"])

    def test_finerbit_html_uses_finerbit_profit_labels(self) -> None:
        with TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            collection = directory / "Transaction Details.csv"
            disbursement = directory / "Disbursement Transaction Details.csv"
            backend = directory / "付款订单.xlsx"
            self.write_finerbit_collection(collection)
            self.write_finerbit_disbursement(disbursement)
            self.write_finerbit_backend(backend)
            job = payout_diff.make_job(
                disbursement,
                backend,
                algorithm="finerbit",
                collection_path=collection,
            )
            with patch(
                "autoexcel.payout_diff.load_finerbit_fee_rates",
                return_value=(Decimal("0.013"), Decimal("0.0125")),
            ):
                job_result = payout_diff.read_job_diff_result(job)

            html = diff_orders.render_stats_html([job_result])

            self.assertIn("finerBit平台（交易金额）", html)
            self.assertIn("tarspay平台（finerBit交易成功金额）", html)
            self.assertIn("tarspay平台（finerBit渠道成本）", html)
            self.assertIn("毛利润（平台手续费-finerBit手续费）", html)
            self.assertIn('<td class="num">3.20</td>', html)
            self.assertNotIn("金额/笔数不一致", html)


if __name__ == "__main__":
    unittest.main()
