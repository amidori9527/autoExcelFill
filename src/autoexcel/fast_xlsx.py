from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
from datetime import date, datetime, timedelta
import os
from pathlib import Path, PurePosixPath
import re
import tempfile
from typing import Callable
import xml.etree.ElementTree as ET
import zipfile

from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.utils.datetime import to_excel


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

ET.register_namespace("", MAIN_NS)
ET.register_namespace("r", REL_NS)


@dataclass(frozen=True)
class SheetEntry:
    name: str
    path: str


@dataclass(frozen=True)
class FastBatchResult:
    changed: list[tuple[str, int]]
    skipped: list[tuple[str, str]]


def _tag(name: str) -> str:
    return f"{{{MAIN_NS}}}{name}"


def _rel_tag(name: str) -> str:
    return f"{{{PKG_REL_NS}}}{name}"


def _cell_ref(column: str, row: int) -> str:
    return f"{column}{row}"


def _split_cell_ref(ref: str) -> tuple[str, int]:
    match = re.fullmatch(r"([A-Z]+)([0-9]+)", ref)
    if not match:
        raise ValueError(f"Unsupported cell reference: {ref}")
    return match.group(1), int(match.group(2))


def _sheet_target_path(target: str) -> str:
    clean = target.lstrip("/")
    if clean.startswith("xl/"):
        return clean
    return str(PurePosixPath("xl") / clean)


def list_sheets(xlsx_path: Path) -> list[SheetEntry]:
    with zipfile.ZipFile(xlsx_path) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    relmap = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall(_rel_tag("Relationship"))
    }
    sheets_node = workbook.find(_tag("sheets"))
    if sheets_node is None:
        return []

    entries: list[SheetEntry] = []
    rel_attr = f"{{{REL_NS}}}id"
    for sheet in sheets_node.findall(_tag("sheet")):
        rel_id = sheet.attrib[rel_attr]
        entries.append(SheetEntry(name=sheet.attrib["name"], path=_sheet_target_path(relmap[rel_id])))
    return entries


def _has_tab_color(root: ET.Element) -> bool:
    sheet_pr = root.find(_tag("sheetPr"))
    return sheet_pr is not None and sheet_pr.find(_tag("tabColor")) is not None


def _cell_value(cell: ET.Element) -> str | None:
    value = cell.find(_tag("v"))
    return value.text if value is not None else None


def _date_from_cell(cell: ET.Element | None) -> date | None:
    if cell is None:
        return None

    value = _cell_value(cell)
    if value is None:
        return None

    if cell.attrib.get("t") == "d":
        try:
            return datetime.fromisoformat(value).date()
        except ValueError:
            return None

    try:
        serial = float(value)
    except ValueError:
        return None

    # Excel stores dates as day serials. Comparing serials avoids needing the style table.
    ordinal = int(serial)
    # Keep this intentionally bounded: target ledgers use normal Excel date serials
    # around 46000 for 2026, while phone/account numbers can be huge numeric cells.
    if ordinal < 1 or ordinal > 100000:
        return None
    return datetime(1899, 12, 30).date() + timedelta(days=ordinal)


def _row_number(row: ET.Element) -> int:
    return int(row.attrib["r"])


def _cell_at(row: ET.Element, column: str) -> ET.Element | None:
    for cell in row.findall(_tag("c")):
        ref = cell.attrib.get("r")
        if ref and _split_cell_ref(ref)[0] == column:
            return cell
    return None


def _set_cell_row(cell: ET.Element, row_number: int) -> None:
    ref = cell.attrib.get("r")
    if not ref:
        return
    column, _ = _split_cell_ref(ref)
    cell.set("r", _cell_ref(column, row_number))


def _set_row_number(row: ET.Element, row_number: int) -> None:
    row.set("r", str(row_number))
    for cell in row.findall(_tag("c")):
        _set_cell_row(cell, row_number)


def _set_numeric_cell_value(cell: ET.Element, value: float) -> None:
    for formula in list(cell.findall(_tag("f"))):
        cell.remove(formula)
    for inline in list(cell.findall(_tag("is"))):
        cell.remove(inline)
    cell.attrib.pop("t", None)
    value_node = cell.find(_tag("v"))
    if value_node is None:
        value_node = ET.SubElement(cell, _tag("v"))
    value_node.text = str(int(value)) if float(value).is_integer() else str(value)


def _set_formula_cell_value(
    cell: ET.Element,
    formula_text: str,
    cached_value: str | None = None,
) -> None:
    for inline in list(cell.findall(_tag("is"))):
        cell.remove(inline)
    cell.attrib.pop("t", None)
    formula = cell.find(_tag("f"))
    if formula is None:
        formula = ET.Element(_tag("f"))
        cell.insert(0, formula)
    formula.text = formula_text
    value = cell.find(_tag("v"))
    if value is None:
        value = ET.SubElement(cell, _tag("v"))
    value.text = cached_value


def _numeric_cell_value(cell: ET.Element | None) -> float | None:
    if cell is None:
        return None
    value = _cell_value(cell)
    if value is None:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _format_number(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else str(value)


def _ensure_cell(row: ET.Element, column: str) -> ET.Element:
    cell = _cell_at(row, column)
    if cell is not None:
        return cell

    row_number = _row_number(row)
    cell = ET.Element(_tag("c"), {"r": _cell_ref(column, row_number)})
    target_index = 0
    target_col_index = range_boundaries(f"{column}1:{column}1")[0]
    for target_index, existing in enumerate(row.findall(_tag("c"))):
        existing_ref = existing.attrib.get("r")
        if not existing_ref:
            continue
        existing_col = range_boundaries(f"{_split_cell_ref(existing_ref)[0]}1:{_split_cell_ref(existing_ref)[0]}1")[0]
        if existing_col > target_col_index:
            row.insert(target_index, cell)
            return cell
    row.append(cell)
    return cell


def _repair_current_row_formulas(row: ET.Element, previous_row: ET.Element | None) -> bool:
    row_number = _row_number(row)
    if row_number <= 1:
        return False

    previous_row_number = row_number - 1
    changed = False

    b_cell = _ensure_cell(row, "B")
    b_formula = b_cell.find(_tag("f"))
    b_cache_value = _numeric_cell_value(_cell_at(previous_row, "K")) if previous_row is not None else None
    expected_b = f"K{previous_row_number}"
    expected_b_cache = _format_number(b_cache_value) if b_cache_value is not None else None
    if b_formula is None or b_formula.text != expected_b or _cell_value(b_cell) != expected_b_cache:
        _set_formula_cell_value(b_cell, expected_b, expected_b_cache)
        changed = True

    n_cell = _ensure_cell(row, "N")
    expected_n = f"M{row_number}-M{previous_row_number}"
    n_formula = n_cell.find(_tag("f"))
    current_m = _numeric_cell_value(_cell_at(row, "M"))
    previous_m = _numeric_cell_value(_cell_at(previous_row, "M")) if previous_row is not None else None
    n_cache_value = current_m - previous_m if current_m is not None and previous_m is not None else None
    expected_n_cache = _format_number(n_cache_value) if n_cache_value is not None else None
    if n_formula is None or n_formula.text != expected_n or _cell_value(n_cell) != expected_n_cache:
        _set_formula_cell_value(n_cell, expected_n, expected_n_cache)
        changed = True

    return changed


def _translate_formulas(row: ET.Element, original_row: int, new_row: int) -> None:
    for cell in row.findall(_tag("c")):
        formula = cell.find(_tag("f"))
        if formula is None or not formula.text:
            continue

        ref = cell.attrib.get("r")
        if not ref:
            continue
        column, _ = _split_cell_ref(ref)
        origin = _cell_ref(column, original_row)
        destination = _cell_ref(column, new_row)
        formula.text = Translator(f"={formula.text}", origin=origin).translate_formula(destination)[1:]


def _make_value_row(source_row: ET.Element, row_number: int) -> ET.Element:
    value_row = deepcopy(source_row)
    _set_row_number(value_row, row_number)
    for cell in value_row.findall(_tag("c")):
        for formula in list(cell.findall(_tag("f"))):
            cell.remove(formula)

        # If a formula has no cached value, leave the cell blank instead of keeping a formula.
        if cell.find(_tag("v")) is None:
            for child in list(cell):
                if child.tag != _tag("v"):
                    cell.remove(child)
    return value_row


def _update_dimension(root: ET.Element, inserted_row: int) -> None:
    dimension = root.find(_tag("dimension"))
    if dimension is None:
        return
    ref = dimension.attrib.get("ref")
    if not ref or ":" not in ref:
        return
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    if max_row >= inserted_row:
        max_row += 1
        dimension.set("ref", f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}")


def _add_current_date_to_sheet_xml(xml_bytes: bytes, current_date: date) -> tuple[bytes, int | None, str | None]:
    root = ET.fromstring(xml_bytes)
    if not _has_tab_color(root):
        return xml_bytes, None, "no tab color"

    sheet_data = root.find(_tag("sheetData"))
    if sheet_data is None:
        return xml_bytes, None, "no sheetData"

    rows = sheet_data.findall(_tag("row"))
    row_by_number = {_row_number(row): row for row in rows}
    source_date = current_date - timedelta(days=1)
    current_serial = to_excel(current_date)

    last_date_row: ET.Element | None = None
    for row in reversed(rows):
        row_date = _date_from_cell(_cell_at(row, "A"))
        if row_date == current_date:
            if _repair_current_row_formulas(row, row_by_number.get(_row_number(row) - 1)):
                return (
                    ET.tostring(root, encoding="utf-8", xml_declaration=True),
                    _row_number(row),
                    None,
                )
            return xml_bytes, None, f"{current_date:%Y-%m-%d} already exists"
        if last_date_row is None and row_date is not None:
            last_date_row = row

    if last_date_row is None:
        return xml_bytes, None, "no date row"

    source_row_number = _row_number(last_date_row)
    last_date = _date_from_cell(_cell_at(last_date_row, "A"))
    if last_date != source_date:
        found = last_date.strftime("%Y-%m-%d") if last_date else "not a date"
        return xml_bytes, None, f"expected last date {source_date:%Y-%m-%d}, found {found}"

    value_row = _make_value_row(last_date_row, source_row_number)

    for row in rows:
        old_row_number = _row_number(row)
        if old_row_number < source_row_number:
            continue
        new_row_number = old_row_number + 1
        _translate_formulas(row, original_row=old_row_number, new_row=new_row_number)
        _set_row_number(row, new_row_number)

    moved_source = last_date_row
    date_cell = _cell_at(moved_source, "A")
    if date_cell is None:
        date_cell = ET.Element(_tag("c"), {"r": _cell_ref("A", source_row_number + 1)})
        moved_source.insert(0, date_cell)
    _set_numeric_cell_value(date_cell, current_serial)
    _repair_current_row_formulas(moved_source, value_row)

    insert_at = list(sheet_data).index(moved_source)
    sheet_data.insert(insert_at, value_row)
    _update_dimension(root, source_row_number)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True), source_row_number, None


def _force_full_calculation(workbook_xml: bytes) -> bytes:
    root = ET.fromstring(workbook_xml)
    calc_pr = root.find(_tag("calcPr"))
    if calc_pr is None:
        calc_pr = ET.SubElement(root, _tag("calcPr"))
    calc_pr.set("calcMode", "auto")
    calc_pr.set("fullCalcOnLoad", "1")
    calc_pr.set("forceFullCalc", "1")
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def add_current_date_to_colored_sheets_fast(
    xlsx_path: Path,
    current_date: date,
    limit_sheets: int,
    progress: Callable[[str], None] | None = None,
) -> FastBatchResult:
    entries = list_sheets(xlsx_path)
    changed: list[tuple[str, int]] = []
    skipped: list[tuple[str, str]] = []
    modified_xml: dict[str, bytes] = {}

    with zipfile.ZipFile(xlsx_path, "r") as archive:
        total_entries = len(entries)
        for index, entry in enumerate(entries, start=1):
            if len(changed) >= limit_sheets:
                break

            if progress is not None:
                progress(f"  scanning sheet {index}/{total_entries}: {entry.name}")

            xml_bytes = archive.read(entry.path)
            new_xml, inserted_row, reason = _add_current_date_to_sheet_xml(xml_bytes, current_date)
            if inserted_row is None:
                if reason != "no tab color":
                    skipped.append((entry.name, reason or "not changed"))
                continue

            modified_xml[entry.path] = new_xml
            changed.append((entry.name, inserted_row))

        if not modified_xml:
            return FastBatchResult(changed=changed, skipped=skipped)

        if progress is not None:
            progress(f"  writing workbook: {len(changed)} sheet(s) changed...")

        modified_xml["xl/workbook.xml"] = _force_full_calculation(archive.read("xl/workbook.xml"))

        fd, tmp_name = tempfile.mkstemp(
            prefix=f"{xlsx_path.stem}.",
            suffix=".tmp.xlsx",
            dir=xlsx_path.parent,
        )
        os.close(fd)
        tmp_path = Path(tmp_name)
        try:
            with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as output:
                for item in archive.infolist():
                    output.writestr(item, modified_xml.get(item.filename, archive.read(item.filename)))
        except Exception:
            if tmp_path.exists():
                tmp_path.unlink()
            raise

    try:
        if progress is not None:
            progress("  replacing original workbook...")
        os.replace(tmp_path, xlsx_path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()

    return FastBatchResult(changed=changed, skipped=skipped)
