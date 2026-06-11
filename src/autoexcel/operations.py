from __future__ import annotations

from copy import copy
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl.formula.translate import Translator
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def copy_row_format(sheet: Worksheet, source_row: int, target_row: int) -> None:
    """Copy row height, cell styles, number formats, comments, and hyperlinks."""
    source_dimension = sheet.row_dimensions[source_row]
    target_dimension = sheet.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden

    for source_cell in sheet[source_row]:
        target_cell = sheet.cell(row=target_row, column=source_cell.column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)
        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)


def copy_row(sheet: Worksheet, source_row: int, target_row: int) -> None:
    copy_row_format(sheet, source_row, target_row)
    for source_cell in sheet[source_row]:
        target_cell = sheet.cell(row=target_row, column=source_cell.column)
        target_cell.value = source_cell.value


def paste_row_values(sheet: Worksheet, target_row: int, values: list[Any], start_column: int = 1) -> None:
    for offset, value in enumerate(values):
        sheet.cell(row=target_row, column=start_column + offset).value = value


def set_formula(sheet: Worksheet, cell_address: str, formula: str) -> None:
    sheet[cell_address] = formula if formula.startswith("=") else f"={formula}"


def translate_row_formulas(sheet: Worksheet, formula_row: int, original_row: int) -> None:
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=formula_row, column=column)
        if not (isinstance(cell.value, str) and cell.value.startswith("=")):
            continue

        origin = sheet.cell(row=original_row, column=column).coordinate
        destination = cell.coordinate
        cell.value = Translator(cell.value, origin=origin).translate_formula(destination)


def find_row_by_date(sheet: Worksheet, target_date: date, date_column: int = 1) -> int:
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row=row, column=date_column).value
        if isinstance(value, datetime):
            value = value.date()
        if value == target_date:
            return row
    raise ValueError(f"Date {target_date:%Y-%m-%d} not found in sheet '{sheet.title}'.")


def cell_date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def find_last_date_row(sheet: Worksheet, date_column: int = 1) -> int:
    for row in range(sheet.max_row, 0, -1):
        if cell_date_value(sheet.cell(row=row, column=date_column).value) is not None:
            return row
    raise ValueError(f"No date row found in sheet '{sheet.title}'.")


def freeze_next_day_row(
    workbook: Workbook,
    values_workbook: Workbook,
    sheet_name: str,
    current_date: date,
) -> int:
    sheet = workbook[sheet_name]
    values_sheet = values_workbook[sheet_name]

    source_date = current_date - timedelta(days=1)
    source_row = find_last_date_row(sheet)
    actual_source_date = cell_date_value(sheet.cell(row=source_row, column=1).value)
    if actual_source_date != source_date:
        found = actual_source_date.strftime("%Y-%m-%d") if actual_source_date else "not a date"
        raise ValueError(
            f"Expected last date row to be {source_date:%Y-%m-%d}; "
            f"found {found} at row {source_row}."
        )

    source_values = [
        values_sheet.cell(row=source_row, column=column).value
        for column in range(1, sheet.max_column + 1)
    ]

    sheet.insert_rows(source_row)
    copy_row_format(sheet, source_row + 1, source_row)
    paste_row_values(sheet, source_row, source_values)
    translate_row_formulas(sheet, formula_row=source_row + 1, original_row=source_row)
    sheet.cell(row=source_row + 1, column=1).value = current_date
    return source_row


def sheet_has_tab_color(sheet: Worksheet) -> bool:
    return sheet.sheet_properties.tabColor is not None


def freeze_colored_sheets_next_day(
    workbook: Workbook,
    values_workbook: Workbook,
    current_date: date,
    limit_sheets: int | None = None,
) -> tuple[list[tuple[str, int]], list[tuple[str, str]]]:
    changed: list[tuple[str, int]] = []
    skipped: list[tuple[str, str]] = []

    for sheet in workbook.worksheets:
        if limit_sheets is not None and len(changed) >= limit_sheets:
            break

        if not sheet_has_tab_color(sheet):
            continue

        if sheet.title not in values_workbook.sheetnames:
            skipped.append((sheet.title, "missing in values workbook"))
            continue

        try:
            find_row_by_date(sheet, current_date)
        except ValueError:
            pass
        else:
            skipped.append((sheet.title, f"{current_date:%Y-%m-%d} already exists"))
            continue

        try:
            inserted_row = freeze_next_day_row(
                workbook=workbook,
                values_workbook=values_workbook,
                sheet_name=sheet.title,
                current_date=current_date,
            )
        except ValueError as exc:
            skipped.append((sheet.title, str(exc)))
            continue

        changed.append((sheet.title, inserted_row))

    return changed, skipped
