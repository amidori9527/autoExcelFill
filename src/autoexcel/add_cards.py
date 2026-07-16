from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
import os
from pathlib import Path, PurePosixPath
import re
import sys
import tempfile
import xml.etree.ElementTree as ET
import zipfile

from openpyxl.utils.cell import get_column_letter, range_boundaries

from autoexcel.fast_xlsx import (
    MAIN_NS,
    PKG_REL_NS,
    REL_NS,
    _cell_at,
    _ensure_cell,
    _force_full_calculation,
    _has_tab_color,
    _row_number,
    _set_row_number,
    _set_numeric_cell_value,
    _sheet_target_path,
    _split_cell_ref,
    _tag,
    _translate_formulas,
)
from autoexcel.main import choose_workbook_from_current_directory, pause_before_exit


CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
WORKSHEET_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
)
CHINESE_RE = re.compile(r"[\u4e00-\u9fff]")
INVALID_SHEET_NAME_RE = re.compile(r"[:\\\\/?*\[\]]")

ET.register_namespace("", CONTENT_TYPES_NS)


@dataclass(frozen=True)
class AddCardsResult:
    created: list[str]
    skipped: list[tuple[str, str]]
    template_name: str


def _rel_tag(name: str) -> str:
    return f"{{{PKG_REL_NS}}}{name}"


def _content_tag(name: str) -> str:
    return f"{{{CONTENT_TYPES_NS}}}{name}"


def _cell_text(cell: ET.Element, shared_strings: list[str]) -> str | None:
    if cell.attrib.get("t") == "s":
        value = cell.find(_tag("v"))
        if value is None or value.text is None:
            return None
        return shared_strings[int(value.text)]
    if cell.attrib.get("t") == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(f".//{_tag('t')}"))
    value = cell.find(_tag("v"))
    return value.text if value is not None else None


def _shared_strings(archive: zipfile.ZipFile) -> list[str]:
    path = "xl/sharedStrings.xml"
    if path not in archive.namelist():
        return []
    root = ET.fromstring(archive.read(path))
    return ["".join(node.text or "" for node in item.findall(f".//{_tag('t')}")) for item in root.findall(_tag("si"))]


def _template_columns(root: ET.Element, shared_strings: list[str]) -> tuple[str, str]:
    sheet_data = root.find(_tag("sheetData"))
    if sheet_data is None:
        raise ValueError("模板没有工作表数据")
    header = next((row for row in sheet_data.findall(_tag("row")) if _row_number(row) == 2), None)
    if header is None:
        raise ValueError("模板缺少第 2 行表头")
    columns: dict[str, str] = {}
    for cell in header.findall(_tag("c")):
        ref = cell.attrib.get("r")
        text = _cell_text(cell, shared_strings)
        if ref and text in {"期初余额", "增量"}:
            columns[text] = _split_cell_ref(ref)[0]
    if "期初余额" not in columns or "增量" not in columns:
        raise ValueError("模板第 2 行缺少“期初余额”或“增量”列")
    return columns["期初余额"], columns["增量"]


def _card_sheet_xml(
    template_xml: bytes,
    shared_strings: list[str],
) -> bytes:
    root = ET.fromstring(template_xml)
    opening_balance_column, increment_column = _template_columns(root, shared_strings)
    sheet_data = root.find(_tag("sheetData"))
    if sheet_data is None:
        raise ValueError("模板没有工作表数据")
    rows = sheet_data.findall(_tag("row"))
    if len(rows) < 3:
        raise ValueError("模板至少需要第 1、2 行和最后一行")

    row_by_number = {_row_number(row): row for row in rows}
    if 1 not in row_by_number or 2 not in row_by_number:
        raise ValueError("模板缺少第 1 行或第 2 行")
    last_row = max(rows, key=_row_number)
    source_last_row_number = _row_number(last_row)
    if source_last_row_number <= 2:
        raise ValueError("模板缺少最后数据行")

    selected_rows = [deepcopy(row_by_number[1]), deepcopy(row_by_number[2]), deepcopy(last_row)]
    copied_last_row = selected_rows[2]
    if source_last_row_number != 3:
        _translate_formulas(copied_last_row, source_last_row_number, 3)
        _set_row_number(copied_last_row, 3)
    _set_numeric_cell_value(_ensure_cell(copied_last_row, opening_balance_column), 0)
    _set_numeric_cell_value(_ensure_cell(copied_last_row, increment_column), 0)

    for row in list(sheet_data):
        sheet_data.remove(row)
    for row in selected_rows:
        sheet_data.append(row)
    dimension = root.find(_tag("dimension"))
    if dimension is not None and dimension.attrib.get("ref"):
        min_col, min_row, max_col, _ = range_boundaries(dimension.attrib["ref"])
        dimension.set("ref", f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}3")
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _valid_card_name(name: str) -> str | None:
    if not name:
        return "empty card number"
    if len(name) > 31:
        return "sheet name exceeds 31 characters"
    if INVALID_SHEET_NAME_RE.search(name):
        return "invalid sheet name character"
    return None


def _next_number(values: list[str], prefix: str) -> int:
    numbers = [int(value[len(prefix):]) for value in values if value.startswith(prefix) and value[len(prefix):].isdigit()]
    return max(numbers, default=0) + 1


def _next_sheet_number(existing_paths: list[str]) -> int:
    used_numbers: set[int] = set()
    for path in existing_paths:
        match = re.fullmatch(r"xl/worksheets/(?:_rels/)?sheet(\d+)\.xml(?:\.rels)?", path)
        if match:
            used_numbers.add(int(match.group(1)))
    return max(used_numbers, default=0) + 1


def parse_card_numbers(text: str) -> list[str]:
    return [card.strip() for card in text.splitlines() if card.strip()]


def read_card_numbers_from_terminal() -> list[str]:
    if not sys.stdin.isatty():
        raise RuntimeError("增卡需要在可输入的终端中运行。")

    print("请粘贴多行卡号；粘贴完成后，在空白行按 Enter 提交。")
    print("示例：1234、3121、1341 每个卡号一行。")
    rows: list[str] = []
    while True:
        raw = input("卡号：" if not rows else "")
        if not raw.strip():
            return parse_card_numbers("\n".join(rows))
        rows.append(raw)


def add_cards_to_workbook(xlsx_path: Path, card_numbers: list[str]) -> AddCardsResult:
    """Insert card sheets before the first non-Chinese sheet with a tab color."""
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在：{xlsx_path}")
    if xlsx_path.suffix.lower() != ".xlsx":
        raise ValueError("请选择 .xlsx 文件")

    cards = [card.strip() for card in card_numbers if card.strip()]
    with zipfile.ZipFile(xlsx_path, "r") as archive:
        workbook_xml = archive.read("xl/workbook.xml")
        workbook = ET.fromstring(workbook_xml)
        rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        sheets = workbook.find(_tag("sheets"))
        if sheets is None:
            raise ValueError("Excel 文件没有工作表")
        rel_attr = f"{{{REL_NS}}}id"
        relation_targets = {
            relation.attrib["Id"]: relation.attrib["Target"]
            for relation in rels.findall(_rel_tag("Relationship"))
        }
        template_sheet = next(
            (
                sheet
                for sheet in sheets.findall(_tag("sheet"))
                if not CHINESE_RE.search(sheet.attrib["name"])
                and _has_tab_color(
                    ET.fromstring(
                        archive.read(_sheet_target_path(relation_targets[sheet.attrib[rel_attr]]))
                    )
                )
            ),
            None,
        )
        if template_sheet is None:
            raise ValueError("未找到不含中文且带颜色的模板工作表")

        template_name = template_sheet.attrib["name"]
        template_path = _sheet_target_path(relation_targets[template_sheet.attrib[rel_attr]])
        new_sheet_xml = _card_sheet_xml(archive.read(template_path), _shared_strings(archive))
        existing_names = {sheet.attrib["name"].casefold() for sheet in sheets.findall(_tag("sheet"))}
        seen_inputs: set[str] = set()
        created: list[str] = []
        skipped: list[tuple[str, str]] = []
        for card in cards:
            key = card.casefold()
            reason = _valid_card_name(card)
            if reason:
                skipped.append((card, reason))
            elif key in seen_inputs:
                skipped.append((card, "duplicate input"))
            elif key in existing_names:
                skipped.append((card, "sheet already exists"))
            else:
                created.append(card)
                seen_inputs.add(key)
                existing_names.add(key)

        if not created:
            return AddCardsResult(created, skipped, template_name)

        existing_paths = [item.filename for item in archive.infolist()]
        sheet_number = _next_sheet_number(existing_paths)
        relationship_number = _next_number(
            [relation.attrib["Id"] for relation in rels.findall(_rel_tag("Relationship"))], "rId"
        )
        sheet_id = max((int(sheet.attrib.get("sheetId", "0")) for sheet in sheets.findall(_tag("sheet"))), default=0) + 1
        new_parts: dict[str, bytes] = {}
        for card in created:
            sheet_path = f"xl/worksheets/sheet{sheet_number}.xml"
            relationship_id = f"rId{relationship_number}"
            ET.SubElement(
                rels,
                _rel_tag("Relationship"),
                {"Id": relationship_id, "Type": WORKSHEET_REL_TYPE, "Target": f"worksheets/sheet{sheet_number}.xml"},
            )
            new_sheet = ET.Element(
                _tag("sheet"),
                {"name": card, "sheetId": str(sheet_id), rel_attr: relationship_id},
            )
            sheets.insert(list(sheets).index(template_sheet), new_sheet)
            new_parts[sheet_path] = new_sheet_xml
            sheet_number += 1
            relationship_number += 1
            sheet_id += 1

        content_types = ET.fromstring(archive.read("[Content_Types].xml"))
        for path in new_parts:
            ET.SubElement(
                content_types,
                _content_tag("Override"),
                {"PartName": f"/{path}", "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"},
            )
        new_parts["xl/workbook.xml"] = _force_full_calculation(
            ET.tostring(workbook, encoding="utf-8", xml_declaration=True)
        )
        new_parts["xl/_rels/workbook.xml.rels"] = ET.tostring(rels, encoding="utf-8", xml_declaration=True)
        new_parts["[Content_Types].xml"] = ET.tostring(content_types, encoding="utf-8", xml_declaration=True)

        fd, temporary_name = tempfile.mkstemp(prefix=f"{xlsx_path.stem}.", suffix=".tmp.xlsx", dir=xlsx_path.parent)
        os.close(fd)
        temporary_path = Path(temporary_name)
        try:
            with zipfile.ZipFile(temporary_path, "w", compression=zipfile.ZIP_DEFLATED) as output:
                for item in archive.infolist():
                    output.writestr(item, new_parts.get(item.filename, archive.read(item.filename)))
                for path, content in new_parts.items():
                    if path not in existing_paths:
                        output.writestr(path, content)
        except Exception:
            if temporary_path.exists():
                temporary_path.unlink()
            raise

    try:
        os.replace(temporary_path, xlsx_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return AddCardsResult(created, skipped, template_name)


def main() -> int:
    try:
        print("Excel 增卡工具")
        print("----------------------------------------")
        selected_file = choose_workbook_from_current_directory()
        cards = read_card_numbers_from_terminal()
        if not cards:
            print("未输入卡号，已取消。")
            return 0
        result = add_cards_to_workbook(selected_file, cards)
    except Exception as error:
        print(f"增卡失败：{error}")
        return 1
    else:
        print(f"模板：{result.template_name}")
        print(f"新增 {len(result.created)} 张卡：{', '.join(result.created) or '无'}")
        if result.skipped:
            print("跳过：" + "；".join(f"{card}（{reason}）" for card, reason in result.skipped))
        return 0
    finally:
        pause_before_exit()


if __name__ == "__main__":
    raise SystemExit(main())
