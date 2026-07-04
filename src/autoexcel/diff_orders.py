from __future__ import annotations

import argparse
import configparser
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from html import escape
import json
from pathlib import Path
import re
import subprocess
import sys
import traceback
from typing import Any, Callable

from openpyxl import load_workbook

from autoexcel.version import version_text
from openpyxl.utils import column_index_from_string


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DIR = PROJECT_ROOT / "workspace" / "diffOrders"
DEFAULT_TEMPLATE = PROJECT_ROOT / "template" / "order_diff.html"
DEFAULT_RESULT_DIR = PROJECT_ROOT / "result"
WORKSPACE_DIR_NAME = "workspace"
DIFF_ORDERS_DIR_NAME = "diffOrders"
TEMPLATE_FILE_NAME = "order_diff.html"
CONFIG_FILE_NAME = "config.ini"
GROUP_CONFIG_FILE_NAME = "conf.ini"
HEADER_KEYWORDS = {
    "TRANSACTION REFERENCE NUMBER",
    "ReferenceId",
    "请求上游订单号",
    "请求上游ID",
}
UPSTREAM_FILE_PREFIX = "TranDetailReport"
FINERBIT_UPSTREAM_FILE_PREFIX = "Transaction Details"
EASYPISA_UPSTREAM_FILE_PREFIX = "TransactionHistoryRecords"
BACKEND_FILE_PREFIX = "收款订单"
DUPLICATE_PAYMENT_FILE_KEYWORD = "重复支付订单"


@dataclass(frozen=True)
class OrderEntry:
    order_id: str
    row_number: int
    amount: Decimal = Decimal("0")
    fee: Decimal = Decimal("0")
    payment_method: str = ""
    sheet_name: str = ""
    extra: dict[str, Any] | None = None


@dataclass(frozen=True)
class DiffResult:
    a_entries: list[OrderEntry]
    b_entries: list[OrderEntry]
    a_only: list[OrderEntry]
    b_only: list[OrderEntry]
    mismatched: list[str]
    a_count: int
    b_count: int
    a_row_count: int
    b_row_count: int
    common_count: int
    a_duplicate_count: int
    b_duplicate_count: int
    a_amount: Decimal
    b_amount: Decimal
    a_fee: Decimal
    b_fee: Decimal
    a_only_amount: Decimal
    b_only_amount: Decimal
    mismatch_amount: Decimal
    duplicate_payment_entries: tuple[OrderEntry, ...] = ()
    repeated_difference_entries: tuple[OrderEntry, ...] = ()
    remaining_difference_entries: tuple[OrderEntry, ...] = ()
    channel_cost: Decimal = Decimal("0")
    special_mode: str = ""


@dataclass(frozen=True)
class DiffJob:
    upstream_path: Path
    backend_path: Path
    duplicate_path: Path | None = None
    platform_mode: str = ""


@dataclass(frozen=True)
class JobDiffResult:
    job: DiffJob
    result: DiffResult


@dataclass(frozen=True)
class SummaryMetric:
    label: str
    count: int
    amount: Decimal
    fee: Decimal | None
    unique_count: int
    note: str
    order_ids: list[str]


def is_frozen_app() -> bool:
    return getattr(sys, "frozen", False)


def get_executable_directory() -> Path:
    if is_frozen_app():
        return Path(sys.executable).resolve().parent
    return PROJECT_ROOT


def format_file_size(path: Path) -> str:
    size = path.stat().st_size
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024 or unit == "GB":
            return f"{size:.1f} {unit}" if unit != "B" else f"{size} {unit}"
        size /= 1024
    return f"{size:.1f} GB"


def list_xlsx_files(directory: Path) -> list[Path]:
    return sorted(
        (
            path
            for path in directory.rglob("*.xlsx")
            if path.is_file() and not path.name.startswith(("~$", ".~", "._"))
        ),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )


def iter_diff_orders_directories() -> list[Path]:
    directories: list[Path] = []
    if is_frozen_app():
        base_directories = (get_executable_directory(), Path.cwd(), PROJECT_ROOT)
    else:
        base_directories = (Path.cwd(), PROJECT_ROOT)
    for base_directory in base_directories:
        workspace_dir = base_directory / WORKSPACE_DIR_NAME
        candidates = [
            workspace_dir / DIFF_ORDERS_DIR_NAME,
            workspace_dir / DIFF_ORDERS_DIR_NAME.lower(),
            workspace_dir / DIFF_ORDERS_DIR_NAME.upper(),
        ]
        if workspace_dir.is_dir():
            candidates.extend(
                path
                for path in workspace_dir.iterdir()
                if path.is_dir() and path.name.lower() == DIFF_ORDERS_DIR_NAME.lower()
            )
        for diff_dir in candidates:
            if diff_dir not in directories:
                directories.append(diff_dir)
    return directories


def find_xlsx_directory() -> tuple[Path, list[Path]]:
    checked: list[Path] = []
    for directory in iter_diff_orders_directories():
        checked.append(directory)
        if not directory.is_dir():
            continue
        files = list_xlsx_files(directory)
        if files:
            return directory, files
    checked_text = "\n  ".join(str(path) for path in checked)
    raise FileNotFoundError(
        f"没有在 workspace/diffOrders 文件夹里找到 xlsx 文件。"
        f"请把 Excel 放入 workspace/diffOrders 后再运行。\n"
        f"已检查目录：\n  {checked_text}"
    )


def parse_date(value: str) -> date:
    raw_value = value.strip()
    current_year = date.today().year

    if re.fullmatch(r"\d{4}", raw_value):
        month = int(raw_value[:2])
        day = int(raw_value[2:])
        return date(current_year, month, day)

    normalized = raw_value.replace("/", "-").replace(".", "-")
    if re.fullmatch(r"\d{1,2}-\d{1,2}", normalized):
        month_text, day_text = normalized.split("-")
        return date(current_year, int(month_text), int(day_text))

    for date_format in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(raw_value, date_format).date()
        except ValueError:
            continue

    raise ValueError("日期格式不对，请输入 2026-06-10、0610、06-10 或 05/12")


def default_target_date() -> date:
    return date.today()


def choose_current_date() -> str:
    default_date = default_target_date().strftime("%Y-%m-%d")
    while True:
        raw_value = input(
            f"请输入目标日期，直接回车默认当天 {default_date}；示例：2026-06-10、0610、06-10、05/12："
        ).strip()
        selected_value = raw_value or default_date
        try:
            return parse_date(selected_value).strftime("%Y-%m-%d")
        except ValueError as error:
            print(error)


def resolve_target_date(value: str | None) -> date:
    if value:
        return parse_date(value)
    return default_target_date()


def match_upstream_workbook(path: Path, date_text: str) -> bool:
    pattern = rf"{re.escape(UPSTREAM_FILE_PREFIX)}[_\-\s]+[^_\-\s]+[_\-\s]+{date_text}"
    if re.search(pattern, path.stem) is not None:
        return True
    return path.stem.startswith(FINERBIT_UPSTREAM_FILE_PREFIX) and date_text in path.stem


def match_backend_workbook(path: Path, date_text: str) -> bool:
    return BACKEND_FILE_PREFIX in path.stem and date_text in path.stem


def match_duplicate_payment_workbook(path: Path, date_text: str) -> bool:
    return DUPLICATE_PAYMENT_FILE_KEYWORD in path.stem and date_text in path.stem


def latest_matching_workbook(
    files: list[Path],
    matcher: Callable[[Path, str], bool],
    label: str,
    date_text: str,
) -> Path:
    matches = [path for path in files if matcher(path, date_text)]
    if not matches:
        available = "\n  ".join(path.name for path in files) or "无"
        raise FileNotFoundError(
            f"没有找到 {label} Excel。目标日期：{date_text}。\n"
            f"上游文件名示例：TranDetailReport_87382398_{date_text}013343.1587508.xlsx\n"
            f"后台文件名示例：收款订单_{date_text}063100.xlsx\n"
            f"当前目录文件：\n  {available}"
        )
    return sorted(matches, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def matching_workbooks(
    files: list[Path],
    matcher: Callable[[Path, str], bool],
    label: str,
    date_text: str,
) -> list[Path]:
    matches = sorted((path for path in files if matcher(path, date_text)), key=lambda path: path.name)
    if not matches:
        available = "\n  ".join(path.name for path in files) or "无"
        raise FileNotFoundError(
            f"没有找到 {label} Excel。目标日期：{date_text}。\n"
            f"上游文件名示例：TranDetailReport_87382398_{date_text}013343.1587508.xlsx\n"
            f"后台文件名示例：收款订单_{date_text}063100.xlsx\n"
            f"当前目录文件：\n  {available}"
        )
    return matches


def matching_workbook_candidates(
    files: list[Path],
    matcher: Callable[[Path, str], bool],
    date_text: str,
) -> list[Path]:
    return sorted((path for path in files if matcher(path, date_text)), key=lambda path: path.name)


def easypaisa_upstream_paths_for_backends(files: list[Path], backend_paths: list[Path]) -> list[Path]:
    backend_dirs = {path.parent for path in backend_paths}
    return sorted(
        (
            path
            for path in files
            if path.parent in backend_dirs and path.stem == EASYPISA_UPSTREAM_FILE_PREFIX
        ),
        key=lambda path: path.name,
    )


def duplicate_path_for_pair(upstream_path: Path, backend_path: Path, duplicate_paths: list[Path]) -> Path | None:
    same_directory = [
        path for path in duplicate_paths if path.parent in {upstream_path.parent, backend_path.parent}
    ]
    if same_directory:
        return same_directory[0]
    if len(duplicate_paths) == 1:
        return duplicate_paths[0]
    return None


def group_directory_for_pair(upstream_path: Path, backend_path: Path) -> Path:
    if upstream_path.parent == backend_path.parent:
        return upstream_path.parent
    return upstream_path.parent


def read_group_platform(group_dir: Path) -> str | None:
    config_path = group_dir / GROUP_CONFIG_FILE_NAME
    if not config_path.exists():
        return None
    parser = configparser.ConfigParser()
    try:
        parser.read(config_path, encoding="utf-8")
    except configparser.Error:
        try:
            text = config_path.read_text(encoding="utf-8")
        except OSError:
            return None
        for line in text.splitlines():
            match = re.match(r"\s*platform\s*=\s*([^\s#;]+)", line, flags=re.IGNORECASE)
            if match:
                return match.group(1).strip().lower() or None
        return None
    for section in ("diff_orders", "platform"):
        if parser.has_option(section, "platform"):
            return parser.get(section, "platform", fallback="").strip().lower() or None
    default_platform = parser.defaults().get("platform", "").strip().lower()
    if default_platform:
        return default_platform
    return None


def platform_mode_for_pair(upstream_path: Path, backend_path: Path, duplicate_path: Path | None) -> str:
    configured_mode = read_group_platform(group_directory_for_pair(upstream_path, backend_path))
    if configured_mode:
        return configured_mode
    if duplicate_path is not None or upstream_path.stem.startswith(FINERBIT_UPSTREAM_FILE_PREFIX):
        return "finerbit"
    return ""


def make_diff_job(upstream_path: Path, backend_path: Path, duplicate_paths: list[Path]) -> DiffJob:
    duplicate_path = duplicate_path_for_pair(upstream_path, backend_path, duplicate_paths)
    return DiffJob(
        upstream_path=upstream_path,
        backend_path=backend_path,
        duplicate_path=duplicate_path,
        platform_mode=platform_mode_for_pair(upstream_path, backend_path, duplicate_path),
    )


def pair_diff_workbooks(
    upstream_paths: list[Path],
    backend_paths: list[Path],
    duplicate_paths: list[Path] | None = None,
) -> list[DiffJob]:
    duplicate_paths = duplicate_paths or []
    upstream_by_key = {key: path for path in upstream_paths if (key := platform_key_from_path(path))}
    backend_by_key = {key: path for path in backend_paths if (key := platform_key_from_path(path))}
    shared_keys = sorted(upstream_by_key.keys() & backend_by_key.keys())
    if shared_keys and len(shared_keys) == len(upstream_paths) == len(backend_paths):
        return [
            make_diff_job(upstream_by_key[key], backend_by_key[key], duplicate_paths)
            for key in shared_keys
        ]

    if len(upstream_paths) == len(backend_paths):
        return [
            make_diff_job(upstream_path, backend_path, duplicate_paths)
            for upstream_path, backend_path in zip(upstream_paths, backend_paths)
        ]
    if len(backend_paths) == 1:
        return [
            make_diff_job(upstream_path, backend_paths[0], duplicate_paths)
            for upstream_path in upstream_paths
        ]
    if len(upstream_paths) == 1:
        return [
            make_diff_job(upstream_paths[0], backend_path, duplicate_paths)
            for backend_path in backend_paths
        ]

    raise ValueError(
        f"自动匹配到 {len(upstream_paths)} 个上游文件、{len(backend_paths)} 个后台文件，无法确定配对关系。"
        "请保证两边文件数量一致，或其中一边只有 1 个文件。"
    )


def find_diff_jobs_by_date(target_date: date) -> tuple[Path, list[DiffJob]]:
    directory, files = find_xlsx_directory()
    date_text = target_date.strftime("%Y%m%d")
    backend_paths = matching_workbooks(files, match_backend_workbook, "后台", date_text)
    upstream_paths = matching_workbook_candidates(files, match_upstream_workbook, date_text)
    upstream_paths.extend(
        path for path in easypaisa_upstream_paths_for_backends(files, backend_paths) if path not in upstream_paths
    )
    if not upstream_paths:
        available = "\n  ".join(path.name for path in files) or "无"
        raise FileNotFoundError(
            f"没有找到 上游 Excel。目标日期：{date_text}。\n"
            f"上游文件名示例：TranDetailReport_87382398_{date_text}013343.1587508.xlsx\n"
            f"或同目录 TransactionHistoryRecords.xlsx\n"
            f"当前目录文件：\n  {available}"
        )
    duplicate_paths = sorted(
        (path for path in files if match_duplicate_payment_workbook(path, date_text)),
        key=lambda path: path.name,
    )
    return directory, pair_diff_workbooks(upstream_paths, backend_paths, duplicate_paths)


def get_template_path() -> Path:
    if is_frozen_app():
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            candidate = Path(meipass) / "template" / TEMPLATE_FILE_NAME
            if candidate.exists():
                return candidate
        candidate = get_executable_directory() / "_internal" / "template" / TEMPLATE_FILE_NAME
        if candidate.exists():
            return candidate
    return DEFAULT_TEMPLATE


def get_result_dir() -> Path:
    if is_frozen_app():
        return get_executable_directory() / "result"
    return DEFAULT_RESULT_DIR


def parse_bool(value: str, option_name: str) -> bool:
    normalized = value.strip().lower()
    if normalized in {"1", "yes", "y", "true", "on"}:
        return True
    if normalized in {"0", "no", "n", "false", "off"}:
        return False
    raise ValueError(f"config.ini 中 {option_name} 必须是 true 或 false")


def get_config_path() -> Path:
    if is_frozen_app():
        candidates = (
            get_executable_directory() / CONFIG_FILE_NAME,
            Path.cwd() / CONFIG_FILE_NAME,
            PROJECT_ROOT / CONFIG_FILE_NAME,
        )
    else:
        candidates = (
            Path.cwd() / CONFIG_FILE_NAME,
            PROJECT_ROOT / CONFIG_FILE_NAME,
        )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def load_diff_config() -> configparser.SectionProxy:
    config = configparser.ConfigParser()
    config_path = get_config_path()
    if config_path.exists():
        config.read(config_path, encoding="utf-8")
    if "diff_orders" not in config:
        config["diff_orders"] = {}
    return config["diff_orders"]


def should_auto_open_html(config: configparser.SectionProxy) -> bool:
    return parse_bool(config.get("auto_open_html", "true"), "diff_orders.auto_open_html")


def choose_diff_jobs_interactive() -> list[DiffJob]:
    if not sys.stdin.isatty():
        raise RuntimeError("交互模式需要可输入的终端；或请使用 --target-date/--a/--b 参数直接运行。")

    target_date = resolve_target_date(choose_current_date())
    directory, jobs = find_diff_jobs_by_date(target_date)

    print("订单差异比对工具")
    print("----------------------------------------")
    print(f"Excel 搜索目录：{directory}")
    print(f"目标日期：{target_date:%Y-%m-%d}")
    print(f"匹配到 {len(jobs)} 组文件：")
    for index, job in enumerate(jobs, start=1):
        print(f"  {index}. 上游：{job.upstream_path.name} ({format_file_size(job.upstream_path)})")
        print(f"     后台：{job.backend_path.name} ({format_file_size(job.backend_path)})")
        if job.platform_mode:
            print(f"     算法：{job.platform_mode}")
        if job.duplicate_path:
            print(f"     重复：{job.duplicate_path.name} ({format_file_size(job.duplicate_path)})")
    print()
    return jobs


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare upstream/backend order IDs.")
    parser.add_argument("--version", action="version", version=version_text("diff-orders"))
    parser.add_argument("--a", type=Path, default=None, help="A (上游) workbook path.")
    parser.add_argument("--b", type=Path, default=None, help="B (后台) workbook path.")
    parser.add_argument("--duplicate", type=Path, default=None, help="代收重复支付订单 workbook path.")
    parser.add_argument("--a-col", default="L", help="A order ID column. Default: L")
    parser.add_argument("--b-col", default="D", help="B order ID column. Default: D")
    parser.add_argument(
        "--target-date",
        help="Target order date. Defaults to today when auto-selecting files.",
    )
    parser.add_argument("--template", type=Path, default=None, help="HTML template path.")
    parser.add_argument("--result-dir", type=Path, default=None, help="HTML result directory.")
    parser.add_argument(
        "--interactive",
        action="store_true",
        help="交互模式：输入目标日期，并自动匹配 workspace/diffOrders 里的 Excel 文件。",
    )
    args = parser.parse_args(argv)
    args.interactive = args.interactive or len(argv if argv is not None else sys.argv[1:]) == 0
    return args


def normalize_order_id(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    text = text.strip()
    if not text or text in HEADER_KEYWORDS:
        return None
    return text


def decimal_value(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def money(value: Decimal) -> str:
    quantized = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return f"{quantized:,.2f}"


def number(value: int) -> str:
    return f"{value:,}"


def extract_platform_name(job: DiffJob) -> str:
    text = f"{job.upstream_path.stem} {job.backend_path.stem} {job.upstream_path.parent.name}"
    match = re.search(r"(?i)\b[a-z]{1,6}\d{2,8}\b", text)
    if match:
        return match.group(0)
    for path in (job.backend_path, job.upstream_path):
        stem = path.stem
        cleaned = re.sub(r"\d{8,}.*$", "", stem).strip("_- ")
        cleaned = cleaned.replace(BACKEND_FILE_PREFIX, "").replace(UPSTREAM_FILE_PREFIX, "").strip("_- ")
        if cleaned:
            return cleaned
    return job.upstream_path.parent.name or "未知平台"


def platform_key_from_path(path: Path) -> str | None:
    text = f"{path.stem} {path.parent.name}"
    match = re.search(r"(?i)\b[a-z]{1,6}\d{2,8}\b", text)
    return match.group(0).lower() if match else None


def read_order_ids(path: Path, column: str) -> list[OrderEntry]:
    return read_order_entries(path, id_col=column)


def read_order_entries(
    path: Path,
    id_col: str,
    amount_col: str | None = None,
    fee_col: str | None = None,
    payment_method_col: str | None = None,
    extra_cols: dict[str, str] | None = None,
) -> list[OrderEntry]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    id_column_index = column_index_from_string(id_col)
    amount_column_index = column_index_from_string(amount_col) if amount_col else None
    fee_column_index = column_index_from_string(fee_col) if fee_col else None
    payment_method_column_index = column_index_from_string(payment_method_col) if payment_method_col else None
    extra_col_indexes = {
        key: column_index_from_string(column)
        for key, column in (extra_cols or {}).items()
    }
    max_column_index = max(
        index
        for index in (
            id_column_index,
            amount_column_index,
            fee_column_index,
            payment_method_column_index,
            *extra_col_indexes.values(),
        )
        if index
    )
    entries: list[OrderEntry] = []
    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(
            worksheet.iter_rows(min_col=1, max_col=max_column_index, values_only=True),
            start=1,
        ):
            order_id = normalize_order_id(row[id_column_index - 1])
            if order_id is None:
                continue
            amount = decimal_value(row[amount_column_index - 1]) if amount_column_index else Decimal("0")
            fee = decimal_value(row[fee_column_index - 1]) if fee_column_index else Decimal("0")
            payment_method = ""
            if payment_method_column_index:
                payment_method = str(row[payment_method_column_index - 1] or "").strip().lower()
            extra = {
                key: row[column_index - 1]
                for key, column_index in extra_col_indexes.items()
            }
            entries.append(
                OrderEntry(
                    order_id=order_id,
                    row_number=row_number,
                    amount=amount,
                    fee=fee,
                    payment_method=payment_method,
                    sheet_name=worksheet.title,
                    extra=extra,
                )
            )
    return entries


def is_finerbit_job(job: DiffJob) -> bool:
    return job.platform_mode == "finerbit"


def is_easypaisa_job(job: DiffJob) -> bool:
    return job.platform_mode in {"easypaisa", "easypasia"}


def is_duplicate_payment_mode(mode: str) -> bool:
    return mode in {"finerbit", "easypaisa"}


def special_platform_label(mode: str) -> str:
    if mode == "easypaisa":
        return "easypaisa"
    return "finerBit"


def read_job_diff_result(job: DiffJob, args: argparse.Namespace) -> JobDiffResult:
    if is_finerbit_job(job):
        a_entries = read_order_entries(
            job.upstream_path,
            id_col="B",
            amount_col="G",
            payment_method_col="D",
            extra_cols={
                "transaction_id": "A",
                "status": "K",
                "created_date": "L",
                "channel_reference_id": "N",
            },
        )
        b_entries = read_order_entries(
            job.backend_path,
            id_col="D",
            amount_col="G",
            fee_col="I",
            payment_method_col="Z",
            extra_cols={
                "platform_order_id": "A",
                "merchant_order_id": "B",
                "channel_order_id": "C",
                "completed_at": "W",
            },
        )
        duplicate_entries = (
            read_order_entries(
                job.duplicate_path,
                id_col="C",
                amount_col="K",
                fee_col="L",
                payment_method_col="I",
                extra_cols={
                    "platform_payment_order_id": "A",
                    "merchant_order_id": "B",
                    "channel_order_id": "D",
                    "merchant_name": "F",
                    "channel_cost": "N",
                    "status": "O",
                    "tid": "Q",
                    "completed_at": "R",
                },
            )
            if job.duplicate_path
            else []
        )
        result = diff_orders_with_duplicate_payments(a_entries, b_entries, duplicate_entries)
        return JobDiffResult(job=job, result=result)

    if is_easypaisa_job(job):
        a_entries = read_order_entries(
            job.upstream_path,
            id_col="I",
            amount_col="P",
            fee_col="R",
            extra_cols={
                "transaction_id": "J",
                "status": "Q",
                "created_date": "L",
                "channel_reference_id": "X",
                "fed": "S",
            },
        )
        b_entries = read_order_entries(
            job.backend_path,
            id_col="D",
            amount_col="G",
            fee_col="I",
            payment_method_col="Z",
            extra_cols={
                "platform_order_id": "A",
                "merchant_order_id": "B",
                "channel_order_id": "C",
                "completed_at": "W",
            },
        )
        duplicate_entries = (
            read_order_entries(
                job.duplicate_path,
                id_col="C",
                amount_col="K",
                fee_col="L",
                payment_method_col="I",
                extra_cols={
                    "platform_payment_order_id": "A",
                    "merchant_order_id": "B",
                    "channel_order_id": "D",
                    "merchant_name": "F",
                    "channel_cost": "N",
                    "status": "O",
                    "tid": "Q",
                    "completed_at": "R",
                },
            )
            if job.duplicate_path
            else []
        )
        result = diff_orders_with_duplicate_payments(a_entries, b_entries, duplicate_entries, special_mode="easypaisa")
        return JobDiffResult(job=job, result=result)

    a_entries = read_order_entries(job.upstream_path, id_col=args.a_col, amount_col="H")
    b_entries = read_order_entries(job.backend_path, id_col=args.b_col, amount_col="G", fee_col="I")
    return JobDiffResult(job=job, result=diff_orders(a_entries, b_entries))


def payment_rate(payment_method: str) -> Decimal:
    normalized = payment_method.strip().lower()
    if "easypaisa" in normalized or "easy paisa" in normalized:
        return Decimal("0.04")
    if "jazzcash" in normalized or "jazz cash" in normalized:
        return Decimal("0.023")
    return Decimal("0")


def calculated_fee_by_payment_method(entries: list[OrderEntry], amount_source: str) -> Decimal:
    total = Decimal("0")
    for entry in entries:
        base_amount = entry.fee if amount_source == "fee" else entry.amount
        total += base_amount * payment_rate(entry.payment_method)
    return total


def calculated_row_rounded_fee_by_payment_method(entries: list[OrderEntry], amount_source: str) -> Decimal:
    total = Decimal("0")
    for entry in entries:
        base_amount = entry.fee if amount_source == "fee" else entry.amount
        total += (base_amount * payment_rate(entry.payment_method)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return total


def calculated_easypaisa_upstream_fee(entries: list[OrderEntry]) -> Decimal:
    return sum((entry.fee + decimal_value(extra_value(entry, "fed")) for entry in entries), Decimal("0"))


def calculated_easypaisa_channel_cost(entries: list[OrderEntry]) -> Decimal:
    return sum((entry.amount * Decimal("0.02") for entry in entries), Decimal("0"))


def grouped_entries(entries: list[OrderEntry]) -> dict[str, tuple[int, Decimal]]:
    grouped: dict[str, tuple[int, Decimal]] = {}
    for entry in entries:
        count, amount = grouped.get(entry.order_id, (0, Decimal("0")))
        grouped[entry.order_id] = (count + 1, amount + entry.amount)
    return grouped


def diff_orders(a_entries: list[OrderEntry], b_entries: list[OrderEntry]) -> DiffResult:
    a_ids = {entry.order_id for entry in a_entries}
    b_ids = {entry.order_id for entry in b_entries}
    a_counter = Counter(entry.order_id for entry in a_entries)
    b_counter = Counter(entry.order_id for entry in b_entries)
    a_only_ids = a_ids - b_ids
    b_only_ids = b_ids - a_ids
    common_ids = a_ids & b_ids
    a_grouped = grouped_entries(a_entries)
    b_grouped = grouped_entries(b_entries)
    mismatched = [
        order_id
        for order_id in sorted(common_ids)
        if a_grouped[order_id][0] != b_grouped[order_id][0] or a_grouped[order_id][1] != b_grouped[order_id][1]
    ]

    return DiffResult(
        a_entries=a_entries,
        b_entries=b_entries,
        a_only=[entry for entry in a_entries if entry.order_id in a_only_ids],
        b_only=[entry for entry in b_entries if entry.order_id in b_only_ids],
        mismatched=mismatched,
        a_count=len(a_ids),
        b_count=len(b_ids),
        a_row_count=len(a_entries),
        b_row_count=len(b_entries),
        common_count=len(common_ids),
        a_duplicate_count=sum(count - 1 for count in a_counter.values() if count > 1),
        b_duplicate_count=sum(count - 1 for count in b_counter.values() if count > 1),
        a_amount=sum((entry.amount for entry in a_entries), Decimal("0")),
        b_amount=sum((entry.amount for entry in b_entries), Decimal("0")),
        a_fee=sum((entry.amount for entry in a_entries), Decimal("0")) * Decimal("0.0126"),
        b_fee=sum((entry.fee for entry in b_entries), Decimal("0")),
        a_only_amount=sum((entry.amount for entry in a_entries if entry.order_id in a_only_ids), Decimal("0")),
        b_only_amount=sum((entry.amount for entry in b_entries if entry.order_id in b_only_ids), Decimal("0")),
        mismatch_amount=sum((a_grouped[order_id][1] for order_id in mismatched), Decimal("0")),
    )


def difference_unique_entries(result: DiffResult) -> list[OrderEntry]:
    return unique_entries(result.a_only + result.b_only)


def upstream_difference_entries(result: DiffResult) -> list[OrderEntry]:
    return unique_entries(result.a_only)


def diff_orders_with_duplicate_payments(
    a_entries: list[OrderEntry],
    b_entries: list[OrderEntry],
    duplicate_entries: list[OrderEntry],
    special_mode: str = "finerbit",
) -> DiffResult:
    base_result = diff_orders(a_entries, b_entries)
    difference_entries = upstream_difference_entries(base_result)
    duplicate_ids = {entry.order_id for entry in duplicate_entries}
    repeated_entries = tuple(entry for entry in difference_entries if entry.order_id in duplicate_ids)
    remaining_entries = tuple(entry for entry in difference_entries if entry.order_id not in duplicate_ids)
    if special_mode == "easypaisa":
        a_fee = calculated_easypaisa_upstream_fee(a_entries)
        channel_cost = calculated_easypaisa_channel_cost(b_entries)
    else:
        a_fee = calculated_fee_by_payment_method(a_entries, "amount")
        channel_cost = calculated_row_rounded_fee_by_payment_method(a_entries, "amount")
    return DiffResult(
        a_entries=base_result.a_entries,
        b_entries=base_result.b_entries,
        a_only=base_result.a_only,
        b_only=base_result.b_only,
        mismatched=base_result.mismatched,
        a_count=base_result.a_count,
        b_count=base_result.b_count,
        a_row_count=base_result.a_row_count,
        b_row_count=base_result.b_row_count,
        common_count=base_result.common_count,
        a_duplicate_count=base_result.a_duplicate_count,
        b_duplicate_count=base_result.b_duplicate_count,
        a_amount=base_result.a_amount,
        b_amount=base_result.b_amount,
        a_fee=a_fee,
        b_fee=base_result.b_fee,
        a_only_amount=base_result.a_only_amount,
        b_only_amount=base_result.b_only_amount,
        mismatch_amount=base_result.mismatch_amount,
        duplicate_payment_entries=tuple(unique_entries(duplicate_entries)),
        repeated_difference_entries=repeated_entries,
        remaining_difference_entries=remaining_entries,
        channel_cost=channel_cost,
        special_mode=special_mode,
    )


def unique_entries(entries: list[OrderEntry]) -> list[OrderEntry]:
    seen: set[str] = set()
    unique: list[OrderEntry] = []
    for entry in entries:
        if entry.order_id in seen:
            continue
        seen.add(entry.order_id)
        unique.append(entry)
    return unique


def unique_order_ids(entries: list[OrderEntry]) -> list[str]:
    return [entry.order_id for entry in unique_entries(entries)]


def render_upstream_order_table(entries: list[OrderEntry]) -> str:
    if not entries:
        return '<div class="empty">无</div>'

    columns = 4
    rows = [
        "<table>",
        "<thead><tr>" + "".join("<th>上游独有订单ID</th>" for _ in range(columns)) + "</tr></thead>",
        "<tbody>",
    ]
    for start in range(0, len(entries), columns):
        cells = []
        for entry in entries[start : start + columns]:
            order_id = escape(entry.order_id, quote=True)
            cells.append(
                "<td>"
                '<label class="order-check">'
                f'<input type="checkbox" data-check-order-id="{order_id}">'
                f'<button class="copy-order" data-order-id="{order_id}" title="点击复制订单号">{order_id}</button>'
                "</label>"
                "</td>"
            )
        cells.extend("<td></td>" for _ in range(columns - len(cells)))
        rows.append("<tr>" + "".join(cells) + "</tr>")
    rows.extend(["</tbody>", "</table>"])
    return "\n".join(rows)


def summary_metrics(job_result: JobDiffResult) -> list[SummaryMetric]:
    result = job_result.result
    if is_duplicate_payment_mode(result.special_mode):
        upstream_note = (
            "TransactionHistoryRecords I列 Order ID"
            if result.special_mode == "easypaisa"
            else "Transaction Details B列 ReferenceId"
        )
        return [
            SummaryMetric("我方TP订单", result.b_row_count, result.b_amount, result.b_fee, result.b_count, "D列 请求上游订单号", unique_order_ids(result.b_entries)),
            SummaryMetric("上游订单", result.a_row_count, result.a_amount, result.a_fee, result.a_count, upstream_note, unique_order_ids(result.a_entries)),
            SummaryMetric("双方共有", result.common_count, Decimal("0"), None, result.common_count, "两边都存在的订单号", []),
            SummaryMetric("上游独有", len(unique_entries(result.a_only)), result.a_only_amount, None, len(unique_entries(result.a_only)), "上游有，我方TP没有", unique_order_ids(result.a_only)),
            SummaryMetric("我方独有", len(unique_entries(result.b_only)), result.b_only_amount, None, len(unique_entries(result.b_only)), "我方TP有，上游没有", unique_order_ids(result.b_only)),
            SummaryMetric("代收重复订单", len(result.duplicate_payment_entries), sum((entry.amount for entry in result.duplicate_payment_entries), Decimal("0")), None, len(result.duplicate_payment_entries), "代收重复支付订单 C列 请求上游ID", unique_order_ids(list(result.duplicate_payment_entries))),
            SummaryMetric("重复订单", len(result.repeated_difference_entries), sum((entry.amount for entry in result.repeated_difference_entries), Decimal("0")), None, len(result.repeated_difference_entries), "上游独有订单中命中代收重复订单", unique_order_ids(list(result.repeated_difference_entries))),
            SummaryMetric("仍存在的差异", len(result.remaining_difference_entries), sum((entry.amount for entry in result.remaining_difference_entries), Decimal("0")), None, len(result.remaining_difference_entries), "完成时间/投诉退款/重复支付后仍需处理", unique_order_ids(list(result.remaining_difference_entries))),
        ]
    return [
        SummaryMetric("上游订单", result.a_row_count, result.a_amount, result.a_fee, result.a_count, "TranDetailReport", unique_order_ids(result.a_entries)),
        SummaryMetric("我方TP订单", result.b_row_count, result.b_amount, result.b_fee, result.b_count, "收款订单；手续费取 I 列", unique_order_ids(result.b_entries)),
        SummaryMetric("上游独有订单", len(unique_entries(result.a_only)), result.a_only_amount, None, len(unique_entries(result.a_only)), "上游有，TP没有", unique_order_ids(result.a_only)),
        SummaryMetric("TP独有订单", len(unique_entries(result.b_only)), result.b_only_amount, None, len(unique_entries(result.b_only)), "TP有，上游没有", unique_order_ids(result.b_only)),
        SummaryMetric("金额/笔数不一致", len(result.mismatched), result.mismatch_amount, None, len(result.mismatched), "双方都有但金额或笔数不同", result.mismatched),
    ]


def render_metric_rows(metrics: list[SummaryMetric], section_key: str) -> tuple[str, dict[str, list[str]]]:
    rows: list[str] = []
    copy_payloads: dict[str, list[str]] = {}
    for index, metric in enumerate(metrics):
        fee_text = "" if metric.fee is None else money(metric.fee)
        copy_key = f"{section_key}-{index}"
        copy_payloads[copy_key] = metric.order_ids
        disabled = " disabled" if not metric.order_ids else ""
        button_text = "复制ID" if metric.order_ids else "无ID"
        rows.append(
            "<tr>"
            f"<td>{escape(metric.label)}</td>"
            f"<td class=\"num\">{number(metric.count)}</td>"
            f"<td class=\"num\">{money(metric.amount)}</td>"
            f"<td class=\"num\">{fee_text}</td>"
            f"<td class=\"num\">{number(metric.unique_count)}</td>"
            f"<td>{escape(metric.note)}</td>"
            f"<td><button class=\"copy-ids\" type=\"button\" data-copy-key=\"{escape(copy_key, quote=True)}\" data-copy-label=\"{escape(metric.label, quote=True)}\"{disabled}>{button_text}</button></td>"
            "</tr>"
        )
    return "".join(rows), copy_payloads


def text_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def extra_value(entry: OrderEntry, key: str) -> Any:
    if not entry.extra:
        return None
    return entry.extra.get(key)


def row_fee(entry: OrderEntry) -> Decimal:
    return (entry.amount * payment_rate(entry.payment_method)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def upstream_entry_fee(entry: OrderEntry, mode: str) -> Decimal:
    if mode == "easypaisa":
        return entry.fee + decimal_value(extra_value(entry, "fed"))
    return row_fee(entry)


def channel_cost_entry_fee(entry: OrderEntry, mode: str) -> Decimal:
    if mode == "easypaisa":
        return entry.amount * Decimal("0.02")
    return row_fee(entry)


def money_cell(value: Decimal | Any) -> str:
    if isinstance(value, Decimal):
        return money(value)
    return escape(text_value(value))


def render_detail_table(headers: list[str], rows: list[list[tuple[Any, bool]]]) -> str:
    if not rows:
        return '<div class="empty-detail">无明细</div>'
    header_html = "".join(f"<th>{escape(header)}</th>" for header in headers)
    body_rows: list[str] = []
    for row in rows:
        cells = []
        for value, is_number in row:
            class_name = ' class="num"' if is_number else ""
            cells.append(f"<td{class_name}>{money_cell(value)}</td>")
        body_rows.append("<tr>" + "".join(cells) + "</tr>")
    return (
        '<div class="detail-table-wrap">'
        '<table class="detail-table">'
        f"<thead><tr>{header_html}</tr></thead>"
        f"<tbody>{''.join(body_rows)}</tbody>"
        "</table>"
        "</div>"
    )


def render_detail_panel(
    section_key: str,
    panel_key: str,
    title: str,
    headers: list[str],
    rows: list[list[tuple[Any, bool]]],
    order_ids: list[str],
) -> tuple[str, dict[str, list[str]]]:
    copy_key = f"{section_key}-{panel_key}"
    disabled = " disabled" if not order_ids else ""
    html = f"""
      <div class="detail-toolbar">
        <strong>{escape(title)}</strong>
        <button class="copy-ids detail-copy" type="button" data-copy-key="{escape(copy_key, quote=True)}" data-copy-label="{escape(title, quote=True)}"{disabled}>复制本页订单ID</button>
      </div>
      {render_detail_table(headers, rows)}
    """
    return html, {copy_key: order_ids}


def tp_detail_rows(entries: list[OrderEntry], mode: str) -> list[list[tuple[Any, bool]]]:
    rows: list[list[tuple[Any, bool]]] = []
    for entry in sorted(unique_entries(entries), key=lambda item: item.order_id):
        rows.append(
            [
                ("我方独有", False),
                (entry.sheet_name, False),
                (entry.row_number, True),
                (entry.order_id, False),
                (entry.amount, True),
                (entry.fee, True),
                (channel_cost_entry_fee(entry, mode), True),
                (entry.payment_method, False),
                (extra_value(entry, "platform_order_id"), False),
                (extra_value(entry, "merchant_order_id"), False),
                (extra_value(entry, "channel_order_id"), False),
                (extra_value(entry, "completed_at"), False),
            ]
        )
    return rows


def upstream_detail_rows(entries: list[OrderEntry], source: str, mode: str, note: str = "") -> list[list[tuple[Any, bool]]]:
    rows: list[list[tuple[Any, bool]]] = []
    for entry in sorted(unique_entries(entries), key=lambda item: item.order_id):
        row = [
            (source, False),
            (entry.row_number, True),
            (entry.order_id, False),
            (entry.amount, True),
            (upstream_entry_fee(entry, mode), True),
            (entry.payment_method, False),
            (extra_value(entry, "transaction_id"), False),
            (extra_value(entry, "status"), False),
        ]
        if note:
            row.append((note, False))
        else:
            row.extend(
                [
                    (extra_value(entry, "created_date"), False),
                    (extra_value(entry, "channel_reference_id"), False),
                ]
            )
        rows.append(row)
    return rows


def final_difference_rows(entries: tuple[OrderEntry, ...], mode: str) -> list[list[tuple[Any, bool]]]:
    rows: list[list[tuple[Any, bool]]] = []
    for entry in sorted(unique_entries(list(entries)), key=lambda item: item.order_id):
        rows.append(
            [
                ("上游独有", False),
                (entry.row_number, True),
                (entry.order_id, False),
                (extra_value(entry, "channel_reference_id"), False),
                (entry.amount, True),
                (upstream_entry_fee(entry, mode), True),
                (entry.payment_method, False),
                (extra_value(entry, "status"), False),
                ("除重复支付之外的差异", False),
            ]
        )
    return rows


def duplicate_detail_rows(result: DiffResult) -> list[list[tuple[Any, bool]]]:
    repeated_ids = {entry.order_id for entry in result.repeated_difference_entries}
    rows: list[list[tuple[Any, bool]]] = []
    for entry in sorted(
        (entry for entry in unique_entries(list(result.duplicate_payment_entries)) if entry.order_id in repeated_ids),
        key=lambda item: item.order_id,
    ):
        rows.append(
            [
                (entry.row_number, True),
                (entry.order_id, False),
                (entry.amount, True),
                (entry.fee, True),
                (extra_value(entry, "channel_cost"), True),
                (entry.payment_method, False),
                (extra_value(entry, "platform_payment_order_id"), False),
                (extra_value(entry, "merchant_order_id"), False),
                (extra_value(entry, "channel_order_id"), False),
                (extra_value(entry, "merchant_name"), False),
                (extra_value(entry, "status"), False),
                (extra_value(entry, "completed_at"), False),
            ]
        )
    return rows


def render_finerbit_tabs(summary_html: str, result: DiffResult, section_key: str) -> tuple[str, dict[str, list[str]]]:
    copy_payloads: dict[str, list[str]] = {}
    panels: list[tuple[str, str, str]] = [("summary", "汇总", summary_html)]

    detail_specs = [
        (
            "tp-only",
            "我方独有",
            ["来源", "工作表", "源表行号", "订单号", "金额(PKR)", "手续费(PKR)", "渠道成本(PKR)", "收款方式名称", "平台订单号", "商户订单号", "渠道订单号", "完成时间"],
            tp_detail_rows(result.b_only, result.special_mode),
            unique_order_ids(result.b_only),
        ),
        (
            "upstream-only",
            "上游独有",
            ["来源", "源表行号", "订单号", "金额(PKR)", "手续费(PKR)", "支付方式", "TransactionId", "状态", "Created Date", "ChannelReferenceId"],
            upstream_detail_rows(result.a_only, "上游独有", result.special_mode),
            unique_order_ids(result.a_only),
        ),
        (
            "repeated",
            "重复订单",
            ["来源", "源表行号", "订单号", "金额(PKR)", "手续费/成本(PKR)", "支付方式", "TransactionId", "状态", "说明"],
            upstream_detail_rows(list(result.repeated_difference_entries), "上游独有", result.special_mode, "差异订单命中代收重复订单"),
            unique_order_ids(list(result.repeated_difference_entries)),
        ),
        (
            "repeated-detail",
            "重复订单明细",
            ["源表行号", "请求上游ID", "订单金额(PKR)", "商户手续费", "渠道成本", "支付方式", "平台支付订单号", "商户订单号", "渠道订单号", "商户名称", "状态", "完成时间"],
            duplicate_detail_rows(result),
            unique_order_ids(list(result.repeated_difference_entries)),
        ),
        (
            "remaining",
            "最终仍存在差异",
            ["来源", "源表行号", "订单号", "tid", "金额(PKR)", "手续费/成本(PKR)", "支付方式", "状态", "后续处理"],
            final_difference_rows(result.remaining_difference_entries, result.special_mode),
            unique_order_ids(list(result.remaining_difference_entries)),
        ),
    ]

    for key, title, headers, rows, order_ids in detail_specs:
        panel_html, panel_payloads = render_detail_panel(section_key, key, title, headers, rows, order_ids)
        panels.append((key, title, panel_html))
        copy_payloads.update(panel_payloads)

    tab_buttons = []
    panel_html_parts = []
    for index, (key, title, panel_html) in enumerate(panels):
        active_class = " active" if index == 0 else ""
        hidden = "" if index == 0 else " hidden"
        panel_id = f"{section_key}-{key}"
        tab_buttons.append(
            f'<button class="sheet-tab{active_class}" type="button" data-tab-target="{escape(panel_id, quote=True)}">{escape(title)}</button>'
        )
        panel_html_parts.append(
            f'<div class="sheet-panel{active_class}" id="{escape(panel_id, quote=True)}"{hidden}>{panel_html}</div>'
        )

    return (
        '<div class="sheet-tabs">'
        f'<div class="sheet-tab-list">{"".join(tab_buttons)}</div>'
        f'{"".join(panel_html_parts)}'
        "</div>",
        copy_payloads,
    )


def display_order_date(results: list[JobDiffResult]) -> str:
    for job_result in results:
        for path in (job_result.job.upstream_path, job_result.job.backend_path):
            match = re.search(r"(20\d{2})(\d{2})(\d{2})", path.stem)
            if match:
                return f"{int(match.group(2))}月{int(match.group(3))}日"
    return "所选日期"


def render_overview(results: list[JobDiffResult]) -> str:
    order_date = display_order_date(results)
    rows: list[str] = []
    for index, job_result in enumerate(results, start=1):
        platform = extract_platform_name(job_result.job)
        rows.append(
            "<tr>"
            f"<td class=\"num\">{index}</td>"
            f"<td><a class=\"detail-link\" href=\"#result-section-{index}\">{escape(platform)}</a></td>"
            f"<td>{escape(job_result.job.upstream_path.name)}</td>"
            f"<td>{escape(job_result.job.backend_path.name)}</td>"
            f"<td><a class=\"detail-link\" href=\"#result-section-{index}\">查看详情</a></td>"
            "</tr>"
        )

    return f"""
    <section class="overview-section">
      <h1>订单差异对账总览</h1>
      <p class="formula">共有 {number(len(results))} 组 {escape(order_date)} 的订单对比</p>
      <table class="overview-table">
        <thead>
          <tr><th>序号</th><th>平台</th><th>上游文件</th><th>TP文件</th><th>定位</th></tr>
        </thead>
        <tbody>{''.join(rows)}</tbody>
      </table>
    </section>
    """


def render_result_section(job_result: JobDiffResult, index: int, total: int) -> tuple[str, dict[str, list[str]]]:
    result = job_result.result
    platform = extract_platform_name(job_result.job)
    metrics = summary_metrics(job_result)
    metric_rows, copy_payloads = render_metric_rows(metrics, f"g{index}")
    title_prefix = f"第 {index} 组：" if total > 1 else ""
    is_special_mode = is_duplicate_payment_mode(result.special_mode)
    special_label = special_platform_label(result.special_mode)
    profit = result.b_fee - result.a_fee if is_special_mode else result.b_fee - result.a_fee - result.channel_cost
    total_summary = (
        f"上游合计 {number(result.a_row_count)} 笔 / {money(result.a_amount)} PKR；"
        f"TP合计 {number(result.b_row_count)} 笔 / {money(result.b_amount)} PKR"
    )
    if is_special_mode:
        difference_count = len(upstream_difference_entries(result))
        conclusion = (
            f"差异订单 {number(difference_count)} 个；"
            f"其中 {number(len(result.repeated_difference_entries))} 个为重复订单，"
            f"{number(len(result.remaining_difference_entries))} 个仍存在差异。"
        )
        upstream_order_note = (
            "TransactionHistoryRecords I列 Order ID"
            if result.special_mode == "easypaisa"
            else "Transaction Details B列 ReferenceId"
        )
        formula_text = (
            f"我方TP订单 D列 请求上游订单号 vs 上游订单 {upstream_order_note}；"
            "差异订单号 vs 代收重复支付订单 C列 请求上游ID"
        )
    else:
        conclusion = (
            f"{platform}：上游独有 {len(unique_entries(result.a_only))} 单；"
            f"TP独有 {len(unique_entries(result.b_only))} 单；"
            f"金额/笔数不一致 {len(result.mismatched)} 单。"
        )
        formula_text = total_summary
    extra_profit_label = "-渠道成本" if is_special_mode else ""
    duplicate_file_note = (
        f"<br>代收重复：{escape(job_result.job.duplicate_path.name)}" if job_result.job.duplicate_path else ""
    )
    profit_tp_amount = result.a_amount if is_special_mode else result.b_amount
    upstream_platform_label = special_label if is_special_mode else platform
    tp_platform_label = "tarspay" if is_special_mode else "TP"
    tp_success_amount_label = (
        f"{tp_platform_label}平台（{special_label}交易成功金额）"
        if is_special_mode
        else f"TP平台（{platform} 交易成功金额）"
    )
    tp_success_count_label = (
        f"{tp_platform_label}平台（{special_label}成功笔数）"
        if is_special_mode
        else f"TP平台（{platform} 成功笔数）"
    )
    channel_cost_header = (
        f'<th class="tp-head">{escape(tp_platform_label)}平台（{escape(special_label)}渠道成本）</th>'
        if is_special_mode
        else ""
    )
    channel_cost_cell = (
        f'<td class="num">{money(result.channel_cost)}</td>' if is_special_mode else ""
    )
    profit_header = (
        f"{escape(tp_platform_label)}平台利润（{escape(special_label)}）"
        if is_special_mode
        else f"TP平台利润（TP手续费-上游手续费{extra_profit_label}）"
    )

    summary_html = f"""
      <table class="summary-table">
        <thead>
          <tr><th>项目</th><th>订单笔数/行数</th><th>订单金额(PKR)</th><th>手续费/成本(PKR)</th><th>唯一订单号数</th><th>备注</th><th>操作</th></tr>
        </thead>
        <tbody>{metric_rows}</tbody>
      </table>
      <div class="conclusion-row">
        <div class="conclusion-label">结论</div>
        <div class="conclusion-text">{escape(conclusion)}</div>
      </div>
      <table class="profit-table">
        <thead>
          <tr>
            <th class="do-head">{escape(upstream_platform_label)}平台（交易金额）</th>
            <th class="tp-head">{escape(tp_success_amount_label)}</th>
            <th class="do-head">{escape(upstream_platform_label)}平台（成功笔数）</th>
            <th class="tp-head">{escape(tp_success_count_label)}</th>
            <th class="do-head">{escape(upstream_platform_label)}平台（手续费）</th>
            {channel_cost_header}
            <th class="tp-head">{escape(tp_platform_label)}平台（手续费）</th>
            <th class="profit-head">{profit_header}</th>
          </tr>
        </thead>
        <tbody>
          <tr>
            <td class="num">{money(result.a_amount)}</td>
            <td class="num">{money(profit_tp_amount)}</td>
            <td class="num">{number(result.a_row_count)}</td>
            <td class="num">{number(result.b_row_count)}</td>
            <td class="num">{money(result.a_fee)}</td>
            {channel_cost_cell}
            <td class="num">{money(result.b_fee)}</td>
            <td class="num">{money(profit)}</td>
          </tr>
        </tbody>
      </table>
      <div class="file-note">上游：{escape(job_result.job.upstream_path.name)}<br>TP：{escape(job_result.job.backend_path.name)}{duplicate_file_note}</div>
    """
    body_html = summary_html
    if is_special_mode:
        body_html, detail_payloads = render_finerbit_tabs(summary_html, result, f"g{index}-detail")
        copy_payloads.update(detail_payloads)

    html = f"""
    <section class="result-section" id="result-section-{index}">
      <h1>{title_prefix}订单差异对账结果（上游 vs TP） - {escape(platform)}</h1>
      <p class="formula">{escape(formula_text)}</p>
      <p class="formula">{escape(total_summary)}</p>
      {body_html}
    </section>
    """
    return html, copy_payloads


def render_stats_html(results: list[JobDiffResult]) -> str:
    title = "订单差异对账结果" if len(results) == 1 else "订单差异批量对账结果"
    overview = render_overview(results)
    section_html: list[str] = []
    copy_payloads: dict[str, list[str]] = {}
    for index, job_result in enumerate(results, start=1):
        section, section_payloads = render_result_section(job_result, index, len(results))
        section_html.append(section)
        copy_payloads.update(section_payloads)
    sections = "".join(section_html)
    copy_payload_json = json.dumps(copy_payloads, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{title}</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif;
      color: #111827;
      background: #fff;
    }}
    main {{
      max-width: 1680px;
      margin: 0 auto;
      padding: 18px 20px 28px;
    }}
    .overview-section {{
      margin-bottom: 36px;
      padding-bottom: 20px;
      border-bottom: 2px solid #d9e2ec;
    }}
    .result-section {{
      margin-bottom: 48px;
      scroll-margin-top: 18px;
    }}
    h1 {{
      margin: 0 0 6px;
      color: #17365d;
      font-size: 26px;
      line-height: 1.25;
      font-weight: 700;
    }}
    .formula {{
      margin: 0 0 18px;
      font-size: 15px;
      color: #1f2937;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
    }}
    th {{
      border: 1px solid #cfd8e3;
      padding: 3px 6px;
      color: #fff;
      background: #1f4e79;
      text-align: left;
      font-size: 17px;
      line-height: 1.25;
      font-weight: 700;
    }}
    td {{
      border: 1px solid #d9e2ec;
      padding: 3px 6px;
      font-size: 16px;
      line-height: 1.25;
      vertical-align: middle;
      word-break: break-all;
    }}
    a.detail-link {{
      color: #1f4e79;
      font-weight: 700;
      text-decoration: none;
    }}
    a.detail-link:hover {{
      text-decoration: underline;
    }}
    .summary-table {{ margin-bottom: 22px; }}
    .overview-table {{ margin-bottom: 6px; }}
    .overview-table th:nth-child(1) {{ width: 7%; }}
    .overview-table th:nth-child(2) {{ width: 13%; }}
    .overview-table th:nth-child(3) {{ width: 36%; }}
    .overview-table th:nth-child(4) {{ width: 36%; }}
    .overview-table th:nth-child(5) {{ width: 8%; }}
    .summary-table th:nth-child(1) {{ width: 15%; }}
    .summary-table th:nth-child(2) {{ width: 12%; }}
    .summary-table th:nth-child(3) {{ width: 15%; }}
    .summary-table th:nth-child(4) {{ width: 15%; }}
    .summary-table th:nth-child(5) {{ width: 13%; }}
    .summary-table th:nth-child(6) {{ width: 22%; }}
    .summary-table th:nth-child(7) {{ width: 8%; }}
    .num {{ text-align: right; }}
    .copy-ids {{
      min-height: 28px;
      width: 100%;
      border: 1px solid #1f4e79;
      border-radius: 4px;
      color: #1f4e79;
      background: #fff;
      font: inherit;
      font-size: 14px;
      cursor: pointer;
    }}
    .copy-ids:hover:not(:disabled) {{
      background: #eaf2f8;
    }}
    .copy-ids:disabled {{
      border-color: #cbd5e1;
      color: #94a3b8;
      cursor: default;
    }}
    .conclusion-row {{
      display: grid;
      grid-template-columns: 220px 1fr;
      gap: 0;
      margin: 0 0 80px;
      font-size: 17px;
    }}
    .conclusion-label {{ padding: 4px 0; }}
    .conclusion-text {{
      padding: 4px 8px;
      color: #8b2f1d;
      background: #fbe2d5;
    }}
    .profit-table th {{
      color: #001f3f;
      background: #fbe4d5;
      font-style: italic;
    }}
    .profit-table .tp-head {{ background: #e2f0d9; }}
    .profit-table .profit-head {{ background: #ffe699; }}
    .file-note {{
      margin-top: 8px;
      color: #6b7280;
      font-size: 12px;
      line-height: 1.4;
    }}
    .sheet-tabs {{
      margin-top: 18px;
    }}
    .sheet-tab-list {{
      display: flex;
      gap: 4px;
      overflow-x: auto;
      border-bottom: 2px solid #1f4e79;
      margin-bottom: 14px;
    }}
    .sheet-tab {{
      flex: 0 0 auto;
      min-height: 34px;
      border: 1px solid #cfd8e3;
      border-bottom: 0;
      padding: 6px 14px;
      color: #1f4e79;
      background: #f8fafc;
      font: inherit;
      font-size: 15px;
      cursor: pointer;
    }}
    .sheet-tab.active {{
      color: #fff;
      background: #1f4e79;
      border-color: #1f4e79;
      font-weight: 700;
    }}
    .sheet-panel[hidden] {{
      display: none;
    }}
    .detail-toolbar {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      margin: 0 0 8px;
      color: #17365d;
      font-size: 16px;
    }}
    .detail-copy {{
      width: auto;
      min-width: 150px;
      padding: 2px 12px;
    }}
    .detail-table-wrap {{
      overflow-x: auto;
      border: 1px solid #d9e2ec;
    }}
    .detail-table {{
      min-width: 1280px;
      table-layout: auto;
      border: 0;
    }}
    .detail-table th,
    .detail-table td {{
      white-space: nowrap;
      font-size: 13px;
    }}
    .empty-detail {{
      padding: 18px;
      border: 1px solid #d9e2ec;
      color: #6b7280;
      background: #f8fafc;
    }}
    .toast {{
      position: fixed;
      left: 50%;
      bottom: 28px;
      transform: translateX(-50%);
      min-width: 220px;
      max-width: calc(100vw - 48px);
      padding: 12px 18px;
      border-radius: 8px;
      background: #111827;
      color: #fff;
      text-align: center;
      opacity: 0;
      pointer-events: none;
      transition: opacity 0.18s ease, transform 0.18s ease;
      box-shadow: 0 10px 25px rgba(17, 24, 39, 0.18);
    }}
    .toast.show {{
      opacity: 1;
      transform: translate(-50%, -6px);
    }}
  </style>
</head>
<body>
  <main>
    {overview}
    {sections}
  </main>
  <div id="toast" class="toast">已复制订单ID</div>
  <script>
    const COPY_PAYLOADS = {copy_payload_json};
    const toast = document.getElementById('toast');
    let toastTimer = null;

    function showMessage(message) {{
      toast.textContent = message;
      toast.classList.add('show');
      window.clearTimeout(toastTimer);
      toastTimer = window.setTimeout(() => toast.classList.remove('show'), 1600);
    }}

    async function copyText(text) {{
      try {{
        await navigator.clipboard.writeText(text);
      }} catch (error) {{
        const input = document.createElement('textarea');
        input.value = text;
        document.body.appendChild(input);
        input.select();
        document.execCommand('copy');
        input.remove();
      }}
    }}

    document.addEventListener('click', async (event) => {{
      const tab = event.target.closest('[data-tab-target]');
      if (tab) {{
        const tabs = tab.closest('.sheet-tabs');
        if (!tabs) return;
        tabs.querySelectorAll('.sheet-tab').forEach((item) => item.classList.remove('active'));
        tabs.querySelectorAll('.sheet-panel').forEach((panel) => {{
          panel.classList.remove('active');
          panel.hidden = true;
        }});
        const panel = document.getElementById(tab.dataset.tabTarget);
        tab.classList.add('active');
        if (panel) {{
          panel.hidden = false;
          panel.classList.add('active');
        }}
        return;
      }}
      const button = event.target.closest('[data-copy-key]');
      if (!button || button.disabled) return;
      const ids = COPY_PAYLOADS[button.dataset.copyKey] || [];
      if (ids.length === 0) {{
        showMessage('没有可复制的订单ID');
        return;
      }}
      await copyText(ids.join('\\n'));
      showMessage(`已复制 ${{ids.length}} 个订单ID`);
    }});
  </script>
</body>
</html>
"""


def render_html(template_path: Path, result: DiffResult) -> str:
    placeholder_job = DiffJob(upstream_path=Path("TranDetailReport.xlsx"), backend_path=Path("收款订单.xlsx"))
    return render_stats_html([JobDiffResult(job=placeholder_job, result=result)])


def render_batch_html(results: list[JobDiffResult]) -> str:
    return render_stats_html(results)


def sanitize_filename_part(value: str) -> str:
    sanitized = re.sub(r"[^\w\u4e00-\u9fff.-]+", "_", value.strip(), flags=re.UNICODE).strip("._-")
    return sanitized or "unknown"


def comparison_folder_name(job: DiffJob) -> str:
    upstream_parent = job.upstream_path.parent.name
    backend_parent = job.backend_path.parent.name
    if upstream_parent == backend_parent:
        return upstream_parent
    return f"{upstream_parent}_{backend_parent}"


def result_filename_group_part(results: list[JobDiffResult]) -> str:
    folder_names: list[str] = []
    seen: set[str] = set()
    for job_result in results:
        folder_name = sanitize_filename_part(comparison_folder_name(job_result.job))
        if folder_name in seen:
            continue
        seen.add(folder_name)
        folder_names.append(folder_name)
    group_part = "_".join(folder_names)
    return group_part[:120].strip("._-") or "unknown"


def write_html_result(result_dir: Path, template_path: Path, job_result: JobDiffResult) -> Path:
    result_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    group_part = result_filename_group_part([job_result])
    output_path = result_dir / f"order_diff_{group_part}_{timestamp}.html"
    output_path.write_text(render_stats_html([job_result]), encoding="utf-8")
    return output_path


def write_batch_html_result(result_dir: Path, results: list[JobDiffResult]) -> Path:
    result_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    group_part = result_filename_group_part(results)
    output_path = result_dir / f"order_diff_batch_{group_part}_{timestamp}.html"
    output_path.write_text(render_batch_html(results), encoding="utf-8")
    return output_path


def open_file(path: Path) -> bool:
    try:
        if sys.platform == "darwin":
            subprocess.run(["open", str(path)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        elif sys.platform == "win32":
            import os

            os.startfile(path)  # type: ignore[attr-defined]
        else:
            subprocess.run(["xdg-open", str(path)], check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return True
    except (OSError, subprocess.SubprocessError):
        return False


def maybe_open_html(path: Path, auto_open_html: bool, label: str) -> None:
    if not auto_open_html:
        print(f"已按 config.ini 配置关闭自动打开 HTML，请手动打开{label}查看")
        return
    if open_file(path):
        print(f"已自动打开 HTML {label}")
    else:
        print(f"自动打开 HTML 失败，请手动打开{label}查看")


def resolve_diff_jobs(args: argparse.Namespace) -> list[DiffJob]:
    if args.a is not None and args.b is not None:
        return [make_diff_job(args.a, args.b, [args.duplicate] if args.duplicate else [])]

    target_date = resolve_target_date(args.target_date)
    directory, jobs = find_diff_jobs_by_date(target_date)

    print("自动匹配订单 Excel")
    print("----------------------------------------")
    print(f"Excel 搜索目录：{directory}")
    print(f"目标日期：{target_date:%Y-%m-%d}")
    print(f"匹配到 {len(jobs)} 组文件：")
    for index, job in enumerate(jobs, start=1):
        print(f"  {index}. 上游：{job.upstream_path.name} ({format_file_size(job.upstream_path)})")
        print(f"     后台：{job.backend_path.name} ({format_file_size(job.backend_path)})")
        if job.platform_mode:
            print(f"     算法：{job.platform_mode}")
        if job.duplicate_path:
            print(f"     重复：{job.duplicate_path.name} ({format_file_size(job.duplicate_path)})")
    print()
    return jobs


def should_pause_before_exit() -> bool:
    return is_frozen_app() and sys.stdin.isatty()


def pause_before_exit() -> None:
    if not should_pause_before_exit():
        return
    try:
        input("\n按回车退出...")
    except EOFError:
        pass


def write_error_log(error: BaseException) -> Path:
    log_path = get_executable_directory() / "diff-orders-error.log"
    log_text = "".join(traceback.format_exception(type(error), error, error.__traceback__))
    log_path.write_text(log_text, encoding="utf-8")
    return log_path


def main() -> None:
    args = parse_args()
    diff_config = load_diff_config()
    auto_open_html = should_auto_open_html(diff_config)
    template_path = args.template or get_template_path()
    result_dir = args.result_dir or get_result_dir()

    if args.interactive:
        jobs = choose_diff_jobs_interactive()
    else:
        jobs = resolve_diff_jobs(args)

    print("正在读取并比对订单，请稍候...")
    job_results: list[JobDiffResult] = []
    for index, job in enumerate(jobs, start=1):
        if len(jobs) > 1:
            print(f"  正在处理第 {index}/{len(jobs)} 组：{job.upstream_path.name} / {job.backend_path.name}")
        job_results.append(read_job_diff_result(job, args))

    if len(job_results) > 1:
        html_path = write_batch_html_result(result_dir, job_results)
        total_upstream_only = sum(len(unique_entries(job_result.result.a_only)) for job_result in job_results)

        print()
        print("订单批量差异比对完成")
        print("----------------------------------------")
        print(f"对比组数: {len(job_results)}")
        print(f"上游独有订单总数: {total_upstream_only}")
        print(f"HTML汇总文件: {html_path}")
        maybe_open_html(html_path, auto_open_html, "汇总文件")
        return

    job_result = job_results[0]
    result = job_result.result
    html_path = write_html_result(result_dir, template_path, job_result)
    upstream_only = unique_entries(result.a_only)

    print()
    print("订单差异比对完成")
    print("----------------------------------------")
    print(f"上游 Excel: {job_result.job.upstream_path} ({result.a_count} unique)")
    print(f"后台 Excel: {job_result.job.backend_path} ({result.b_count} unique)")
    print(f"上游独有订单数: {len(upstream_only)}")
    print(f"HTML结果文件: {html_path}")
    maybe_open_html(html_path, auto_open_html, "结果文件")


if __name__ == "__main__":
    exit_code = 0
    try:
        main()
    except KeyboardInterrupt:
        print("\n已取消。")
        exit_code = 130
    except (FileNotFoundError, RuntimeError, ValueError) as error:
        print(f"错误：{error}")
        exit_code = 1
    except Exception as error:
        try:
            log_path = write_error_log(error)
            print(f"程序执行失败：{error}")
            print(f"详细错误已写入：{log_path}")
        except Exception:
            print("程序执行失败，并且写入错误日志时也失败：")
            traceback.print_exc()
        exit_code = 1
    finally:
        pause_before_exit()
    sys.exit(exit_code)
