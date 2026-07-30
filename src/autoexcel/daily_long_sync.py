from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
import os
from pathlib import Path
import re
import stat
import tempfile
import xml.etree.ElementTree as ET
import zipfile

from openpyxl.utils.cell import get_column_letter, range_boundaries

from autoexcel.fast_xlsx import list_sheets
from autoexcel.flow_sync import (
    LogCallback,
    WALLET_FLOW_HEADERS,
    _cell_at,
    _cell_has_value,
    _cell_text,
    _inline_text,
    _load_shared_strings,
    _numeric_cell_value,
    _related_part,
    _requested_source_shared_indices,
    _row_number,
    _set_row_number,
    _split_cell_ref,
    _tag,
    _validate_wallet_flow_header,
    _write_replaced_archive,
)


CUMULATIVE_HEADERS = (
    "是否可疑订单",
    "完成时间",
    "账单时间 ",
    "交易流水号",
    "钱包ID",
    "用户钱包ID",
    "钱包类型",
    "列1",
    "服务商",
    "交易类型",
    "期初余额",
    "交易金额",
    "渠道成本",
    "期末余额",
    "是否B2B回款",
    "是否补单",
    "提取时间",
    "列2",
)
SOURCE_TO_TARGET_COLUMNS = {
    "A": "C",
    "B": "D",
    "C": "E",
    "D": "F",
    "E": "G",
    "F": "H",
    "G": "I",
    "I": "J",
    "J": "K",
    "K": "L",
    "L": "M",
    "M": "N",
    "N": "O",
}
TARGET_NUMERIC_COLUMNS = frozenset(("K", "L", "M", "N"))
DAILY_SUMMARY_HEADERS = (
    "钱包ID",
    "求和项:交易金额",
    "计数项:交易类型",
)
BROKEN_SUPPLEMENT_FORMULA = re.compile(
    rb'(<(?:[A-Za-z_][\w.-]*:)?c\b[^>]*\br="P(\d+)"[^>]*>'
    rb"<(?:[A-Za-z_][\w.-]*:)?f\b[^>]*>[^<]*"
    rb"TP\xe4\xbb\xa3\xe6\x94\xb6!L:L,)"
    rb"#REF!"
    rb"([^<]*</(?:[A-Za-z_][\w.-]*:)?f>)"
)


@dataclass(frozen=True)
class DailyLongSyncResult:
    source_rows: int
    inserted_rows: int
    skipped_rows: int
    old_summary_row: int
    new_summary_row: int
    repaired_formula_rows: int


def _daily_detail_rows(
    root: ET.Element,
    shared_strings: dict[int, str],
) -> list[ET.Element]:
    sheet_data = root.find(_tag("sheetData"))
    if sheet_data is None:
        raise ValueError("“长款(当日)”缺少 sheetData")
    rows = sheet_data.findall(_tag("row"))
    rows_by_number = {_row_number(row): row for row in rows}
    header = rows_by_number.get(1)
    if header is None:
        raise ValueError("“长款(当日)”缺少第 1 行表头")
    _validate_wallet_flow_header(header, shared_strings, "长款(当日)")

    summary_row = next(
        (
            _row_number(row)
            for row in rows
            if tuple(
                _cell_text(_cell_at(row, column), shared_strings)
                for column in ("B", "C", "D")
            )
            == DAILY_SUMMARY_HEADERS
        ),
        None,
    )
    if summary_row is None:
        raise ValueError("“长款(当日)”未识别到 B:D 列的数据透视表头")

    detail_rows: list[ET.Element] = []
    for row_number in range(2, summary_row):
        row = rows_by_number.get(row_number)
        if row is None:
            continue
        transaction_id = _cell_text(_cell_at(row, "F"), shared_strings)
        has_detail_content = any(
            _cell_has_value(_cell_at(row, get_column_letter(column)))
            for column in range(3, 16)
        )
        if not transaction_id:
            if has_detail_content:
                raise ValueError(
                    f"“长款(当日)”第 {row_number} 行存在明细内容但缺少交易流水号"
                )
            continue
        detail_rows.append(row)
    if not detail_rows:
        raise ValueError("“长款(当日)”第 2 行开始没有可识别的明细")
    return detail_rows


def _cell_token(cell: ET.Element | None) -> tuple[str, str | int] | None:
    if cell is None:
        return None
    if cell.attrib.get("t") == "s":
        value = cell.find(_tag("v"))
        if value is None or value.text is None:
            return None
        return ("shared", int(value.text))
    if cell.attrib.get("t") == "inlineStr":
        return ("text", _inline_text(cell))
    value = cell.find(_tag("v"))
    if value is None or value.text is None:
        return None
    return ("text", value.text)


def _target_rows_and_keys(
    archive: zipfile.ZipFile,
    sheet_part: str,
    summary_row: int,
) -> tuple[
    ET.Element,
    ET.Element,
    list[ET.Element],
    list[tuple[str, str | int]],
    set[int],
    int,
]:
    template_row: ET.Element | None = None
    total_row: ET.Element | None = None
    trailing_rows: list[ET.Element] = []
    key_tokens: list[tuple[str, str | int]] = []
    shared_indices: set[int] = set()
    max_row = 0
    with archive.open(sheet_part) as stream:
        for _event, element in ET.iterparse(stream, events=("end",)):
            if element.tag != _tag("row"):
                continue
            row_number = _row_number(element)
            max_row = max(max_row, row_number)
            if 2 <= row_number < summary_row:
                token = _cell_token(_cell_at(element, "D"))
                if token is not None:
                    key_tokens.append(token)
                    if token[0] == "shared":
                        shared_indices.add(int(token[1]))
            if row_number == summary_row - 1:
                template_row = deepcopy(element)
            elif row_number == summary_row:
                total_row = deepcopy(element)
            elif row_number > summary_row:
                if any(
                    _cell_has_value(cell)
                    or cell.find(_tag("f")) is not None
                    for cell in element.findall(_tag("c"))
                ):
                    raise ValueError("“长款累计”的汇总行之后存在数据或公式")
                trailing_rows.append(deepcopy(element))
            element.clear()

    if template_row is None or total_row is None:
        raise ValueError("“长款累计”缺少末条明细或汇总行")
    return (
        template_row,
        total_row,
        trailing_rows,
        key_tokens,
        shared_indices,
        max_row,
    )


def _table_details(
    archive: zipfile.ZipFile,
    sheet_part: str,
) -> tuple[str, ET.Element, int]:
    table_part = _related_part(archive, sheet_part, "/table")
    if table_part not in archive.namelist():
        raise ValueError("“长款累计”的 Excel 表格文件不存在")
    table_root = ET.fromstring(archive.read(table_part))
    table_ref = table_root.attrib.get("ref")
    if not table_ref:
        raise ValueError("“长款累计”的 Excel 表格缺少范围")
    min_col, min_row, max_col, max_row = range_boundaries(table_ref)
    if (min_col, min_row, max_col) != (1, 1, 18):
        raise ValueError("“长款累计”的 Excel 表格范围必须从 A1 延伸到 R列")
    columns = table_root.find(_tag("tableColumns"))
    actual_headers = (
        tuple(column.attrib.get("name", "") for column in columns)
        if columns is not None
        else ()
    )
    if actual_headers != CUMULATIVE_HEADERS:
        raise ValueError("“长款累计”的 Excel 表格表头与预期不一致")
    if table_root.attrib.get("totalsRowCount") != "1":
        raise ValueError("“长款累计”的 Excel 表格未启用末行汇总")
    return table_part, table_root, max_row


def _set_cell_value(
    target_cell: ET.Element,
    source_cell: ET.Element | None,
    source_shared_strings: dict[int, str],
    numeric: bool,
) -> None:
    for child in list(target_cell):
        target_cell.remove(child)
    target_cell.attrib.pop("t", None)
    text = _cell_text(source_cell, source_shared_strings)
    if not text:
        return
    numeric_value = _numeric_cell_value(text) if numeric else None
    if numeric_value is not None:
        ET.SubElement(target_cell, _tag("v")).text = numeric_value
        return
    target_cell.set("t", "inlineStr")
    inline = ET.SubElement(target_cell, _tag("is"))
    ET.SubElement(inline, _tag("t")).text = text


def _set_formula(cell: ET.Element, formula: str) -> None:
    for child in list(cell):
        cell.remove(child)
    cell.set("t", "str")
    ET.SubElement(cell, _tag("f")).text = formula


def _make_cumulative_row(
    template_row: ET.Element,
    source_row: ET.Element,
    row_number: int,
    source_shared_strings: dict[int, str],
) -> ET.Element:
    row = deepcopy(template_row)
    _set_row_number(row, row_number)
    cells = {
        _split_cell_ref(cell.attrib["r"])[0]: cell
        for cell in row.findall(_tag("c"))
        if cell.attrib.get("r")
    }
    for target_column, source_column in SOURCE_TO_TARGET_COLUMNS.items():
        target_cell = cells.get(target_column)
        if target_cell is None:
            raise ValueError(
                f"“长款累计”末条明细缺少 {target_column} 列模板单元格"
            )
        _set_cell_value(
            target_cell,
            _cell_at(source_row, source_column),
            source_shared_strings,
            target_column in TARGET_NUMERIC_COLUMNS,
        )
    _set_formula(
        cells["O"],
        f'IF(COUNTIF(B2B回款!B:B,D{row_number}),"YES","NO")',
    )
    _set_formula(
        cells["P"],
        f'IF(COUNTIF(TP代收!L:L,D{row_number}),"YES","NO")',
    )
    _set_formula(cells["Q"], f'TEXT(INT(B{row_number}),"YYYYMMDD")')
    return row


def _update_dimension(
    sheet_xml: bytes,
    old_summary_row: int,
    new_summary_row: int,
) -> bytes:
    pattern = re.compile(rb'(<dimension\b[^>]*\bref=")([^"]+)(")')
    match = pattern.search(sheet_xml)
    if match is None:
        return sheet_xml
    ref = match.group(2).decode("ascii")
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    if max_row < old_summary_row:
        raise ValueError("“长款累计”的工作表范围与汇总行不一致")
    updated_max_row = max_row + new_summary_row - old_summary_row
    updated_ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{updated_max_row}"
    ).encode("ascii")
    return sheet_xml[: match.start(2)] + updated_ref + sheet_xml[match.end(2) :]


def _repair_supplement_formulas(sheet_xml: bytes) -> tuple[bytes, int]:
    def replace(match: re.Match[bytes]) -> bytes:
        return (
            match.group(1)
            + f"D{match.group(2).decode('ascii')}".encode("ascii")
            + match.group(3)
        )

    return BROKEN_SUPPLEMENT_FORMULA.subn(replace, sheet_xml)


def _updated_table_xml(
    table_root: ET.Element,
    new_summary_row: int,
) -> bytes:
    table_root.set("ref", f"A1:R{new_summary_row}")
    auto_filter = table_root.find(_tag("autoFilter"))
    if auto_filter is None:
        raise ValueError("“长款累计”的 Excel 表格缺少筛选范围")
    auto_filter.set("ref", f"A1:R{new_summary_row - 1}")
    columns = table_root.find(_tag("tableColumns"))
    supplement = next(
        (
            column
            for column in columns or ()
            if column.attrib.get("name") == "是否补单"
        ),
        None,
    )
    if supplement is None:
        raise ValueError("“长款累计”的 Excel 表格缺少“是否补单”列")
    formula = supplement.find(_tag("calculatedColumnFormula"))
    if formula is None:
        formula = ET.SubElement(supplement, _tag("calculatedColumnFormula"))
    formula.text = 'IF(COUNTIF(TP代收!L:L,D2),"YES","NO")'
    return ET.tostring(
        table_root,
        encoding="utf-8",
        xml_declaration=True,
    )


def _updated_workbook_xml(archive: zipfile.ZipFile) -> bytes:
    root = ET.fromstring(archive.read("xl/workbook.xml"))
    calc_pr = root.find(_tag("calcPr"))
    if calc_pr is None:
        calc_pr = ET.SubElement(root, _tag("calcPr"))
    calc_pr.set("fullCalcOnLoad", "1")
    calc_pr.set("forceFullCalc", "1")
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def sync_daily_long(workbook_path: Path, progress: LogCallback | None = None) -> DailyLongSyncResult:
    if not workbook_path.is_file():
        raise FileNotFoundError(f"工作簿不存在：{workbook_path}")
    if workbook_path.suffix.lower() != ".xlsx":
        raise ValueError("工作簿必须是 .xlsx 文件")

    entries = {entry.name: entry for entry in list_sheets(workbook_path)}
    missing = [
        name for name in ("长款(当日)", "长款累计") if name not in entries
    ]
    if missing:
        raise ValueError(f"工作簿缺少工作表：{'、'.join(missing)}")

    daily_entry = entries["长款(当日)"]
    cumulative_entry = entries["长款累计"]
    if progress is not None:
        progress("正在动态识别“长款(当日)”有效明细…")

    with zipfile.ZipFile(workbook_path, "r") as archive:
        daily_root = ET.fromstring(archive.read(daily_entry.path))
        daily_shared_indices = _requested_source_shared_indices(daily_root)
        table_part, table_root, old_summary_row = _table_details(
            archive, cumulative_entry.path
        )
        (
            template_row,
            total_row,
            trailing_rows,
            key_tokens,
            target_shared_indices,
            _last_physical_row,
        ) = _target_rows_and_keys(
            archive,
            cumulative_entry.path,
            old_summary_row,
        )
        shared_strings = _load_shared_strings(
            archive,
            daily_shared_indices | target_shared_indices,
        )
        source_rows = _daily_detail_rows(daily_root, shared_strings)
        existing_keys = {
            shared_strings[int(value)] if kind == "shared" else str(value)
            for kind, value in key_tokens
        }

        new_source_rows: list[ET.Element] = []
        seen_keys = set(existing_keys)
        for source_row in source_rows:
            transaction_id = _cell_text(
                _cell_at(source_row, "F"),
                shared_strings,
            )
            if transaction_id in seen_keys:
                continue
            seen_keys.add(transaction_id)
            new_source_rows.append(source_row)

        inserted_rows = len(new_source_rows)
        new_summary_row = old_summary_row + inserted_rows
        new_rows = [
            _make_cumulative_row(
                template_row,
                source_row,
                old_summary_row + offset,
                shared_strings,
            )
            for offset, source_row in enumerate(new_source_rows)
        ]
        _set_row_number(total_row, new_summary_row)
        for row in trailing_rows:
            _set_row_number(row, _row_number(row) + inserted_rows)

        if progress is not None:
            progress(
                f"识别 {len(source_rows)} 条，去重后新增 {inserted_rows} 条；"
                "正在修复“是否补单”公式…"
            )
        sheet_xml = archive.read(cumulative_entry.path)
        summary_start_pattern = re.compile(
            rb'<(?:[A-Za-z_][\w.-]*:)?row\b[^>]*\br="'
            + str(old_summary_row).encode("ascii")
            + rb'"[^>]*>'
        )
        summary_start = summary_start_pattern.search(sheet_xml)
        sheet_data_end = re.search(
            rb"</(?:[A-Za-z_][\w.-]*:)?sheetData>",
            sheet_xml[summary_start.end() :] if summary_start else b"",
        )
        if summary_start is None or sheet_data_end is None:
            raise ValueError("未能在“长款累计”中定位汇总行 XML")
        sheet_data_end_start = summary_start.end() + sheet_data_end.start()
        inserted_xml = b"".join(
            ET.tostring(row, encoding="utf-8") for row in new_rows
        )
        total_xml = ET.tostring(total_row, encoding="utf-8")
        trailing_xml = b"".join(
            ET.tostring(row, encoding="utf-8") for row in trailing_rows
        )
        sheet_xml = (
            sheet_xml[: summary_start.start()]
            + inserted_xml
            + total_xml
            + trailing_xml
            + sheet_xml[sheet_data_end_start:]
        )
        sheet_xml = _update_dimension(
            sheet_xml,
            old_summary_row,
            new_summary_row,
        )
        sheet_xml, repaired_formula_rows = _repair_supplement_formulas(sheet_xml)

        replacements = {
            table_part: _updated_table_xml(table_root, new_summary_row),
            "xl/workbook.xml": _updated_workbook_xml(archive),
        }
        file_mode = stat.S_IMODE(workbook_path.stat().st_mode)
        fd, temporary_name = tempfile.mkstemp(
            prefix=f"{workbook_path.stem}.",
            suffix=".tmp.xlsx",
            dir=workbook_path.parent,
        )
        os.close(fd)
        temporary_path = Path(temporary_name)
        try:
            if progress is not None:
                progress("正在写回工作簿，请勿打开 Excel/WPS…")
            _write_replaced_archive(
                archive,
                temporary_path,
                cumulative_entry.path,
                sheet_xml,
                replacements,
            )
            os.chmod(temporary_path, file_mode)
        except Exception:
            if temporary_path.exists():
                temporary_path.unlink()
            raise

    try:
        os.replace(temporary_path, workbook_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()

    return DailyLongSyncResult(
        source_rows=len(source_rows),
        inserted_rows=inserted_rows,
        skipped_rows=len(source_rows) - inserted_rows,
        old_summary_row=old_summary_row,
        new_summary_row=new_summary_row,
        repaired_formula_rows=repaired_formula_rows,
    )
