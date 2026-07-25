from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
import os
from pathlib import Path
import re
import shutil
import stat
import tempfile
from typing import Callable
import xml.etree.ElementTree as ET
import zipfile

from openpyxl.formula.translate import Translator
from openpyxl.utils.cell import get_column_letter, range_boundaries

from autoexcel.fast_xlsx import MAIN_NS, list_sheets


LogCallback = Callable[[str], None]
SUMMARY_HEADERS = ("支付通道", "平台钱包ID")
TP_PAYOUT_HEADERS = (
    "平台订单号",
    "商户订单号",
    "渠道订单号",
    "商户名称",
    "商户ID",
    "付款金额(৳)",
    "手续费(৳)",
    "到账金额(৳)",
    "渠道成本(৳)",
    "创建时间",
    "更新时间",
    "支付方式名称",
    "支付方式",
    "支付通道",
    "平台钱包ID",
    "用户钱包ID",
    "来源",
    "拨号状态",
    "换卡次数",
    "交易状态",
    "回调状态",
    "失败信息",
)
TP_COLLECTION_HEADERS = (
    "平台订单号",
    "商户订单号",
    "渠道订单号",
    "请求上游订单号",
    "商户名称",
    "商户ID",
    "交易金额(৳)",
    "实收金额(৳)",
    "手续费(৳)",
    "到账金额(৳)",
    "钱包类型",
    "trxId",
    "收款人钱包id",
    "用户钱包ID",
    "用户 IP",
    "设备编号",
    "USER_ID",
    "创建时间",
    "完成时间",
    "支付通道",
    "收款方式",
    "收款方式名称",
    "交易状态",
    "通知状态",
    "产品名称",
    "用户ID",
)
TP_COLLECTION_STATUSES = ("支付成功", "部分支付")
WALLET_FLOW_HEADERS = (
    "平台订单号",
    "商户订单号",
    "是否可疑订单",
    "完成时间",
    "账单时间",
    "交易流水号",
    "钱包ID",
    "用户钱包ID",
    "钱包类型",
    "服务商",
    "交易类型",
    "期初余额",
    "交易金额",
    "渠道成本",
    "期末余额",
)


@dataclass(frozen=True)
class TpPayoutSyncResult:
    removed_rows: int
    inserted_rows: int
    old_detail_end_row: int
    new_detail_end_row: int
    summary_row: int | None
    shifted_rows: int


@dataclass(frozen=True)
class TpCollectionSyncResult(TpPayoutSyncResult):
    successful_rows: int
    partial_rows: int


def _tag(name: str) -> str:
    return f"{{{MAIN_NS}}}{name}"


def _row_number(row: ET.Element) -> int:
    return int(row.attrib["r"])


def _split_cell_ref(ref: str) -> tuple[str, int]:
    match = re.fullmatch(r"([A-Z]+)([0-9]+)", ref)
    if not match:
        raise ValueError(f"不支持的单元格地址：{ref}")
    return match.group(1), int(match.group(2))


def _cell_at(row: ET.Element | None, column: str) -> ET.Element | None:
    if row is None:
        return None
    for cell in row.findall(_tag("c")):
        ref = cell.attrib.get("r")
        if ref and _split_cell_ref(ref)[0] == column:
            return cell
    return None


def _inline_text(cell: ET.Element) -> str:
    inline = cell.find(_tag("is"))
    if inline is None:
        return ""
    return "".join(node.text or "" for node in inline.iter(_tag("t")))


def _cell_text(cell: ET.Element | None, shared_strings: dict[int, str]) -> str:
    if cell is None:
        return ""
    if cell.attrib.get("t") == "inlineStr":
        return _inline_text(cell)
    value = cell.find(_tag("v"))
    if value is None or value.text is None:
        return ""
    if cell.attrib.get("t") == "s":
        return shared_strings.get(int(value.text), "")
    return value.text


def _cell_has_value(cell: ET.Element | None) -> bool:
    if cell is None:
        return False
    if cell.attrib.get("t") == "inlineStr":
        return bool(_inline_text(cell))
    value = cell.find(_tag("v"))
    return value is not None and bool(value.text)


def _shared_string_indices(cells: list[ET.Element]) -> set[int]:
    indices: set[int] = set()
    for cell in cells:
        if cell.attrib.get("t") != "s":
            continue
        value = cell.find(_tag("v"))
        if value is not None and value.text is not None:
            indices.add(int(value.text))
    return indices


def _load_shared_strings(
    archive: zipfile.ZipFile,
    indices: set[int],
) -> dict[int, str]:
    if not indices or "xl/sharedStrings.xml" not in archive.namelist():
        return {}

    found: dict[int, str] = {}
    current_index = -1
    with archive.open("xl/sharedStrings.xml") as stream:
        for _event, element in ET.iterparse(stream, events=("end",)):
            if element.tag != _tag("si"):
                continue
            current_index += 1
            if current_index in indices:
                found[current_index] = "".join(
                    node.text or "" for node in element.iter(_tag("t"))
                )
            element.clear()
            if len(found) == len(indices):
                break
    return found


def _header_values(
    row: ET.Element,
    shared_strings: dict[int, str],
    column_count: int = len(TP_PAYOUT_HEADERS),
) -> tuple[str, ...]:
    return tuple(
        _cell_text(_cell_at(row, get_column_letter(column)), shared_strings)
        for column in range(1, column_count + 1)
    )


def _validate_header(
    row: ET.Element,
    shared_strings: dict[int, str],
    label: str,
) -> None:
    actual = _header_values(row, shared_strings)
    if actual != TP_PAYOUT_HEADERS:
        raise ValueError(f"{label}表头与 TP代付 需要的 A:V 字段不一致")


def _validate_collection_header(
    row: ET.Element,
    shared_strings: dict[int, str],
    label: str,
) -> None:
    actual = _header_values(row, shared_strings, len(TP_COLLECTION_HEADERS))
    differences = [
        index
        for index, (value, expected) in enumerate(
            zip(actual, TP_COLLECTION_HEADERS)
        )
        if value != expected
    ]
    if differences == [12] and actual[12] == "平台钱包ID":
        return
    if differences:
        raise ValueError(f"{label}表头与 TP代收 需要的 A:Z 字段不一致")


def _validate_wallet_flow_header(
    row: ET.Element,
    shared_strings: dict[int, str],
    label: str,
) -> None:
    actual = _header_values(row, shared_strings, len(WALLET_FLOW_HEADERS))
    if actual != WALLET_FLOW_HEADERS:
        raise ValueError(f"{label}表头与钱包流水需要的 A:O 字段不一致")


def _continuous_detail_end(rows_by_number: dict[int, ET.Element], label: str) -> int:
    row_number = 2
    while _cell_has_value(_cell_at(rows_by_number.get(row_number), "A")):
        row_number += 1
    detail_end = row_number - 1
    if detail_end < 2:
        raise ValueError(f"{label}第 2 行开始没有可识别的平台订单号")
    return detail_end


def _continuous_key_end(
    rows_by_number: dict[int, ET.Element],
    label: str,
    start_row: int,
    key_column: str,
) -> int:
    row_number = start_row
    while _cell_has_value(
        _cell_at(rows_by_number.get(row_number), key_column)
    ):
        row_number += 1
    detail_end = row_number - 1
    if detail_end < start_row:
        raise ValueError(
            f"{label}第 {start_row} 行开始没有可识别的交易流水号"
        )
    return detail_end


def _set_row_number(row: ET.Element, row_number: int) -> None:
    old_row_number = _row_number(row)
    for cell in row.findall(_tag("c")):
        ref = cell.attrib.get("r")
        if not ref:
            continue
        column, _ = _split_cell_ref(ref)
        formula = cell.find(_tag("f"))
        if formula is not None and formula.text:
            formula.text = Translator(
                f"={formula.text}",
                origin=f"{column}{old_row_number}",
            ).translate_formula(f"{column}{row_number}")[1:]
        cell.set("r", f"{column}{row_number}")
    row.set("r", str(row_number))


def _summary_header_row(
    rows: list[ET.Element],
    detail_end: int,
    shared_strings: dict[int, str],
) -> int | None:
    trailing_content = False
    for row in rows:
        row_number = _row_number(row)
        if row_number <= detail_end:
            continue
        cells = row.findall(_tag("c"))
        if any(_cell_has_value(cell) for cell in cells):
            trailing_content = True
        if (
            _cell_text(_cell_at(row, "E"), shared_strings) == SUMMARY_HEADERS[0]
            and _cell_text(_cell_at(row, "F"), shared_strings) == SUMMARY_HEADERS[1]
        ):
            return row_number
    if trailing_content:
        raise ValueError(
            "TP代付明细后存在内容，"
            "但未识别到“支付通道 / 平台钱包ID”汇总表头"
        )
    return None


def _collection_summary_header_row(
    rows: list[ET.Element],
    detail_end: int,
    shared_strings: dict[int, str],
) -> int | None:
    trailing_content = False
    for row in rows:
        row_number = _row_number(row)
        if row_number <= detail_end:
            continue
        cells = row.findall(_tag("c"))
        if any(_cell_has_value(cell) for cell in cells):
            trailing_content = True
        if (
            _cell_text(_cell_at(row, "D"), shared_strings)
            == "求和项:实收金额(৳)"
            and _cell_text(_cell_at(row, "E"), shared_strings)
            == "求和项:手续费(৳)"
            and _cell_text(_cell_at(row, "F"), shared_strings)
            == "计数项:平台订单号"
        ):
            return row_number
    if trailing_content:
        raise ValueError(
            "TP代收明细后存在内容，但未识别到 D:F 列的汇总表头"
        )
    return None


def _wallet_summary_header_row(
    rows: list[ET.Element],
    detail_end: int,
    shared_strings: dict[int, str],
) -> int | None:
    trailing_content = False
    for row in rows:
        row_number = _row_number(row)
        if row_number <= detail_end:
            continue
        cells = row.findall(_tag("c"))
        if any(_cell_has_value(cell) for cell in cells):
            trailing_content = True
        if (
            _cell_text(_cell_at(row, "B"), shared_strings) == "钱包ID"
            and _cell_text(_cell_at(row, "C"), shared_strings)
            == "求和项:交易金额"
            and _cell_text(_cell_at(row, "D"), shared_strings)
            == "计数项:交易类型"
        ):
            return row_number
    if trailing_content:
        raise ValueError(
            "长款(当日)明细后存在内容，但未识别到 B:D 列的汇总表头"
        )
    return None


def _make_target_row(
    source_row: ET.Element,
    row_number: int,
    template_attributes: dict[str, str],
    styles_by_column: dict[str, str],
    source_shared_strings: dict[int, str],
) -> ET.Element:
    row = deepcopy(source_row)
    row.attrib.clear()
    row.attrib.update(template_attributes)
    row.set("r", str(row_number))
    for cell in row.findall(_tag("c")):
        ref = cell.attrib.get("r")
        if not ref:
            continue
        column, _ = _split_cell_ref(ref)
        cell.set("r", f"{column}{row_number}")
        if cell.attrib.get("t") == "s":
            value = cell.find(_tag("v"))
            shared_index = int(value.text) if value is not None and value.text else -1
            text = source_shared_strings.get(shared_index)
            if text is None:
                raise ValueError(f"订单文件共享字符串索引无效：{shared_index}")
            for child in list(cell):
                cell.remove(child)
            cell.set("t", "inlineStr")
            inline = ET.SubElement(cell, _tag("is"))
            ET.SubElement(inline, _tag("t")).text = text
        style = styles_by_column.get(column)
        if style is None:
            cell.attrib.pop("s", None)
        else:
            cell.set("s", style)
    return row


def _replace_detail_rows(
    target_root: ET.Element,
    target_rows: list[ET.Element],
    source_groups: list[tuple[list[ET.Element], dict[int, str]]],
    old_detail_end: int,
    summary_row: int | None,
    max_column: str,
) -> TpPayoutSyncResult:
    target_sheet_data = target_root.find(_tag("sheetData"))
    if target_sheet_data is None:
        raise ValueError("工作簿缺少 sheetData")

    target_by_number = {_row_number(row): row for row in target_rows}
    old_count = old_detail_end - 1
    new_count = sum(len(rows) for rows, _shared_strings in source_groups)
    template_row = target_by_number[2]
    template_attributes = dict(template_row.attrib)
    styles_by_column = {
        _split_cell_ref(cell.attrib["r"])[0]: cell.attrib["s"]
        for cell in template_row.findall(_tag("c"))
        if "r" in cell.attrib and "s" in cell.attrib
    }

    new_rows: list[ET.Element] = []
    output_row_number = 2
    for source_rows, source_shared_strings in source_groups:
        for source_row in source_rows:
            new_rows.append(
                _make_target_row(
                    source_row,
                    output_row_number,
                    template_attributes,
                    styles_by_column,
                    source_shared_strings,
                )
            )
            output_row_number += 1

    shift = max(0, new_count - old_count)
    leading_rows = [row for row in target_rows if _row_number(row) < 2]
    trailing_rows = [row for row in target_rows if _row_number(row) > old_detail_end]
    if shift:
        for row in trailing_rows:
            _set_row_number(row, _row_number(row) + shift)

    sheet_data_attributes = dict(target_sheet_data.attrib)
    target_sheet_data.clear()
    target_sheet_data.attrib.update(sheet_data_attributes)
    for row in (*leading_rows, *new_rows, *trailing_rows):
        target_sheet_data.append(row)

    new_detail_end = new_count + 1
    auto_filter = target_root.find(_tag("autoFilter"))
    if auto_filter is not None:
        auto_filter.set("ref", f"A1:{max_column}{new_detail_end}")

    dimension = target_root.find(_tag("dimension"))
    if dimension is not None:
        ref = dimension.attrib.get("ref")
        if ref and ":" in ref:
            min_col, min_row, dimension_max_col, _max_row = range_boundaries(ref)
            max_row = max(
                _row_number(row)
                for row in target_sheet_data.findall(_tag("row"))
            )
            dimension.set(
                "ref",
                f"{get_column_letter(min_col)}{min_row}:"
                f"{get_column_letter(dimension_max_col)}{max_row}",
            )

    return TpPayoutSyncResult(
        removed_rows=old_count,
        inserted_rows=new_count,
        old_detail_end_row=old_detail_end,
        new_detail_end_row=new_detail_end,
        summary_row=summary_row + shift if summary_row is not None else None,
        shifted_rows=shift,
    )


def _replace_tp_payout_roots(
    target_root: ET.Element,
    source_root: ET.Element,
    target_shared_strings: dict[int, str],
    source_shared_strings: dict[int, str],
) -> TpPayoutSyncResult:
    target_sheet_data = target_root.find(_tag("sheetData"))
    source_sheet_data = source_root.find(_tag("sheetData"))
    if target_sheet_data is None or source_sheet_data is None:
        raise ValueError("工作簿缺少 sheetData")

    target_rows = target_sheet_data.findall(_tag("row"))
    source_rows = source_sheet_data.findall(_tag("row"))
    target_by_number = {_row_number(row): row for row in target_rows}
    source_by_number = {_row_number(row): row for row in source_rows}
    target_header = target_by_number.get(1)
    source_header = source_by_number.get(1)
    if target_header is None or source_header is None:
        raise ValueError("目标工作表或付款订单缺少第 1 行表头")
    _validate_header(target_header, target_shared_strings, "TP代付")
    _validate_header(source_header, source_shared_strings, "付款订单")

    old_detail_end = _continuous_detail_end(target_by_number, "TP代付")
    source_detail_end = _continuous_detail_end(source_by_number, "付款订单")
    for row_number in range(2, source_detail_end + 1):
        if not _cell_has_value(_cell_at(source_by_number.get(row_number), "A")):
            raise ValueError(f"付款订单第 {row_number} 行缺少平台订单号")

    summary_row = _summary_header_row(
        target_rows,
        old_detail_end,
        target_shared_strings,
    )
    return _replace_detail_rows(
        target_root,
        target_rows,
        [
            (
                [
                    source_by_number[row_number]
                    for row_number in range(2, source_detail_end + 1)
                ],
                source_shared_strings,
            )
        ],
        old_detail_end,
        summary_row,
        "V",
    )


def _collection_source_rows(
    root: ET.Element,
    shared_strings: dict[int, str],
    label: str,
) -> tuple[str, list[ET.Element], list[str]]:
    sheet_data = root.find(_tag("sheetData"))
    if sheet_data is None:
        raise ValueError(f"{label}缺少 sheetData")
    rows = sheet_data.findall(_tag("row"))
    rows_by_number = {_row_number(row): row for row in rows}
    header = rows_by_number.get(1)
    if header is None:
        raise ValueError(f"{label}缺少第 1 行表头")
    _validate_collection_header(header, shared_strings, label)
    detail_end = _continuous_detail_end(rows_by_number, label)

    statuses: set[str] = set()
    order_ids: list[str] = []
    detail_rows: list[ET.Element] = []
    for row_number in range(2, detail_end + 1):
        row = rows_by_number[row_number]
        order_id = _cell_text(_cell_at(row, "A"), shared_strings)
        status = _cell_text(_cell_at(row, "W"), shared_strings)
        if not order_id:
            raise ValueError(f"{label}第 {row_number} 行缺少平台订单号")
        if status not in TP_COLLECTION_STATUSES:
            raise ValueError(
                f"{label}第 {row_number} 行交易状态“{status}”不受支持"
            )
        statuses.add(status)
        order_ids.append(order_id)
        detail_rows.append(row)

    if len(statuses) != 1:
        raise ValueError(f"{label}同时包含多种交易状态，无法自动分类")
    return statuses.pop(), detail_rows, order_ids


def _replace_tp_collection_roots(
    target_root: ET.Element,
    first_source_root: ET.Element,
    second_source_root: ET.Element,
    target_shared_strings: dict[int, str],
    first_source_shared_strings: dict[int, str],
    second_source_shared_strings: dict[int, str],
) -> TpCollectionSyncResult:
    target_sheet_data = target_root.find(_tag("sheetData"))
    if target_sheet_data is None:
        raise ValueError("工作簿缺少 sheetData")
    target_rows = target_sheet_data.findall(_tag("row"))
    target_by_number = {_row_number(row): row for row in target_rows}
    target_header = target_by_number.get(1)
    if target_header is None:
        raise ValueError("TP代收缺少第 1 行表头")
    _validate_collection_header(target_header, target_shared_strings, "TP代收")
    old_detail_end = _continuous_detail_end(target_by_number, "TP代收")
    summary_row = _collection_summary_header_row(
        target_rows,
        old_detail_end,
        target_shared_strings,
    )

    sources_by_status: dict[
        str, tuple[list[ET.Element], dict[int, str], list[str]]
    ] = {}
    for label, root, shared_strings in (
        ("收款订单文件 1", first_source_root, first_source_shared_strings),
        ("收款订单文件 2", second_source_root, second_source_shared_strings),
    ):
        status, rows, order_ids = _collection_source_rows(
            root,
            shared_strings,
            label,
        )
        if status in sources_by_status:
            raise ValueError(
                f"两个收款订单文件都是“{status}”，"
                "需要分别上传支付成功和部分支付文件"
            )
        sources_by_status[status] = (rows, shared_strings, order_ids)

    missing_statuses = [
        status for status in TP_COLLECTION_STATUSES if status not in sources_by_status
    ]
    if missing_statuses:
        raise ValueError(f"缺少“{missing_statuses[0]}”收款订单文件")

    seen_order_ids: set[str] = set()
    for status in TP_COLLECTION_STATUSES:
        for order_id in sources_by_status[status][2]:
            if order_id in seen_order_ids:
                raise ValueError(f"平台订单号重复：{order_id}")
            seen_order_ids.add(order_id)

    base_result = _replace_detail_rows(
        target_root,
        target_rows,
        [
            (
                sources_by_status[status][0],
                sources_by_status[status][1],
            )
            for status in TP_COLLECTION_STATUSES
        ],
        old_detail_end,
        summary_row,
        "Z",
    )
    return TpCollectionSyncResult(
        **base_result.__dict__,
        successful_rows=len(sources_by_status["支付成功"][0]),
        partial_rows=len(sources_by_status["部分支付"][0]),
    )


def _replace_wallet_flow_roots(
    target_root: ET.Element,
    source_root: ET.Element,
    target_shared_strings: dict[int, str],
    source_shared_strings: dict[int, str],
) -> TpPayoutSyncResult:
    target_sheet_data = target_root.find(_tag("sheetData"))
    source_sheet_data = source_root.find(_tag("sheetData"))
    if target_sheet_data is None or source_sheet_data is None:
        raise ValueError("工作簿缺少 sheetData")

    target_rows = target_sheet_data.findall(_tag("row"))
    source_rows = source_sheet_data.findall(_tag("row"))
    target_by_number = {_row_number(row): row for row in target_rows}
    source_by_number = {_row_number(row): row for row in source_rows}
    target_header = target_by_number.get(1)
    source_header = source_by_number.get(2)
    if target_header is None or source_header is None:
        raise ValueError("长款(当日)或平台钱包流水记录缺少表头")
    _validate_wallet_flow_header(
        target_header,
        target_shared_strings,
        "长款(当日)",
    )
    _validate_wallet_flow_header(
        source_header,
        source_shared_strings,
        "平台钱包流水记录",
    )

    old_detail_end = _continuous_key_end(
        target_by_number,
        "长款(当日)",
        2,
        "F",
    )
    source_detail_end = _continuous_key_end(
        source_by_number,
        "平台钱包流水记录",
        3,
        "F",
    )
    summary_row = _wallet_summary_header_row(
        target_rows,
        old_detail_end,
        target_shared_strings,
    )
    return _replace_detail_rows(
        target_root,
        target_rows,
        [
            (
                [
                    source_by_number[row_number]
                    for row_number in range(3, source_detail_end + 1)
                ],
                source_shared_strings,
            )
        ],
        old_detail_end,
        summary_row,
        "N",
    )


def _requested_target_shared_indices(root: ET.Element) -> set[int]:
    sheet_data = root.find(_tag("sheetData"))
    if sheet_data is None:
        return set()
    rows = sheet_data.findall(_tag("row"))
    rows_by_number = {_row_number(row): row for row in rows}
    detail_end = _continuous_detail_end(rows_by_number, "TP代付")
    cells = list(rows_by_number[1].findall(_tag("c")))
    for row in rows:
        if _row_number(row) <= detail_end:
            continue
        for column in ("E", "F"):
            cell = _cell_at(row, column)
            if cell is not None:
                cells.append(cell)
    return _shared_string_indices(cells)


def _requested_source_shared_indices(root: ET.Element) -> set[int]:
    sheet_data = root.find(_tag("sheetData"))
    if sheet_data is None:
        return set()
    return _shared_string_indices(
        [
            cell
            for row in sheet_data.findall(_tag("row"))
            for cell in row.findall(_tag("c"))
        ]
    )


def _requested_collection_target_shared_indices(root: ET.Element) -> set[int]:
    sheet_data = root.find(_tag("sheetData"))
    if sheet_data is None:
        return set()
    rows = sheet_data.findall(_tag("row"))
    rows_by_number = {_row_number(row): row for row in rows}
    detail_end = _continuous_detail_end(rows_by_number, "TP代收")
    cells = list(rows_by_number[1].findall(_tag("c")))
    for row in rows:
        if _row_number(row) <= detail_end:
            continue
        for column in ("D", "E", "F"):
            cell = _cell_at(row, column)
            if cell is not None:
                cells.append(cell)
    return _shared_string_indices(cells)


def _requested_wallet_target_shared_indices(root: ET.Element) -> set[int]:
    sheet_data = root.find(_tag("sheetData"))
    if sheet_data is None:
        return set()
    rows = sheet_data.findall(_tag("row"))
    rows_by_number = {_row_number(row): row for row in rows}
    detail_end = _continuous_key_end(
        rows_by_number,
        "长款(当日)",
        2,
        "F",
    )
    cells = list(rows_by_number[1].findall(_tag("c")))
    for row in rows:
        if _row_number(row) <= detail_end:
            continue
        for column in ("B", "C", "D"):
            cell = _cell_at(row, column)
            if cell is not None:
                cells.append(cell)
    return _shared_string_indices(cells)


def _write_replaced_archive(
    archive: zipfile.ZipFile,
    output_path: Path,
    replaced_part: str,
    replacement: bytes,
) -> None:
    with zipfile.ZipFile(output_path, "w") as output:
        output.comment = archive.comment
        for item in archive.infolist():
            if item.filename == replaced_part:
                output.writestr(item, replacement)
                continue
            with archive.open(item, "r") as source, output.open(item, "w") as destination:
                shutil.copyfileobj(source, destination, length=1024 * 1024)


def sync_tp_payout(
    workbook_path: Path,
    payment_orders_path: Path,
    progress: LogCallback | None = None,
) -> TpPayoutSyncResult:
    for label, path in (
        ("工作簿", workbook_path),
        ("付款订单文件", payment_orders_path),
    ):
        if not path.is_file():
            raise FileNotFoundError(f"{label}不存在：{path}")
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"{label}必须是 .xlsx 文件")
    if workbook_path.resolve() == payment_orders_path.resolve():
        raise ValueError("工作簿和付款订单文件不能是同一个文件")

    target_entry = next(
        (entry for entry in list_sheets(workbook_path) if entry.name == "TP代付"),
        None,
    )
    if target_entry is None:
        raise ValueError("工作簿中没有名为“TP代付”的工作表")
    source_entries = list_sheets(payment_orders_path)
    if not source_entries:
        raise ValueError("付款订单文件中没有工作表")
    source_entry = source_entries[0]

    if progress is not None:
        progress("正在识别 TP代付 历史明细范围…")
    with zipfile.ZipFile(workbook_path, "r") as target_archive:
        target_root = ET.fromstring(target_archive.read(target_entry.path))
        target_shared = _load_shared_strings(
            target_archive,
            _requested_target_shared_indices(target_root),
        )

        if progress is not None:
            progress("正在读取付款订单明细…")
        with zipfile.ZipFile(payment_orders_path, "r") as source_archive:
            source_root = ET.fromstring(source_archive.read(source_entry.path))
            source_shared = _load_shared_strings(
                source_archive,
                _requested_source_shared_indices(source_root),
            )

        result = _replace_tp_payout_roots(
            target_root,
            source_root,
            target_shared,
            source_shared,
        )
        replacement = ET.tostring(
            target_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        file_mode = stat.S_IMODE(workbook_path.stat().st_mode)
        fd, temporary_name = tempfile.mkstemp(
            prefix=f"{workbook_path.stem}.",
            suffix=".tmp.xlsx",
            dir=workbook_path.parent,
        )
        os.close(fd)
        temporary_path = Path(temporary_name)
        try:
            if progress is not None:
                progress("正在写回工作簿，请勿打开 Excel/WPS…")
            _write_replaced_archive(
                target_archive,
                temporary_path,
                target_entry.path,
                replacement,
            )
            os.chmod(temporary_path, file_mode)
        except Exception:
            if temporary_path.exists():
                temporary_path.unlink()
            raise

    try:
        os.replace(temporary_path, workbook_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()

    return result


def sync_tp_collection(
    workbook_path: Path,
    first_collection_orders_path: Path,
    second_collection_orders_path: Path,
    progress: LogCallback | None = None,
) -> TpCollectionSyncResult:
    paths = (
        ("工作簿", workbook_path),
        ("收款订单文件 1", first_collection_orders_path),
        ("收款订单文件 2", second_collection_orders_path),
    )
    for label, path in paths:
        if not path.is_file():
            raise FileNotFoundError(f"{label}不存在：{path}")
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"{label}必须是 .xlsx 文件")
    resolved_paths = [path.resolve() for _label, path in paths]
    if len(set(resolved_paths)) != len(resolved_paths):
        raise ValueError("工作簿和两个收款订单文件必须是三个不同的文件")

    target_entry = next(
        (entry for entry in list_sheets(workbook_path) if entry.name == "TP代收"),
        None,
    )
    if target_entry is None:
        raise ValueError("工作簿中没有名为“TP代收”的工作表")

    source_paths = (
        first_collection_orders_path,
        second_collection_orders_path,
    )
    source_entries = []
    for index, source_path in enumerate(source_paths, start=1):
        entries = list_sheets(source_path)
        if not entries:
            raise ValueError(f"收款订单文件 {index} 中没有工作表")
        source_entries.append(entries[0])

    if progress is not None:
        progress("正在识别 TP代收 历史明细和汇总区域…")
    with zipfile.ZipFile(workbook_path, "r") as target_archive:
        target_root = ET.fromstring(target_archive.read(target_entry.path))
        target_shared = _load_shared_strings(
            target_archive,
            _requested_collection_target_shared_indices(target_root),
        )

        source_roots: list[ET.Element] = []
        source_shared_strings: list[dict[int, str]] = []
        for index, (source_path, source_entry) in enumerate(
            zip(source_paths, source_entries),
            start=1,
        ):
            if progress is not None:
                progress(f"正在读取收款订单文件 {index}…")
            with zipfile.ZipFile(source_path, "r") as source_archive:
                source_root = ET.fromstring(source_archive.read(source_entry.path))
                source_roots.append(source_root)
                source_shared_strings.append(
                    _load_shared_strings(
                        source_archive,
                        _requested_source_shared_indices(source_root),
                    )
                )

        if progress is not None:
            progress("正在校验交易状态和重复平台订单号…")
        result = _replace_tp_collection_roots(
            target_root,
            source_roots[0],
            source_roots[1],
            target_shared,
            source_shared_strings[0],
            source_shared_strings[1],
        )
        replacement = ET.tostring(
            target_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        file_mode = stat.S_IMODE(workbook_path.stat().st_mode)
        fd, temporary_name = tempfile.mkstemp(
            prefix=f"{workbook_path.stem}.",
            suffix=".tmp.xlsx",
            dir=workbook_path.parent,
        )
        os.close(fd)
        temporary_path = Path(temporary_name)
        try:
            if progress is not None:
                progress("正在写回工作簿，请勿打开 Excel/WPS…")
            _write_replaced_archive(
                target_archive,
                temporary_path,
                target_entry.path,
                replacement,
            )
            os.chmod(temporary_path, file_mode)
        except Exception:
            if temporary_path.exists():
                temporary_path.unlink()
            raise

    try:
        os.replace(temporary_path, workbook_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()

    return result


def sync_wallet_flow(
    workbook_path: Path,
    wallet_flow_path: Path,
    progress: LogCallback | None = None,
) -> TpPayoutSyncResult:
    for label, path in (
        ("工作簿", workbook_path),
        ("平台钱包流水记录", wallet_flow_path),
    ):
        if not path.is_file():
            raise FileNotFoundError(f"{label}不存在：{path}")
        if path.suffix.lower() != ".xlsx":
            raise ValueError(f"{label}必须是 .xlsx 文件")
    if workbook_path.resolve() == wallet_flow_path.resolve():
        raise ValueError("工作簿和平台钱包流水记录不能是同一个文件")

    target_entry = next(
        (
            entry
            for entry in list_sheets(workbook_path)
            if entry.name == "长款(当日)"
        ),
        None,
    )
    if target_entry is None:
        raise ValueError("工作簿中没有名为“长款(当日)”的工作表")
    source_entries = list_sheets(wallet_flow_path)
    if not source_entries:
        raise ValueError("平台钱包流水记录中没有工作表")
    source_entry = source_entries[0]

    if progress is not None:
        progress("正在识别 长款(当日) 历史明细和汇总区域…")
    with zipfile.ZipFile(workbook_path, "r") as target_archive:
        target_root = ET.fromstring(target_archive.read(target_entry.path))
        target_shared = _load_shared_strings(
            target_archive,
            _requested_wallet_target_shared_indices(target_root),
        )

        if progress is not None:
            progress("正在读取全部平台钱包流水记录…")
        with zipfile.ZipFile(wallet_flow_path, "r") as source_archive:
            source_root = ET.fromstring(source_archive.read(source_entry.path))
            source_shared = _load_shared_strings(
                source_archive,
                _requested_source_shared_indices(source_root),
            )

        result = _replace_wallet_flow_roots(
            target_root,
            source_root,
            target_shared,
            source_shared,
        )
        replacement = ET.tostring(
            target_root,
            encoding="utf-8",
            xml_declaration=True,
        )

        file_mode = stat.S_IMODE(workbook_path.stat().st_mode)
        fd, temporary_name = tempfile.mkstemp(
            prefix=f"{workbook_path.stem}.",
            suffix=".tmp.xlsx",
            dir=workbook_path.parent,
        )
        os.close(fd)
        temporary_path = Path(temporary_name)
        try:
            if progress is not None:
                progress("正在写回工作簿，请勿打开 Excel/WPS…")
            _write_replaced_archive(
                target_archive,
                temporary_path,
                target_entry.path,
                replacement,
            )
            os.chmod(temporary_path, file_mode)
        except Exception:
            if temporary_path.exists():
                temporary_path.unlink()
            raise

    try:
        os.replace(temporary_path, workbook_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()

    return result
