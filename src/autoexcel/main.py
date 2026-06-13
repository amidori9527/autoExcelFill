from __future__ import annotations

import argparse
import configparser
import re
import sys
import traceback
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from autoexcel.fast_xlsx import add_current_date_to_colored_sheets_fast
from autoexcel.operations import freeze_colored_sheets_next_day, freeze_next_day_row
from autoexcel.workbook_io import list_sheet_names, preview_sheet


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_FILE = PROJECT_ROOT / "data.xlsx"
EXAMPLE_SHEET = "example"
WORKSPACE_DIR_NAME = "workspace"
CONFIG_FILE_NAME = "config.ini"


@dataclass
class FillSummary:
    workbook: Path
    current_date: date
    log_path: Path
    changed: list[tuple[str, int]] = field(default_factory=list)
    skipped_count: int = 0
    batch_count: int = 0

    @property
    def changed_count(self) -> int:
        return len(self.changed)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Inspect and automate data.xlsx.")
    parser.add_argument("--workbook", type=Path, default=None)
    parser.add_argument("--sheet", default=EXAMPLE_SHEET)
    parser.add_argument("--rows", type=int, default=10)
    parser.add_argument(
        "--interactive",
        action="store_true",
        help="Choose an xlsx file from the current directory and use fill defaults.",
    )
    parser.add_argument("--list-sheets", action="store_true")
    parser.add_argument(
        "--copy-sheet-from",
        help="Create --sheet by copying this source sheet when --sheet does not exist.",
    )
    parser.add_argument(
        "--freeze-next-day-row",
        action="store_true",
        help="Insert a value row before the last date row, then advance that last row to current date.",
    )
    parser.add_argument(
        "--colored-sheets",
        action="store_true",
        help="Apply --freeze-next-day-row to every sheet with a tab color.",
    )
    parser.add_argument("--previous-date", default="2026-03-26", help="Deprecated; kept for old commands.")
    parser.add_argument(
        "--source-date",
        default="2026-03-27",
        help="Deprecated; kept for old commands.",
    )
    parser.add_argument(
        "--current-date",
        help="Current date to add. If omitted, defaults to today's date.",
    )
    parser.add_argument(
        "--limit-sheets",
        type=int,
        help="Stop after this many sheets are successfully changed.",
    )
    parser.add_argument(
        "--fast-xml",
        action="store_true",
        help="Use direct xlsx XML editing for large files. Supports colored sheet batches.",
    )
    parser.add_argument(
        "--run-until-done",
        action="store_true",
        help="Repeat fast XML batches until no more sheets are changed.",
    )
    args = parser.parse_args(argv)
    args.interactive = args.interactive or len(argv if argv is not None else sys.argv[1:]) == 0
    return args


def parse_bool(value: str, option_name: str) -> bool:
    normalized = value.strip().lower()
    if normalized in {"1", "yes", "y", "true", "on"}:
        return True
    if normalized in {"0", "no", "n", "false", "off"}:
        return False
    raise ValueError(f"config.ini 中 {option_name} 必须是 true 或 false")


def get_config_path() -> Path:
    if is_frozen_app():
        candidates = (
            get_executable_directory() / CONFIG_FILE_NAME,
            Path.cwd() / CONFIG_FILE_NAME,
            PROJECT_ROOT / CONFIG_FILE_NAME,
        )
    else:
        candidates = (
            Path.cwd() / CONFIG_FILE_NAME,
            PROJECT_ROOT / CONFIG_FILE_NAME,
        )
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def load_config() -> configparser.SectionProxy:
    config = configparser.ConfigParser()
    config_path = get_config_path()
    if config_path.exists():
        config.read(config_path, encoding="utf-8")
    if "fill" not in config:
        config["fill"] = {}
    return config["fill"]


def apply_config_defaults(args: argparse.Namespace, config: configparser.SectionProxy) -> argparse.Namespace:
    workbook_value = config.get("workbook", "").strip()
    if args.workbook is None and workbook_value:
        workbook_path = Path(workbook_value).expanduser()
        if not workbook_path.is_absolute():
            workbook_path = find_existing_workspace_directory() / workbook_path
        args.workbook = workbook_path

    if args.current_date is None:
        target_date = config.get("target_date", "").strip()
        if target_date:
            args.current_date = target_date

    if args.limit_sheets is None:
        limit_sheets = config.get("limit_sheets", "").strip()
        if limit_sheets:
            args.limit_sheets = int(limit_sheets)

    if not args.colored_sheets:
        args.colored_sheets = parse_bool(config.get("colored_sheets", "true"), "colored_sheets")

    if not args.fast_xml:
        args.fast_xml = parse_bool(config.get("fast_xml", "true"), "fast_xml")

    if not args.run_until_done:
        args.run_until_done = parse_bool(config.get("run_until_done", "true"), "run_until_done")

    return args


def parse_date(value: str) -> date:
    raw_value = value.strip()
    current_year = date.today().year

    if re.fullmatch(r"\d{4}", raw_value):
        month = int(raw_value[:2])
        day = int(raw_value[2:])
        return date(current_year, month, day)

    normalized = raw_value.replace("/", "-").replace(".", "-")
    if re.fullmatch(r"\d{1,2}-\d{1,2}", normalized):
        month_text, day_text = normalized.split("-")
        return date(current_year, int(month_text), int(day_text))

    for date_format in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(raw_value, date_format).date()
        except ValueError:
            continue

    raise ValueError("日期格式不对，请输入 2026-06-10、0610、06-10 或 05/12")


def resolve_current_date(args: argparse.Namespace) -> date:
    if args.current_date:
        return parse_date(args.current_date)
    return date.today()


def choose_date_mode() -> str:
    today_text = date.today().strftime("%Y-%m-%d")
    while True:
        raw_value = input(f"目标日期：1=今天({today_text})，2=手动输入，直接回车默认 1：").strip()
        if raw_value in {"", "1"}:
            return today_text
        if raw_value == "2":
            return choose_current_date()
        print("请输入 1 或 2。")


def format_file_size(path: Path) -> str:
    size = path.stat().st_size
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024 or unit == "GB":
            return f"{size:.1f} {unit}" if unit != "B" else f"{size} {unit}"
        size /= 1024
    return f"{size:.1f} GB"


def list_current_xlsx_files(directory: Path) -> list[Path]:
    return sorted(
        (
            path
            for path in directory.glob("*.xlsx")
            if path.is_file() and not path.name.startswith(("~$", ".~", "._"))
        ),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )


def get_executable_directory() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return PROJECT_ROOT


def iter_workspace_directories() -> list[Path]:
    directories: list[Path] = []
    if getattr(sys, "frozen", False):
        base_directories = (get_executable_directory(), Path.cwd(), PROJECT_ROOT)
    else:
        base_directories = (Path.cwd(), PROJECT_ROOT)

    for base_directory in base_directories:
        workspace_directory = base_directory / WORKSPACE_DIR_NAME
        if workspace_directory not in directories:
            directories.append(workspace_directory)
    return directories


def find_existing_workspace_directory() -> Path:
    checked: list[Path] = []
    for directory in iter_workspace_directories():
        checked.append(directory)
        if directory.is_dir():
            return directory
    checked_text = "\n  ".join(str(path) for path in checked)
    raise FileNotFoundError(f"没有找到 workspace 文件夹。已检查目录：\n  {checked_text}")


def find_workbook_directory() -> tuple[Path, list[Path]]:
    checked: list[Path] = []
    for directory in iter_workspace_directories():
        checked.append(directory)
        if not directory.is_dir():
            continue
        files = list_current_xlsx_files(directory)
        if files:
            return directory, files
    checked_text = "\n  ".join(str(path) for path in checked)
    raise FileNotFoundError(
        f"没有在 workspace 文件夹里找到 xlsx 文件。请把 Excel 放入 workspace 后再运行。\n"
        f"已检查目录：\n  {checked_text}"
    )


def choose_workbook_from_current_directory() -> Path:
    if not sys.stdin.isatty():
        raise RuntimeError("交互模式需要可输入的终端；或请使用 --workbook 等参数直接运行。")

    workbook_directory, files = find_workbook_directory()

    print("Excel 数据填报工具")
    print("----------------------------------------")
    print(f"Excel 搜索目录：{workbook_directory}")
    print("请选择要处理的 workbook：")
    for index, path in enumerate(files, start=1):
        print(f"  {index}. {path.name} ({format_file_size(path)})")

    default_index = 1
    while True:
        raw_value = input(f"输入编号或文件路径，直接回车默认 {default_index}：").strip()
        if not raw_value:
            return files[default_index - 1]

        unquoted = raw_value.strip("'\"")
        if unquoted.isdigit():
            selected_index = int(unquoted)
            if 1 <= selected_index <= len(files):
                return files[selected_index - 1]
            print(f"编号超出范围，请输入 1 到 {len(files)}。")
            continue

        selected_path = Path(unquoted).expanduser()
        if not selected_path.is_absolute():
            selected_path = workbook_directory / selected_path
        if selected_path.is_file() and selected_path.suffix.lower() == ".xlsx":
            return selected_path
        print("没有找到这个 xlsx 文件，请重新输入。")


def choose_current_date() -> str:
    default_date = date.today().strftime("%Y-%m-%d")
    while True:
        raw_value = input(
            f"请输入目标日期，直接回车默认今天 {default_date}；示例：2026-06-10、0610、06-10、05/12："
        ).strip()
        selected_value = raw_value or default_date
        try:
            return parse_date(selected_value).strftime("%Y-%m-%d")
        except ValueError as error:
            print(error)
            continue


def apply_interactive_fill_defaults(
    args: argparse.Namespace,
    config: configparser.SectionProxy,
) -> argparse.Namespace:
    select_workbook = parse_bool(config.get("select_workbook", "true"), "select_workbook")
    if select_workbook or args.workbook is None:
        args.workbook = choose_workbook_from_current_directory()
    args.current_date = choose_date_mode()
    args.freeze_next_day_row = True
    args.limit_sheets = args.limit_sheets or 20

    print()
    print(f"配置文件：{get_config_path()}")
    print("将使用以下配置处理：")
    print(f"  workbook: {args.workbook}")
    print(f"  current-date: {args.current_date}")
    print(f"  colored-sheets: {'yes' if args.colored_sheets else 'no'}")
    print(f"  fast-xml: {'yes' if args.fast_xml else 'no'}")
    print(f"  limit-sheets: {args.limit_sheets}")
    print(f"  run-until-done: {'yes' if args.run_until_done else 'no'}")
    print()
    return args


def is_frozen_app() -> bool:
    return getattr(sys, "frozen", False)


def should_pause_before_exit() -> bool:
    return is_frozen_app() and sys.stdin.isatty()


def pause_before_exit() -> None:
    if not should_pause_before_exit():
        return
    try:
        input("\n按回车退出...")
    except EOFError:
        pass


def write_error_log(error: BaseException) -> Path:
    log_path = get_executable_directory() / "autoexcel-fill-error.log"
    log_text = "".join(traceback.format_exception(type(error), error, error.__traceback__))
    log_path.write_text(log_text, encoding="utf-8")
    return log_path


def create_process_log_path() -> Path:
    log_dir = get_executable_directory() / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return log_dir / f"autoexcel-fill-{timestamp}.log"


def append_log(log_path: Path, message: str) -> None:
    with log_path.open("a", encoding="utf-8") as log_file:
        log_file.write(message.rstrip() + "\n")


def print_fill_summary(summary: FillSummary) -> None:
    print()
    print("处理完成")
    print("----------------------------------------")
    print(f"Workbook: {summary.workbook}")
    print(f"目标日期: {summary.current_date:%Y-%m-%d}")
    print(f"处理批次: {summary.batch_count}")
    print(f"成功处理 sheet 数: {summary.changed_count}")
    print(f"跳过 sheet 数: {summary.skipped_count}")
    if summary.changed:
        print("已处理 sheet:")
        for sheet_name, row in summary.changed:
            print(f"  - {sheet_name}，行 {row}")
    else:
        print("已处理 sheet: 无")
    print(f"详细日志: {summary.log_path}")


def main() -> None:
    args = parse_args()
    config = load_config()
    args = apply_config_defaults(args, config)

    if args.interactive:
        args = apply_interactive_fill_defaults(args, config)

    if args.workbook is None:
        args.workbook = DATA_FILE

    if args.list_sheets:
        for sheet_name in list_sheet_names(args.workbook):
            print(sheet_name)
        return

    if args.freeze_next_day_row:
        current_date = resolve_current_date(args)
        log_path = create_process_log_path()
        summary = FillSummary(workbook=args.workbook, current_date=current_date, log_path=log_path)
        append_log(log_path, "Excel 数据填报详细日志")
        append_log(log_path, f"开始时间: {datetime.now():%Y-%m-%d %H:%M:%S}")
        append_log(log_path, f"Workbook: {args.workbook}")
        append_log(log_path, f"目标日期: {current_date:%Y-%m-%d}")
        append_log(log_path, f"配置文件: {get_config_path()}")
        append_log(log_path, f"colored_sheets: {args.colored_sheets}")
        append_log(log_path, f"fast_xml: {args.fast_xml}")
        append_log(log_path, f"limit_sheets: {args.limit_sheets or 20}")
        append_log(log_path, f"run_until_done: {args.run_until_done}")
        append_log(log_path, "")

        if args.fast_xml:
            if not args.colored_sheets:
                raise ValueError("--fast-xml currently requires --colored-sheets")
            batch_number = 0
            while True:
                batch_number += 1
                summary.batch_count = batch_number
                print(f"Batch {batch_number}: 正在读取并处理 workbook，请不要打开 Excel 文件...")
                sys.stdout.flush()
                append_log(log_path, f"Batch {batch_number}: 开始处理")
                result = add_current_date_to_colored_sheets_fast(
                    xlsx_path=args.workbook,
                    current_date=current_date,
                    limit_sheets=args.limit_sheets or 20,
                    progress=lambda message: print(message, flush=True),
                )
                changed_count = len(result.changed)
                summary.changed.extend(result.changed)
                summary.skipped_count += len(result.skipped)
                print(f"Batch {batch_number}: changed {changed_count}, skipped {len(result.skipped)}.")
                append_log(log_path, f"Batch {batch_number}: changed {changed_count}, skipped {len(result.skipped)}")
                for sheet_name, row in result.changed:
                    append_log(log_path, f"  changed: {sheet_name} row {row}")
                for sheet_name, reason in result.skipped:
                    append_log(log_path, f"  skipped: {sheet_name} ({reason})")
                append_log(log_path, "")

                if not args.run_until_done or changed_count == 0:
                    if changed_count == 0:
                        append_log(log_path, "没有更多可处理的彩色标签 sheet。")
                    append_log(log_path, f"结束时间: {datetime.now():%Y-%m-%d %H:%M:%S}")
                    print_fill_summary(summary)
                    break
            return

        workbook = load_workbook(args.workbook, data_only=False)
        if args.sheet not in workbook.sheetnames and args.copy_sheet_from:
            if args.copy_sheet_from not in workbook.sheetnames:
                available = ", ".join(workbook.sheetnames)
                raise ValueError(
                    f"Source sheet '{args.copy_sheet_from}' not found. Available sheets: {available}"
                )
            copied_sheet = workbook.copy_worksheet(workbook[args.copy_sheet_from])
            copied_sheet.title = args.sheet
            workbook.save(args.workbook)

            workbook = load_workbook(args.workbook, data_only=False)

        values_workbook = load_workbook(args.workbook, data_only=True)
        if args.colored_sheets:
            changed, skipped = freeze_colored_sheets_next_day(
                workbook=workbook,
                values_workbook=values_workbook,
                current_date=current_date,
                limit_sheets=args.limit_sheets,
            )
            workbook.save(args.workbook)
            summary.batch_count = 1
            summary.changed.extend(changed)
            summary.skipped_count += len(skipped)
            append_log(log_path, f"Changed {len(changed)} colored sheets.")
            for sheet_name, row in changed:
                append_log(log_path, f"  changed: {sheet_name} row {row}")
            append_log(log_path, f"Skipped {len(skipped)} colored sheets.")
            for sheet_name, reason in skipped:
                append_log(log_path, f"  skipped: {sheet_name} ({reason})")
            append_log(log_path, f"结束时间: {datetime.now():%Y-%m-%d %H:%M:%S}")
            print_fill_summary(summary)
            return

        inserted_row = freeze_next_day_row(
            workbook=workbook,
            values_workbook=values_workbook,
            sheet_name=args.sheet,
            current_date=current_date,
        )
        workbook.save(args.workbook)
        summary.batch_count = 1
        summary.changed.append((args.sheet, inserted_row))
        append_log(log_path, f"changed: {args.sheet} row {inserted_row}")
        append_log(log_path, f"结束时间: {datetime.now():%Y-%m-%d %H:%M:%S}")
        print_fill_summary(summary)
        return

    preview = preview_sheet(args.workbook, args.sheet, max_rows=args.rows)

    print(f"Workbook: {args.workbook}")
    print(f"Sheet: {preview.sheet_name}")
    print(f"Size: {preview.max_row} rows x {preview.max_column} columns")
    print("Preview:")
    for row in preview.rows:
        print(row)


if __name__ == "__main__":
    exit_code = 0
    try:
        main()
    except KeyboardInterrupt:
        print("\n已取消。")
        exit_code = 130
    except (FileNotFoundError, RuntimeError, ValueError) as error:
        print(f"错误：{error}")
        exit_code = 1
    except Exception as error:
        try:
            log_path = write_error_log(error)
            print(f"程序执行失败：{error}")
            print(f"详细错误已写入：{log_path}")
        except Exception:
            print("程序执行失败，并且写入错误日志时也失败：")
            traceback.print_exc()
        exit_code = 1
    finally:
        pause_before_exit()
    sys.exit(exit_code)
