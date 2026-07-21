from __future__ import annotations

import csv
from dataclasses import replace
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re
from typing import Iterable, Iterator

from openpyxl import load_workbook

from autoexcel import diff_orders


PAYOUT_MODE = "payout"
PAYOUT_FINERBIT_MODE = "payout_finerbit"
UPSTREAM_REQUIRED_HEADERS = {
    "TRANS_ID",
    "TRX_STATUS",
    "TRX_AMT",
    "FEE",
    "FED",
}
BACKEND_REQUIRED_HEADERS = {
    "transactionId",
    "交易状态",
    "付款金额(PKR)",
    "手续费(PKR)",
    "支付方式名称",
}
FINERBIT_COLLECTION_REQUIRED_HEADERS = {
    "ChannelName",
    "PurchaseAmount",
    "TransactionStatus",
    "Created Date",
}
FINERBIT_DISBURSEMENT_REQUIRED_HEADERS = {
    "Created Date Time",
    "Reference Id",
    "Received Amount",
    "Status",
}
FINERBIT_BACKEND_REQUIRED_HEADERS = {
    "平台订单号",
    "付款金额(PKR)",
    "手续费(PKR)",
    "渠道成本(PKR)",
    "交易状态",
}
SUPPORTED_UPSTREAM_SUFFIXES = {".csv", ".xlsx"}
SUPPORTED_BACKEND_SUFFIXES = {".xlsx"}


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).casefold()


def header_map(row: Iterable[object]) -> dict[str, int]:
    return {
        normalize_header(value): index
        for index, value in enumerate(row)
        if value not in (None, "")
    }


def contains_headers(row: Iterable[object], required: set[str]) -> bool:
    available = header_map(row)
    return all(normalize_header(header) in available for header in required)


def _csv_rows(path: Path) -> Iterator[tuple[int, list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        for row_number, row in enumerate(csv.reader(stream), start=1):
            yield row_number, row


def _xlsx_rows(
    path: Path, max_rows_per_sheet: int | None = None
) -> Iterator[tuple[str, int, tuple[object, ...]]]:
    with path.open("rb") as stream:
        workbook = load_workbook(stream, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                worksheet.reset_dimensions()
                for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                    if max_rows_per_sheet is not None and row_number > max_rows_per_sheet:
                        break
                    yield worksheet.title, row_number, row
        finally:
            workbook.close()


def file_kind(path: Path) -> str | None:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        for row_number, row in _csv_rows(path):
            if contains_headers(row, UPSTREAM_REQUIRED_HEADERS):
                return "upstream"
            if row_number >= 20:
                break
        return None
    if suffix != ".xlsx":
        return None

    for _sheet_name, _row_number, row in _xlsx_rows(path, max_rows_per_sheet=20):
        if contains_headers(row, UPSTREAM_REQUIRED_HEADERS):
            return "upstream"
        if contains_headers(row, BACKEND_REQUIRED_HEADERS):
            return "backend"
    return None


def validate_file(path: Path, label: str, allowed_suffixes: set[str]) -> None:
    if not path.is_file():
        raise FileNotFoundError(f"{label}不存在：{path}")
    if path.suffix.lower() not in allowed_suffixes:
        suffixes = " / ".join(sorted(allowed_suffixes))
        raise ValueError(f"{label}必须是 {suffixes} 文件")


def _xlsx_contains_headers(path: Path, required_headers: set[str]) -> bool:
    return any(
        contains_headers(row, required_headers)
        for _sheet_name, _row_number, row in _xlsx_rows(
            path, max_rows_per_sheet=20
        )
    )


def make_job(
    upstream_path: Path,
    backend_path: Path,
    algorithm: str = "zee",
    collection_path: Path | None = None,
) -> diff_orders.DiffJob:
    normalized_algorithm = algorithm.strip().lower()
    if normalized_algorithm == "finerbit":
        validate_file(upstream_path, "finerBit代付上游账单", SUPPORTED_UPSTREAM_SUFFIXES)
        validate_file(backend_path, "我方付款订单", SUPPORTED_BACKEND_SUFFIXES)
        if collection_path is None:
            raise ValueError("代付 finerBit 必须选择上游代收账单")
        validate_file(collection_path, "finerBit上游代收账单", SUPPORTED_UPSTREAM_SUFFIXES)
        if not _xlsx_contains_headers(
            upstream_path, FINERBIT_DISBURSEMENT_REQUIRED_HEADERS
        ):
            raise ValueError(f"{upstream_path.name} 未找到 finerBit 代付上游必需表头")
        if not _xlsx_contains_headers(
            backend_path, FINERBIT_BACKEND_REQUIRED_HEADERS
        ):
            raise ValueError(f"{backend_path.name} 未找到我方付款订单必需表头")
        if not _xlsx_contains_headers(
            collection_path, FINERBIT_COLLECTION_REQUIRED_HEADERS
        ):
            raise ValueError(f"{collection_path.name} 未找到 finerBit 上游代收必需表头")
        return diff_orders.DiffJob(
            upstream_path=upstream_path,
            backend_path=backend_path,
            platform_mode="finerbit",
            collection_path=collection_path,
        )

    if normalized_algorithm != "zee":
        raise ValueError(f"不支持的代付算法：{algorithm}")
    validate_file(upstream_path, "代付上游账单", SUPPORTED_UPSTREAM_SUFFIXES)
    validate_file(backend_path, "我方付款订单", SUPPORTED_BACKEND_SUFFIXES)
    if file_kind(upstream_path) != "upstream":
        raise ValueError(f"{upstream_path.name} 未找到代付上游必需表头")
    if file_kind(backend_path) != "backend":
        raise ValueError(f"{backend_path.name} 未找到我方付款订单必需表头")
    return diff_orders.DiffJob(
        upstream_path=upstream_path,
        backend_path=backend_path,
        platform_mode="zee",
    )


def find_jobs_in_directory(directory: Path) -> list[diff_orders.DiffJob]:
    if not directory.is_dir():
        raise FileNotFoundError(f"订单目录不存在：{directory}")

    classified: dict[Path, dict[str, list[Path]]] = {}
    all_upstream: list[Path] = []
    all_backend: list[Path] = []
    for path in sorted(directory.rglob("*")):
        if not path.is_file() or path.name.startswith(("~$", ".~", "._")):
            continue
        if path.suffix.lower() not in SUPPORTED_UPSTREAM_SUFFIXES:
            continue
        kind = file_kind(path)
        if kind is None:
            continue
        classified.setdefault(path.parent, {"upstream": [], "backend": []})[kind].append(path)
        (all_upstream if kind == "upstream" else all_backend).append(path)

    jobs: list[diff_orders.DiffJob] = []
    for group_directory, files in classified.items():
        upstream_paths = files["upstream"]
        backend_paths = files["backend"]
        if not upstream_paths or not backend_paths:
            continue
        if len(upstream_paths) != 1 or len(backend_paths) != 1:
            raise ValueError(
                f"{group_directory} 识别到 {len(upstream_paths)} 个上游账单和 "
                f"{len(backend_paths)} 个我方付款订单，无法确定配对。"
            )
        jobs.append(make_job(upstream_paths[0], backend_paths[0]))

    if jobs:
        return jobs
    if len(all_upstream) == 1 and len(all_backend) == 1:
        return [make_job(all_upstream[0], all_backend[0])]
    raise FileNotFoundError(
        "目录中未找到可配对的代付上游账单和我方付款订单。"
    )


def _read_csv_table(path: Path, required_headers: set[str]) -> Iterator[tuple[int, list[str], dict[str, int]]]:
    mapping: dict[str, int] | None = None
    for row_number, row in _csv_rows(path):
        if mapping is None:
            if contains_headers(row, required_headers):
                mapping = header_map(row)
            continue
        yield row_number, row, mapping
    if mapping is None:
        raise ValueError(f"{path.name} 未找到必需表头")


def _read_xlsx_table(
    path: Path, required_headers: set[str]
) -> Iterator[tuple[str, int, tuple[object, ...], dict[str, int]]]:
    found_header = False
    with path.open("rb") as stream:
        workbook = load_workbook(stream, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                worksheet.reset_dimensions()
                mapping: dict[str, int] | None = None
                for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                    if mapping is None:
                        if contains_headers(row, required_headers):
                            mapping = header_map(row)
                            found_header = True
                        continue
                    yield worksheet.title, row_number, row, mapping
        finally:
            workbook.close()
    if not found_header:
        raise ValueError(f"{path.name} 未找到必需表头")


def _row_value(row: Iterable[object], mapping: dict[str, int], header: str) -> object:
    values = row if isinstance(row, (list, tuple)) else tuple(row)
    index = mapping[normalize_header(header)]
    return values[index] if index < len(values) else None


def read_upstream_entries(path: Path) -> list[diff_orders.OrderEntry]:
    entries: list[diff_orders.OrderEntry] = []
    if path.suffix.lower() == ".csv":
        rows = (
            ("CSV", row_number, row, mapping)
            for row_number, row, mapping in _read_csv_table(path, UPSTREAM_REQUIRED_HEADERS)
        )
    else:
        rows = _read_xlsx_table(path, UPSTREAM_REQUIRED_HEADERS)

    for sheet_name, row_number, row, mapping in rows:
        status = str(_row_value(row, mapping, "TRX_STATUS") or "").strip().upper()
        if status != "COMPLETED":
            continue
        order_id = diff_orders.normalize_order_id(_row_value(row, mapping, "TRANS_ID"))
        if order_id is None:
            continue
        fee = diff_orders.decimal_value(_row_value(row, mapping, "FEE"))
        fed = diff_orders.decimal_value(_row_value(row, mapping, "FED"))
        entries.append(
            diff_orders.OrderEntry(
                order_id=order_id,
                row_number=row_number,
                amount=diff_orders.decimal_value(_row_value(row, mapping, "TRX_AMT")),
                fee=fee + fed,
                sheet_name=sheet_name,
                extra={"status": status},
            )
        )
    return entries


def read_backend_entries(path: Path) -> list[diff_orders.OrderEntry]:
    entries: list[diff_orders.OrderEntry] = []
    for sheet_name, row_number, row, mapping in _read_xlsx_table(
        path, BACKEND_REQUIRED_HEADERS
    ):
        status = str(_row_value(row, mapping, "交易状态") or "").strip()
        if status != "上游已打款":
            continue
        order_id = diff_orders.normalize_order_id(_row_value(row, mapping, "transactionId"))
        if order_id is None:
            continue
        entries.append(
            diff_orders.OrderEntry(
                order_id=order_id,
                row_number=row_number,
                amount=diff_orders.decimal_value(_row_value(row, mapping, "付款金额(PKR)")),
                fee=diff_orders.decimal_value(_row_value(row, mapping, "手续费(PKR)")),
                payment_method=str(
                    _row_value(row, mapping, "支付方式名称") or ""
                ).strip(),
                sheet_name=sheet_name,
                extra={"status": status},
            )
        )
    return entries


def payout_channel_cost(entries: list[diff_orders.OrderEntry]) -> Decimal:
    total = Decimal("0")
    for entry in entries:
        method = entry.payment_method.strip().lower()
        rate = Decimal("0.0034") if method in {"jazzcash", "jc"} else Decimal("0.007")
        total += (entry.amount * rate).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    return total


def load_finerbit_fee_rates() -> tuple[Decimal, Decimal]:
    return diff_orders.load_payout_finerbit_fee_rates()


def normalize_report_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or "").strip()
    for pattern, date_format in (
        (r"\d{2}-\d{2}-\d{4}", "%d-%m-%Y"),
        (r"\d{4}-\d{2}-\d{2}", "%Y-%m-%d"),
        (r"\d{2}/\d{2}/\d{4}", "%d/%m/%Y"),
        (r"\d{4}/\d{2}/\d{2}", "%Y/%m/%d"),
    ):
        match = re.search(pattern, text)
        if match:
            return datetime.strptime(match.group(), date_format).date().isoformat()
    raise ValueError(f"无法识别账单日期：{value}")


def read_finerbit_collection_amounts(
    path: Path,
) -> tuple[Decimal, Decimal, str]:
    easypaisa_amount = Decimal("0")
    jazzcash_amount = Decimal("0")
    report_dates: set[str] = set()
    for _sheet_name, _row_number, row, mapping in _read_xlsx_table(
        path, FINERBIT_COLLECTION_REQUIRED_HEADERS
    ):
        status = str(_row_value(row, mapping, "TransactionStatus") or "").strip().lower()
        if status != "success":
            continue
        report_dates.add(normalize_report_date(_row_value(row, mapping, "Created Date")))
        channel = str(_row_value(row, mapping, "ChannelName") or "").strip().lower()
        amount = diff_orders.decimal_value(_row_value(row, mapping, "PurchaseAmount"))
        if "easypaisa" in channel or "easy paisa" in channel:
            easypaisa_amount += amount
        elif "jazzcash" in channel or "jazz cash" in channel:
            jazzcash_amount += amount
    if len(report_dates) != 1:
        dates = "、".join(sorted(report_dates)) or "无"
        raise ValueError(f"上游代收账单应只有一个成功交易日期，当前为：{dates}")
    return easypaisa_amount, jazzcash_amount, next(iter(report_dates))


def read_finerbit_disbursement_entries(
    path: Path, report_date: str
) -> list[diff_orders.OrderEntry]:
    entries: list[diff_orders.OrderEntry] = []
    for sheet_name, row_number, row, mapping in _read_xlsx_table(
        path, FINERBIT_DISBURSEMENT_REQUIRED_HEADERS
    ):
        status = str(_row_value(row, mapping, "Status") or "").strip().lower()
        if status != "success":
            continue
        created_at = _row_value(row, mapping, "Created Date Time")
        if normalize_report_date(created_at) != report_date:
            continue
        order_id = diff_orders.normalize_order_id(_row_value(row, mapping, "Reference Id"))
        if order_id is None:
            continue
        entries.append(
            diff_orders.OrderEntry(
                order_id=order_id,
                row_number=row_number,
                amount=diff_orders.decimal_value(
                    _row_value(row, mapping, "Received Amount")
                ),
                sheet_name=sheet_name,
                extra={"status": status},
            )
        )
    return entries


def read_finerbit_backend_entries(path: Path) -> list[diff_orders.OrderEntry]:
    entries: list[diff_orders.OrderEntry] = []
    for sheet_name, row_number, row, mapping in _read_xlsx_table(
        path, FINERBIT_BACKEND_REQUIRED_HEADERS
    ):
        status = str(_row_value(row, mapping, "交易状态") or "").strip()
        if status != "上游已打款":
            continue
        order_id = diff_orders.normalize_order_id(_row_value(row, mapping, "平台订单号"))
        if order_id is None:
            continue
        entries.append(
            diff_orders.OrderEntry(
                order_id=order_id,
                row_number=row_number,
                amount=diff_orders.decimal_value(
                    _row_value(row, mapping, "付款金额(PKR)")
                ),
                fee=diff_orders.decimal_value(
                    _row_value(row, mapping, "手续费(PKR)")
                ),
                sheet_name=sheet_name,
                extra={
                    "status": status,
                    "channel_cost": _row_value(row, mapping, "渠道成本(PKR)"),
                },
            )
        )
    return entries


def finerbit_fee(
    easypaisa_amount: Decimal,
    jazzcash_amount: Decimal,
    easypaisa_rate: Decimal,
    jazzcash_rate: Decimal,
) -> Decimal:
    return easypaisa_amount * easypaisa_rate + jazzcash_amount * jazzcash_rate


def read_job_diff_result(job: diff_orders.DiffJob) -> diff_orders.JobDiffResult:
    if job.platform_mode == "finerbit":
        if job.collection_path is None:
            raise ValueError("代付 finerBit 必须选择上游代收账单")
        backend_entries = read_finerbit_backend_entries(job.backend_path)
        easypaisa_amount, jazzcash_amount, report_date = read_finerbit_collection_amounts(
            job.collection_path
        )
        upstream_entries = read_finerbit_disbursement_entries(
            job.upstream_path, report_date
        )
        easypaisa_rate, jazzcash_rate = load_finerbit_fee_rates()
        result = diff_orders.diff_orders(upstream_entries, backend_entries)
        result = replace(
            result,
            mismatched=[],
            mismatch_amount=Decimal("0"),
            a_fee=finerbit_fee(
                easypaisa_amount,
                jazzcash_amount,
                easypaisa_rate,
                jazzcash_rate,
            ),
            b_fee=sum((entry.fee for entry in backend_entries), Decimal("0")),
            channel_cost=sum(
                (
                    diff_orders.decimal_value(entry.extra.get("channel_cost"))
                    for entry in backend_entries
                ),
                Decimal("0"),
            ),
            special_mode=PAYOUT_FINERBIT_MODE,
        )
        return diff_orders.JobDiffResult(job=job, result=result)

    upstream_entries = read_upstream_entries(job.upstream_path)
    backend_entries = read_backend_entries(job.backend_path)
    result = diff_orders.diff_orders(upstream_entries, backend_entries)
    result = replace(
        result,
        a_fee=sum((entry.fee for entry in upstream_entries), Decimal("0")),
        b_fee=sum((entry.fee for entry in backend_entries), Decimal("0")),
        channel_cost=payout_channel_cost(backend_entries),
        special_mode=PAYOUT_MODE,
    )
    return diff_orders.JobDiffResult(job=job, result=result)


def write_html_results(
    results: list[diff_orders.JobDiffResult], result_dir: Path
) -> Path:
    result_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    group_part = diff_orders.result_filename_group_part(results)
    batch_part = "batch_" if len(results) > 1 else ""
    output_path = result_dir / f"payout_order_diff_{batch_part}{group_part}_{timestamp}.html"
    output_path.write_text(diff_orders.render_stats_html(results), encoding="utf-8")
    return output_path
