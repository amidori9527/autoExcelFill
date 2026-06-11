from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class SheetPreview:
    sheet_name: str
    max_row: int
    max_column: int
    rows: list[list[Any]]


def load_sheet(workbook_path: Path, sheet_name: str) -> Worksheet:
    workbook = load_workbook(workbook_path, data_only=False)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available}")
    return workbook[sheet_name]


def list_sheet_names(workbook_path: Path) -> list[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    return list(workbook.sheetnames)


def preview_sheet(workbook_path: Path, sheet_name: str, max_rows: int = 10) -> SheetPreview:
    sheet = load_sheet(workbook_path, sheet_name)
    rows = [
        list(row)
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(sheet.max_row, max_rows),
            values_only=True,
        )
    ]
    return SheetPreview(
        sheet_name=sheet.title,
        max_row=sheet.max_row,
        max_column=sheet.max_column,
        rows=rows,
    )
